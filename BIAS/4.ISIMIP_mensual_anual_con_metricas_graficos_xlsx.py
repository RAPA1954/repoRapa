# =============================================================================
#  ISIMIP BIAS CORRECTION — Análisis MENSUAL y ANUAL
#  Métricas de ajuste + Gráficos PNG (FDC, Histograma, QQ-plot) + Excel
#  ✨ VERSIÓN ACTUALIZADA CON CRITERIOS CORRECTOS PARA QUANTILE MAPPING
# =============================================================================

import os
import warnings
import numpy as np
import pandas as pd
from scipy.stats import gamma as gamma_dist, pearsonr
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')

plt.rcParams.update({
    'font.family'       : 'DejaVu Sans',
    'axes.unicode_minus': False,
    'figure.dpi'        : 150,
})

# =============================================================================
#  CONFIGURACIÓN USUARIO  ← EDITAR AQUÍ
# =============================================================================
OBS_XLSX   = r'C:/1.PYTHON/Descarga_Python/Pd_Historica_SMN.xlsx'
GCM_XLSX   = r'C:/1.PYTHON/Descarga_Python/Tabla_CESM2_historico_futuro.xlsx'
COL_OBS    = 'prec'
COL_GCM    = 'pr_mm'
MODELO     = 'CESM2'
CAL_INI    = '1961-01-01'
CAL_FIN    = '2024-12-31'
FUT_INI    = '2030-01-01'
FUT_FIN    = '2100-12-31'
OUT_DIR    = rf'C:/1.PYTHON/Descarga_Python/{MODELO}_salidas_ISIMIP'
N_QUANTILES = 100
UMBRAL_MM   = 0.1
# =============================================================================

os.makedirs(OUT_DIR, exist_ok=True)
PNG_DIR = os.path.join(OUT_DIR, 'graficos')
os.makedirs(PNG_DIR, exist_ok=True)

SALIDA_XLSX = os.path.join(OUT_DIR, f'{MODELO}_ISIMIP_mensual_anual.xlsx')

SEP = '=' * 78
SEP2 = '-' * 78

# ─────────────────────────────────────────────────────────────────────────────
#  FUNCIONES AUXILIARES
# ─────────────────────────────────────────────────────────────────────────────

def titulo_consola(texto, nivel=1):
    if nivel == 1:
        print(f'\n{SEP}')
        print(f'  {texto}')
        print(SEP)
    else:
        print(f'\n{SEP2}')
        print(f'  {texto}')
        print(SEP2)

def isimip_bias_correction(obs_cal, gcm_cal, gcm_fut,
                            distribucion='empirica', n_quantiles=100, umbral=0.1):
    """
    ISIMIP Quantile Mapping con preservación de tendencia futura.
    """
    obs_cal = obs_cal.dropna()
    gcm_cal = gcm_cal.dropna()
    gcm_fut_clean = gcm_fut.dropna()

    if distribucion == 'empirica':
        quantiles   = np.linspace(0, 1, n_quantiles + 1)
        q_obs       = np.quantile(obs_cal,  quantiles)
        q_gcm_cal   = np.quantile(gcm_cal,  quantiles)
        q_gcm_fut   = np.quantile(gcm_fut_clean, quantiles)
        transfer    = q_obs - q_gcm_cal
        q_corr      = q_gcm_fut + transfer
        gcm_corr    = np.interp(gcm_fut_clean.values, q_gcm_fut, q_corr,
                                left=q_corr[0], right=q_corr[-1])

    elif distribucion == 'gamma':
        obs_wet     = obs_cal[obs_cal > umbral]
        gcm_cal_wet = gcm_cal[gcm_cal > umbral]
        gcm_fut_wet = gcm_fut_clean[gcm_fut_clean > umbral]

        if len(obs_wet) < 10 or len(gcm_cal_wet) < 10 or len(gcm_fut_wet) < 10:
            return isimip_bias_correction(obs_cal, gcm_cal, gcm_fut,
                                          distribucion='empirica',
                                          n_quantiles=n_quantiles)

        a_obs,  _, sc_obs  = gamma_dist.fit(obs_wet,     floc=0)
        a_cal,  _, sc_cal  = gamma_dist.fit(gcm_cal_wet, floc=0)
        a_fut,  _, sc_fut  = gamma_dist.fit(gcm_fut_wet, floc=0)

        delta_scale = sc_obs - sc_cal
        sc_corr     = max(sc_fut + delta_scale, 1e-6)

        p           = gamma_dist.cdf(gcm_fut_wet.values, a_fut,  scale=sc_fut)
        corr_wet    = gamma_dist.ppf(p,                  a_obs,  scale=sc_corr)

        gcm_corr_s               = gcm_fut_clean.copy().astype(float)
        gcm_corr_s[gcm_fut_clean >  umbral] = corr_wet
        gcm_corr_s[gcm_fut_clean <= umbral] = 0.0
        gcm_corr = gcm_corr_s.values
    else:
        raise ValueError("distribucion debe ser 'empirica' o 'gamma'")

    return pd.Series(np.clip(gcm_corr, 0, None), index=gcm_fut_clean.index)


def calcular_metricas(obs, sim, nombre=''):
    """Calcula todas las métricas para evaluación de QM."""
    df  = pd.DataFrame({'obs': obs, 'sim': sim}).dropna()
    o   = df['obs'].values
    s   = df['sim'].values
    n   = len(o)

    if n < 3:
        return None

    nse   = 1 - np.sum((o - s)**2) / np.sum((o - np.mean(o))**2)
    r, _  = pearsonr(o, s)
    alpha = np.std(s)  / np.std(o)
    beta  = np.mean(s) / np.mean(o)
    kge   = 1 - np.sqrt((r-1)**2 + (alpha-1)**2 + (beta-1)**2)
    pbias = (np.sum(s - o) / np.sum(o)) * 100  # (SIM - OBS) / OBS
    rmse  = np.sqrt(np.mean((o - s)**2))
    mae   = np.mean(np.abs(o - s))
    r2    = r**2

    # Métricas específicas para QM
    media_obs = np.mean(o)
    media_sim = np.mean(s)
    error_media = abs(media_obs - media_sim) / media_obs * 100 if media_obs != 0 else 0

    std_obs = np.std(o)
    std_sim = np.std(s)
    error_std = abs(std_obs - std_sim) / std_obs * 100 if std_obs != 0 else 0

    return {
        'Nombre'         : nombre,
        'n'              : n,
        'NSE'            : round(nse,   4),
        'KGE'            : round(kge,   4),
        'r (Pearson)'    : round(r,     4),
        'R²'             : round(r2,    4),
        'PBIAS (%)'      : round(pbias, 2),
        'RMSE (mm)'      : round(rmse,  4),
        'MAE (mm)'       : round(mae,   4),
        'Media OBS'      : round(media_obs, 4),
        'Media SIM'      : round(media_sim, 4),
        'Error Media (%)': round(error_media, 2),
        'StdDev OBS'     : round(std_obs,  4),
        'StdDev SIM'     : round(std_sim,  4),
        'Error StdDev (%)': round(error_std, 2),
    }


def imprimir_metricas(m, escala=''):
    if m is None:
        print('  !! Insuficientes datos para calcular métricas.')
        return
    print(f'\n  ── {m["Nombre"]} ({escala}) ──')
    print(f'     n               : {m["n"]}')
    print(f'     NSE             : {m["NSE"]:>8.4f}   (ℹ️ informativo: QM no preserva sincronía)')
    print(f'     KGE             : {m["KGE"]:>8.4f}   (ℹ️ informativo: QM no preserva sincronía)')
    print(f'     r  (Pearson)    : {m["r (Pearson)"]:>8.4f}   (ℹ️ informativo: QM no preserva sincronía)')
    print(f'     R²              : {m["R²"]:>8.4f}')
    print(f'')
    print(f'     ✅ PBIAS (%)       : {m["PBIAS (%)"]:>8.2f}   ← MÉTRICA CLAVE para QM')
    print(f'     ✅ Error Media (%) : {m["Error Media (%)"]:>8.2f}   ← MÉTRICA CLAVE para QM')
    print(f'     ✅ Error StdDev(%) : {m["Error StdDev (%)"]:>8.2f}   ← MÉTRICA CLAVE para QM')
    print(f'')
    print(f'     Media OBS/SIM   : {m["Media OBS"]:>8.4f} / {m["Media SIM"]:>8.4f}')
    print(f'     StdDev OBS/SIM  : {m["StdDev OBS"]:>8.4f} / {m["StdDev SIM"]:>8.4f}')
    print(f'     RMSE (mm)       : {m["RMSE (mm)"]:>8.4f}')
    print(f'     MAE  (mm)       : {m["MAE (mm)"]:>8.4f}')


def clasificar_metrica_qm(m, escala='mensual'):
    """
    Clasifica desempeño según criterios CORRECTOS para Quantile Mapping.
    
    Criterios basados en:
    - PBIAS: sesgo absoluto (0% es perfecto)
    - Error de media: < 2% es excelente
    - Error de varianza: < 5% es excelente
    
    Retorna: (diagnóstico_icon, recomendacion_texto)
    """
    if m is None:
        return '⚠️  SIN DATOS', 'No se pudo evaluar. Insuficientes registros.'

    pbias = abs(m['PBIAS (%)'])
    err_media = m['Error Media (%)']
    err_std = m['Error StdDev (%)']

    # Criterios para QM (Quantile Mapping):
    # Excelente: PBIAS <5% y ambos errores <3%
    # Muy bueno: PBIAS <10% y ambos errores <5%
    # Bueno: PBIAS <15% y ambos errores <10%
    # Aceptable: PBIAS <25% y ambos errores <15%
    # Deficiente: fuera de eso

    if pbias < 5 and err_media < 3 and err_std < 3:
        nivel = 'EXCELENTE'
        icon = '✅✅✅'
        desc = ('El ajuste es excelente. El método ISIMIP es ALTAMENTE RECOMENDADO '
                'para proyecciones de cambio climático.')

    elif pbias < 10 and err_media < 5 and err_std < 5:
        nivel = 'MUY BUENO'
        icon = '✅✅'
        desc = ('El ajuste es muy bueno. El método ISIMIP es RECOMENDADO. '
                'La corrección de distribución es robusta.')

    elif pbias < 15 and err_media < 10 and err_std < 10:
        nivel = 'BUENO'
        icon = '✅'
        desc = ('El ajuste es bueno. El método ISIMIP es ACEPTABLE. '
                'Revisa gráficos FDC y QQ-plot para más detalles.')

    elif pbias < 25 and err_media < 15 and err_std < 15:
        nivel = 'ACEPTABLE'
        icon = '⚠️'
        desc = ('El ajuste es aceptable pero moderado. Considera complementar con '
                'otros métodos o validación estacional/extremos.')

    else:
        nivel = 'DEFICIENTE'
        icon = '❌'
        desc = ('El ajuste no es satisfactorio. Revisa calidad de datos, modelo, '
                'escala o considera otros métodos de corrección.')

    return icon, nivel, desc


# ─────────────────────────────────────────────────────────────────────────────
#  FUNCIONES DE GRÁFICOS (sin errores de fuentes)
# ─────────────────────────────────────────────────────────────────────────────

COLORES = {
    'OBS'  : '#1a6e9e',
    'SIN'  : '#d94f3d',
    'EMP'  : '#2ca02c',
    'GAM'  : '#ff7f0e',
}

def _guardar(fig, ruta):
    """Guarda figura con manejo seguro de tight_layout."""
    try:
        fig.tight_layout()
    except Exception:
        pass
    fig.savefig(ruta, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f'     PNG guardado → {ruta}')
    return ruta


def plot_fdc(series_dict, ruta_png, titulo):
    """Curva de Duración de Caudales (FDC / excedencia)."""
    fig, ax = plt.subplots(figsize=(10, 5))
    estilos = ['-', '--', '-.', ':']
    colores = [COLORES['OBS'], COLORES['SIN'], COLORES['EMP'], COLORES['GAM']]

    for i, (lbl, s) in enumerate(series_dict.items()):
        data = s.dropna().sort_values(ascending=False).values
        n    = len(data)
        prob = np.arange(1, n + 1) / (n + 1) * 100
        ax.plot(prob, data,
                linestyle=estilos[i % 4],
                color=colores[i % 4],
                linewidth=1.8,
                label=lbl)

    ax.set_xlabel('Probabilidad de excedencia (%)', fontsize=11)
    ax.set_ylabel('Precipitación (mm)', fontsize=11)
    ax.set_title(titulo, fontsize=12, fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(True, linestyle='--', alpha=0.5)
    ax.xaxis.set_major_formatter(mticker.FormatStrFormatter('%g%%'))
    return _guardar(fig, ruta_png)


def plot_hist(series_dict, ruta_png, titulo):
    """Histogramas de densidad superpuestos."""
    fig, ax = plt.subplots(figsize=(10, 5))
    colores = [COLORES['OBS'], COLORES['SIN'], COLORES['EMP'], COLORES['GAM']]

    for i, (lbl, s) in enumerate(series_dict.items()):
        data = s.dropna().values
        ax.hist(data, bins=30, density=True, alpha=0.45,
                color=colores[i % 4], label=lbl, edgecolor='white')

    ax.set_xlabel('Precipitación (mm)', fontsize=11)
    ax.set_ylabel('Densidad', fontsize=11)
    ax.set_title(titulo, fontsize=12, fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(True, linestyle='--', alpha=0.5)
    return _guardar(fig, ruta_png)


def plot_qq(obs, sims_dict, ruta_png, titulo):
    """QQ-plot: cuantiles de OBS vs cuantiles de cada SIM."""
    quantiles = np.linspace(0, 1, 101)
    q_obs = np.quantile(obs.dropna(), quantiles)

    colores = [COLORES['SIN'], COLORES['EMP'], COLORES['GAM']]
    fig, ax = plt.subplots(figsize=(7, 7))

    lim_max = q_obs[-1]
    for i, (lbl, s) in enumerate(sims_dict.items()):
        q_sim = np.quantile(s.dropna(), quantiles)
        ax.scatter(q_obs, q_sim, s=18, alpha=0.7,
                   color=colores[i % 3], label=lbl)
        lim_max = max(lim_max, q_sim[-1])

    ax.plot([0, lim_max], [0, lim_max], 'k--', linewidth=1.2, label='1:1 (perfecto)')
    ax.set_xlabel('Cuantiles OBS (mm)', fontsize=11)
    ax.set_ylabel('Cuantiles SIM (mm)', fontsize=11)
    ax.set_title(titulo, fontsize=12, fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(True, linestyle='--', alpha=0.5)
    return _guardar(fig, ruta_png)


# ─────────────────────────────────────────────────────────────────────────────
#  FUNCIÓN EXPORTAR EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def df_to_sheet(ws, df, header_fill='1a6e9e'):
    """Escribe un DataFrame en una hoja openpyxl con formato."""
    fill   = PatternFill('solid', fgColor=header_fill)
    font_h = Font(bold=True, color='FFFFFF')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if r_idx == 1:
                cell.fill = fill
                cell.font = font_h
        if r_idx == 1:
            ws.row_dimensions[1].height = 20

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)


# =============================================================================
#  INICIO EJECUCIÓN
# =============================================================================

titulo_consola(f'ISIMIP BIAS CORRECTION — Análisis MENSUAL y ANUAL\n  Modelo: {MODELO}')

# ─── PASO 1: Lectura de datos ─────────────────────────────────────────────────
titulo_consola('PASO 1 — Lectura de archivos de entrada', nivel=2)

print(f'  OBS  → {OBS_XLSX}')
obs_raw = pd.read_excel(OBS_XLSX, parse_dates=['fecha'], index_col='fecha')
print(f'  GCM  → {GCM_XLSX}')
gcm_raw = pd.read_excel(GCM_XLSX, parse_dates=['fecha'], index_col='fecha')

obs_d = obs_raw[COL_OBS].sort_index()
gcm_d = gcm_raw[COL_GCM].sort_index()

obs_cal_d = obs_d[CAL_INI:CAL_FIN]
gcm_cal_d = gcm_d[CAL_INI:CAL_FIN]
gcm_fut_d = gcm_d[FUT_INI:FUT_FIN]

print(f'\n  Período calibración : {CAL_INI} → {CAL_FIN}')
print(f'  Período futuro      : {FUT_INI} → {FUT_FIN}')
print(f'  Datos diarios OBS   : {len(obs_cal_d):,} registros')
print(f'  Datos diarios GCM   : {len(gcm_cal_d):,} (cal)  |  {len(gcm_fut_d):,} (fut)')

# ─── PASO 2: Agregación mensual y anual ──────────────────────────────────────
titulo_consola('PASO 2 — Agregación a escala MENSUAL y ANUAL (suma mm)', nivel=2)

obs_men  = obs_cal_d.resample('ME').sum()
gcm_cal_m = gcm_cal_d.resample('ME').sum()
gcm_fut_m = gcm_fut_d.resample('ME').sum()

obs_anu  = obs_cal_d.resample('YE').sum()
gcm_cal_a = gcm_cal_d.resample('YE').sum()
gcm_fut_a = gcm_fut_d.resample('YE').sum()

print(f'  Mensual — OBS_cal: {len(obs_men)} meses  |  GCM_cal: {len(gcm_cal_m)} meses  |  GCM_fut: {len(gcm_fut_m)} meses')
print(f'  Anual   — OBS_cal: {len(obs_anu)} años   |  GCM_cal: {len(gcm_cal_a)} años   |  GCM_fut: {len(gcm_fut_a)} años')

# ─── PASO 3: Corrección ISIMIP — Futuro ──────────────────────────────────────
titulo_consola('PASO 3 — Corrección ISIMIP al período FUTURO', nivel=2)
print('  (No hay OBS futuro; corrección sin evaluación directa)')

print('  [Mensual] Empírico...', end=' ')
gcm_fut_m_emp = isimip_bias_correction(obs_men, gcm_cal_m, gcm_fut_m, 'empirica', N_QUANTILES)
print('OK')

print('  [Mensual] Gamma...', end=' ')
gcm_fut_m_gam = isimip_bias_correction(obs_men, gcm_cal_m, gcm_fut_m, 'gamma', N_QUANTILES, UMBRAL_MM)
print('OK')

print('  [Anual]   Empírico...', end=' ')
gcm_fut_a_emp = isimip_bias_correction(obs_anu, gcm_cal_a, gcm_fut_a, 'empirica', N_QUANTILES)
print('OK')

print('  [Anual]   Gamma...', end=' ')
gcm_fut_a_gam = isimip_bias_correction(obs_anu, gcm_cal_a, gcm_fut_a, 'gamma', N_QUANTILES, UMBRAL_MM)
print('OK')

# ─── PASO 4: Retroproyección en calibración ──────────────────────────────────
titulo_consola('PASO 4 — Retroproyección en CALIBRACIÓN (para poder comparar con OBS)', nivel=2)
print('  Se corrige el propio GCM_cal usando GCM_cal como "futuro" y se compara con OBS_cal.')

print('  [Mensual] Empírico...', end=' ')
gcm_cal_m_emp = isimip_bias_correction(obs_men, gcm_cal_m, gcm_cal_m, 'empirica', N_QUANTILES)
print('OK')

print('  [Mensual] Gamma...', end=' ')
gcm_cal_m_gam = isimip_bias_correction(obs_men, gcm_cal_m, gcm_cal_m, 'gamma', N_QUANTILES, UMBRAL_MM)
print('OK')

print('  [Anual]   Empírico...', end=' ')
gcm_cal_a_emp = isimip_bias_correction(obs_anu, gcm_cal_a, gcm_cal_a, 'empirica', N_QUANTILES)
print('OK')

print('  [Anual]   Gamma...', end=' ')
gcm_cal_a_gam = isimip_bias_correction(obs_anu, gcm_cal_a, gcm_cal_a, 'gamma', N_QUANTILES, UMBRAL_MM)
print('OK')

# ─── PASO 5: Métricas de ajuste ──────────────────────────────────────────────
titulo_consola('PASO 5 — MÉTRICAS DE AJUSTE en Calibración', nivel=2)

print("""
  ╔════════════════════════════════════════════════════════════════════════════╗
  ║  NOTA CRUCIAL SOBRE EVALUACIÓN DE QUANTILE MAPPING (QM/ISIMIP)             ║
  ╠════════════════════════════════════════════════════════════════════════════╣
  ║                                                                            ║
  ║  QM es un método de CORRECCIÓN DE DISTRIBUCIÓN, no de predicción:          ║
  ║                                                                            ║
  ║  ✅ CORRIGE (y estas métricas mejoran):                                   ║
  ║     • Sesgo sistemático (PBIAS)                                            ║
  ║     • Media y varianza de la serie                                         ║
  ║     • Distribución de cuantiles                                            ║
  ║     • Extremos (p5, p95, etc.)                                             ║
  ║                                                                            ║
  ║  ❌ NO PRESERVA (y estas métricas típicamente empeoran):                  ║
  ║     • Sincronía temporal día-a-día (NSE baja)                              ║
  ║     • Correlación interanual perfecta (r baja)                             ║
  ║     • KGE como métrica convencional                                        ║
  ║                                                                            ║
  ║  📊 MÉTRICAS CLAVE para evaluar QM:                                       ║
  ║     1. PBIAS (%)      → Sesgo de volumen (0% = perfecto)                   ║
  ║     2. Error Media (%) → Error en media anual/mensual (<3% = excelente)    ║
  ║     3. Error StdDev(%) → Error en variabilidad (<5% = excelente)           ║
  ║                                                                            ║
  ║  📈 VALIDACIÓN VISUAL (revisa siempre):                                   ║
  ║     • FDC: ¿líneas OBS, empírico y gamma coinciden?                        ║
  ║     • QQ-plot: ¿puntos siguen diagonal 1:1?                                ║
  ║     • Histograma: ¿distribuciones se solapan?                              ║
  ║                                                                            ║
  ╚════════════════════════════════════════════════════════════════════════════╝
""")

# ── Escala Mensual ──
titulo_consola('  ESCALA MENSUAL', nivel=2)
m_sin_m = calcular_metricas(obs_men, gcm_cal_m,     'Sin corrección')
m_emp_m = calcular_metricas(obs_men, gcm_cal_m_emp, 'ISIMIP Empírico')
m_gam_m = calcular_metricas(obs_men, gcm_cal_m_gam, 'ISIMIP Gamma')

for m in [m_sin_m, m_emp_m, m_gam_m]:
    imprimir_metricas(m, 'Mensual')

# ── Escala Anual ──
titulo_consola('  ESCALA ANUAL', nivel=2)
m_sin_a = calcular_metricas(obs_anu, gcm_cal_a,     'Sin corrección')
m_emp_a = calcular_metricas(obs_anu, gcm_cal_a_emp, 'ISIMIP Empírico')
m_gam_a = calcular_metricas(obs_anu, gcm_cal_a_gam, 'ISIMIP Gamma')

for m in [m_sin_a, m_emp_a, m_gam_a]:
    imprimir_metricas(m, 'Anual')

# ─── PASO 6: Conclusión automática ───────────────────────────────────────────
titulo_consola('PASO 6 — CONCLUSIÓN: ¿Es correcto usar ISIMIP en este caso?')

# Evaluar ambas escalas
print('\n  EVALUACIÓN POR ESCALA (Criterios correctos para QM):')
print('  ' + '─' * 76)

# Mensual empírico
icon_m, nivel_m, desc_m = clasificar_metrica_qm(m_emp_m, 'mensual')
print(f'\n  ESCALA MENSUAL (ISIMIP Empírico):')
print(f'    Diagnóstico: {icon_m} {nivel_m}')
print(f'    PBIAS: {m_emp_m["PBIAS (%)"]:.2f}% | Error Media: {m_emp_m["Error Media (%)"]:.2f}% | Error StdDev: {m_emp_m["Error StdDev (%)"]:.2f}%')

# Anual empírico
icon_a, nivel_a, desc_a = clasificar_metrica_qm(m_emp_a, 'anual')
print(f'\n  ESCALA ANUAL (ISIMIP Empírico):')
print(f'    Diagnóstico: {icon_a} {nivel_a}')
print(f'    PBIAS: {m_emp_a["PBIAS (%)"]:.2f}% | Error Media: {m_emp_a["Error Media (%)"]:.2f}% | Error StdDev: {m_emp_a["Error StdDev (%)"]:.2f}%')

# Determinar recomendación global (prioriza anual)
if m_emp_a['PBIAS (%)'] < 15 and m_emp_a['Error Media (%)'] < 10:
    icon_global = icon_a
    nivel_global = nivel_a
    desc_global = desc_a
    escala_recomendada = 'ANUAL'
elif m_emp_m['PBIAS (%)'] < 15 and m_emp_m['Error Media (%)'] < 10:
    icon_global = icon_m
    nivel_global = nivel_m
    desc_global = desc_m
    escala_recomendada = 'MENSUAL'
else:
    icon_global = '❌'
    nivel_global = 'DEFICIENTE'
    desc_global = ('El ajuste no es satisfactorio en ninguna escala. '
                   'Revisa calidad de datos o considera otros métodos.')
    escala_recomendada = 'NINGUNA'

print(f'\n  ' + '─' * 76)
print(f'\n  ╔══════════════════════════════════════════════════════════════════════════════╗')
print(f'  ║  RECOMENDACIÓN FINAL: {icon_global} {nivel_global:<60} ║')
print(f'  ║  Escala prioritaria: {escala_recomendada:<55} ║')
print(f'  ╚══════════════════════════════════════════════════════════════════════════════╝')
print(f'\n  {desc_global}')

print(f"""
  ╔════════════════════════════════════════════════════════════════════════════╗
  ║  CRITERIOS DE CLASIFICACIÓN PARA QUANTILE MAPPING                         ║
  ╠════════════════════════════════════════════════════════════════════════════╣
  │  ✅✅✅ EXCELENTE  │ PBIAS < 5%  y  Error Media < 3%  y  Error Std < 3%   │
  │  ✅✅  MUY BUENO  │ PBIAS < 10% y  Error Media < 5%  y  Error Std < 5%   │
  │  ✅  BUENO      │ PBIAS < 15% y  Error Media < 10% y  Error Std < 10%  │
  │  ⚠️  ACEPTABLE   │ PBIAS < 25% y  Error Media < 15% y  Error Std < 15%  │
  │  ❌  DEFICIENTE  │ Fuera de los rangos anteriores                        │
  ╚════════════════════════════════════════════════════════════════════════════╝
""")

print(f"""
  RECOMENDACIONES PRÁCTICAS:
  ─────────────────────────────────────────────────────────────────────────────
  ✔ Si necesitas PROYECCIONES DE CAUDAL PROMEDIO (anual/mensual):
    → La corrección ISIMIP es ADECUADA incluso con PBIAS <5%

  ✔ Si necesitas ESTUDIOS DE TENDENCIA O CAMBIO CLIMÁTICO:
    → ISIMIP es ROBUSTO para detectar cambios en distribución

  ✔ Si necesitas ANÁLISIS DE EXTREMOS (sequías, inundaciones):
    → Revisa gráficos FDC/QQ-plot y valida con datos observados

  ✔ Si necesitas SIMULACIONES DÍA-A-DÍA (variabilidad temporal):
    → ISIMIP NO es la mejor opción; considera otros métodos

  ✔ SIEMPRE:
    → Valida visualmente con FDC, QQ-plot e histogramas
    → Complementa con análisis estacional/mensual si es crítico
  ─────────────────────────────────────────────────────────────────────────────
""")

# ─── PASO 7: Gráficos PNG ────────────────────────────────────────────────────
titulo_consola('PASO 7 — Generando gráficos PNG (FDC, Histograma, QQ-plot)', nivel=2)
pngs = []

# ── Mensual ──
print('\n  [MENSUAL]')

series_fdc_m = {
    'OBS'            : obs_men,
    'GCM sin corr.'  : gcm_cal_m,
    'ISIMIP Empírico': gcm_cal_m_emp,
    'ISIMIP Gamma'   : gcm_cal_m_gam,
}
p = plot_fdc(series_fdc_m,
             os.path.join(PNG_DIR, f'{MODELO}_FDC_mensual.png'),
             f'FDC Mensual — Calibración — {MODELO}')
pngs.append({'Gráfico': 'FDC Mensual', 'Ruta': p})

p = plot_hist(series_fdc_m,
              os.path.join(PNG_DIR, f'{MODELO}_Hist_mensual.png'),
              f'Histograma Mensual — Calibración — {MODELO}')
pngs.append({'Gráfico': 'Histograma Mensual', 'Ruta': p})

sims_qq_m = {
    'GCM sin corr.'  : gcm_cal_m,
    'ISIMIP Empírico': gcm_cal_m_emp,
    'ISIMIP Gamma'   : gcm_cal_m_gam,
}
p = plot_qq(obs_men, sims_qq_m,
            os.path.join(PNG_DIR, f'{MODELO}_QQ_mensual.png'),
            f'QQ-plot Mensual — Calibración — {MODELO}')
pngs.append({'Gráfico': 'QQ-plot Mensual', 'Ruta': p})

# ── Anual ──
print('\n  [ANUAL]')

series_fdc_a = {
    'OBS'            : obs_anu,
    'GCM sin corr.'  : gcm_cal_a,
    'ISIMIP Empírico': gcm_cal_a_emp,
    'ISIMIP Gamma'   : gcm_cal_a_gam,
}
p = plot_fdc(series_fdc_a,
             os.path.join(PNG_DIR, f'{MODELO}_FDC_anual.png'),
             f'FDC Anual — Calibración — {MODELO}')
pngs.append({'Gráfico': 'FDC Anual', 'Ruta': p})

p = plot_hist(series_fdc_a,
              os.path.join(PNG_DIR, f'{MODELO}_Hist_anual.png'),
              f'Histograma Anual — Calibración — {MODELO}')
pngs.append({'Gráfico': 'Histograma Anual', 'Ruta': p})

sims_qq_a = {
    'GCM sin corr.'  : gcm_cal_a,
    'ISIMIP Empírico': gcm_cal_a_emp,
    'ISIMIP Gamma'   : gcm_cal_a_gam,
}
p = plot_qq(obs_anu, sims_qq_a,
            os.path.join(PNG_DIR, f'{MODELO}_QQ_anual.png'),
            f'QQ-plot Anual — Calibración — {MODELO}')
pngs.append({'Gráfico': 'QQ-plot Anual', 'Ruta': p})

# ─── PASO 8: Exportar Excel ──────────────────────────────────────────────────
titulo_consola('PASO 8 — Exportando resultados a Excel (.xlsx)', nivel=2)

wb = Workbook()
wb.remove(wb.active)

# ── Hoja: calibracion_mensual ──
ws = wb.create_sheet('calibracion_mensual')
df_cal_m = pd.DataFrame({
    'obs_mm_mes'         : obs_men,
    'gcm_sin_corr_mm_mes': gcm_cal_m,
    'gcm_emp_cal_mm_mes' : gcm_cal_m_emp,
    'gcm_gam_cal_mm_mes' : gcm_cal_m_gam,
})
df_to_sheet(ws, df_cal_m)
print('  Hoja "calibracion_mensual" ... OK')

# ── Hoja: futuro_mensual ──
ws = wb.create_sheet('futuro_mensual')
df_fut_m = pd.DataFrame({
    'gcm_sin_corr_mm_mes': gcm_fut_m,
    'gcm_emp_fut_mm_mes' : gcm_fut_m_emp,
    'gcm_gam_fut_mm_mes' : gcm_fut_m_gam,
})
df_to_sheet(ws, df_fut_m)
print('  Hoja "futuro_mensual" ........ OK')

# ── Hoja: calibracion_anual ──
ws = wb.create_sheet('calibracion_anual')
df_cal_a = pd.DataFrame({
    'obs_mm_año'         : obs_anu,
    'gcm_sin_corr_mm_año': gcm_cal_a,
    'gcm_emp_cal_mm_año' : gcm_cal_a_emp,
    'gcm_gam_cal_mm_año' : gcm_cal_a_gam,
})
df_to_sheet(ws, df_cal_a)
print('  Hoja "calibracion_anual" ..... OK')

# ── Hoja: futuro_anual ──
ws = wb.create_sheet('futuro_anual')
df_fut_a = pd.DataFrame({
    'gcm_sin_corr_mm_año': gcm_fut_a,
    'gcm_emp_fut_mm_año' : gcm_fut_a_emp,
    'gcm_gam_fut_mm_año' : gcm_fut_a_gam,
})
df_to_sheet(ws, df_fut_a)
print('  Hoja "futuro_anual" .......... OK')

# ── Hoja: metricas ──
ws = wb.create_sheet('metricas')
filas_met = []
for escala, lista in [('Mensual', [m_sin_m, m_emp_m, m_gam_m]),
                       ('Anual',   [m_sin_a, m_emp_a, m_gam_a])]:
    for m in lista:
        if m:
            row = {'Escala': escala}
            row.update(m)
            filas_met.append(row)
df_met = pd.DataFrame(filas_met)
df_to_sheet(ws, df_met, header_fill='1a6e9e')
print('  Hoja "metricas" .............. OK')

# ── Hoja: graficos ──
ws = wb.create_sheet('graficos')
df_pngs = pd.DataFrame(pngs)
df_to_sheet(ws, df_pngs, header_fill='2ca02c')
print('  Hoja "graficos" .............. OK')

# ── Hoja: resumen ──
ws = wb.create_sheet('resumen')
ws.column_dimensions['A'].width = 95
lineas = [
    f'═══════════════════════════════════════════════════════════════════════════════════════════════════',
    f'RESUMEN EJECUTIVO — ISIMIP Bias Correction | Modelo: {MODELO}',
    f'═══════════════════════════════════════════════════════════════════════════════════════════════════',
    '',
    f'PERÍODO DE ANÁLISIS:',
    f'  Calibración : {CAL_INI} → {CAL_FIN}',
    f'  Futuro      : {FUT_INI} → {FUT_FIN}',
    '',
    f'DIAGNÓSTICO FINAL:',
    f'  {icon_global} {nivel_global}',
    f'  Escala prioritaria: {escala_recomendada}',
    '',
    f'  {desc_global}',
    '',
    f'MÉTRICAS CLAVE (ISIMIP Empírico):',
    f'',
    f'  ESCALA MENSUAL:',
    f'    PBIAS         : {m_emp_m["PBIAS (%)"]:.2f}%',
    f'    Error Media   : {m_emp_m["Error Media (%)"]:.2f}%',
    f'    Error StdDev  : {m_emp_m["Error StdDev (%)"]:.2f}%',
    f'',
    f'  ESCALA ANUAL:',
    f'    PBIAS         : {m_emp_a["PBIAS (%)"]:.2f}%',
    f'    Error Media   : {m_emp_a["Error Media (%)"]:.2f}%',
    f'    Error StdDev  : {m_emp_a["Error StdDev (%)"]:.2f}%',
    '',
    f'CRITERIOS USADOS (QM):',
    f'  ✅✅✅ EXCELENTE : PBIAS<5% y Error Media<3% y Error Std<3%',
    f'  ✅✅  MUY BUENO : PBIAS<10% y Error Media<5% y Error Std<5%',
    f'  ✅  BUENO     : PBIAS<15% y Error Media<10% y Error Std<10%',
    f'  ⚠️  ACEPTABLE  : PBIAS<25% y Error Media<15% y Error Std<15%',
    f'  ❌  DEFICIENTE : Fuera de los rangos anteriores',
    '',
    f'RECOMENDACIONES:',
    f'  • Si necesitas caudal promedio (anual/mensual) → ISIMIP es ADECUADO',
    f'  • Si necesitas proyecciones de tendencia → ISIMIP es ROBUSTO',
    f'  • Si necesitas análisis de extremos → Valida con gráficos FDC/QQ',
    f'  • Si necesitas variabilidad día-a-día → Considera otros métodos',
    '',
    f'ARCHIVOS GENERADOS:',
] + [f'  {p["Gráfico"]} → {os.path.basename(p["Ruta"])}' for p in pngs]

for i, linea in enumerate(lineas, 1):
    c = ws.cell(row=i, column=1, value=linea)
    if i in [1, 2, 3]:
        c.font = Font(bold=True, size=11, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1a6e9e')
    elif 'DIAGNÓSTICO' in linea or 'MÉTRICAS' in linea or 'CRITERIOS' in linea:
        c.font = Font(bold=True, size=10)
    elif any(x in linea for x in ['✅', '⚠️', '❌']):
        c.font = Font(size=10)

wb.save(SALIDA_XLSX)
print(f'\n  ✔ Excel guardado → {SALIDA_XLSX}')

# ─── FIN ─────────────────────────────────────────────────────────────────────
titulo_consola(f'PROCESO COMPLETADO EXITOSAMENTE')
print(f"""
  ╔═══════════════════════════════════════════════════════════════════════════╗
  ║  ARCHIVOS GENERADOS                                                       ║
  ╠═══════════════════════════════════════════════════════════════════════════╣
  │  Excel  → {SALIDA_XLSX}
  │  PNGs   → {PNG_DIR}
  │    - GFDL-ESM4_FDC_mensual.png
  │    - GFDL-ESM4_Hist_mensual.png
  │    - GFDL-ESM4_QQ_mensual.png
  │    - GFDL-ESM4_FDC_anual.png
  │    - GFDL-ESM4_Hist_anual.png
  │    - GFDL-ESM4_QQ_anual.png
  ╚═══════════════════════════════════════════════════════════════════════════╝

  PRÓXIMOS PASOS:
  ─────────────────────────────────────────────────────────────────────────────
  1. Abre el Excel → revisa hojas "metricas" y "resumen"
  2. Abre los gráficos PNG → valida FDC, QQ-plot e histogramas
  3. Si FDC/QQ muestran buen ajuste → ISIMIP es CONFIABLE para el futuro
  4. Usa los datos de "futuro_anual" o "futuro_mensual" en tus análisis
  ─────────────────────────────────────────────────────────────────────────────
""")