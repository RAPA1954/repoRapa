# =============================================================================
#  ISIMIP3 BIAS CORRECTION — Análisis MENSUAL y ANUAL
#  Implementación de la versión ISIMIP3 (Lange 2019/2021)
#
#  Mejoras respecto a versiones anteriores (ISIMIP2b):
#   1. Corrección de frecuencia de días secos por cuantil empírico
#   2. Ajuste paramétrico de distribución gamma (días húmedos)
#   3. Corrección de colas extremas con Generalized Pareto Distribution (GPD)
#   4. Preservación del cambio de señal futuro (delta scaling aditivo/multiplicativo)
#   5. Tres modos:
#       • 'empirica'   → ISIMIP2b (mapeo cuantil empírico, referencia)
#       • 'isimip3'    → ISIMIP3 estándar (gamma + preservación de tendencia)
#       • 'isimip3+'   → ISIMIP3 con corrección de extremos GPD
#
#  Referencia: Lange (2019) https://doi.org/10.5194/gmd-12-3055-2019
#              ISIMIP3 protocol https://www.isimip.org/protocol/
# =============================================================================

import os
import warnings
import numpy as np
import pandas as pd
from scipy.stats   import gamma as gamma_dist, genpareto, pearsonr
from scipy.special import ndtri
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')

plt.rcParams.update({
    'font.family'       : 'DejaVu Sans',
    'axes.unicode_minus': False,
    'figure.dpi'        : 150,
})

# =============================================================================
#  CONFIGURACIÓN USUARIO  ← EDITAR AQUÍ
# =============================================================================
OBS_XLSX    = r'C:/1.PYTHON/Descarga_Python/Pd_Historica_SMN.xlsx'
GCM_XLSX    = r'C:/1.PYTHON/Descarga_Python/Tabla_CESM2_historico_futuro.xlsx'
COL_OBS     = 'prec'
COL_GCM     = 'pr_mm'
MODELO      = 'CESM2'
CAL_INI     = '1961-01-01'
CAL_FIN     = '2024-12-31'
FUT_INI     = '2030-01-01'
FUT_FIN     = '2100-12-31'
OUT_DIR     = rf'C:/1.PYTHON/Descarga_Python/{MODELO}_salidas_ISIMIP3'
N_QUANTILES = 100        # Cuantiles para mapeo empírico
UMBRAL_MM   = 0.1        # mm — umbral para distinguir día seco/húmedo
P_EXTREMO   = 0.95       # Percentil a partir del cual se ajusta la cola GPD (ISIMIP3+)
# =============================================================================

os.makedirs(OUT_DIR, exist_ok=True)
PNG_DIR = os.path.join(OUT_DIR, 'graficos')
os.makedirs(PNG_DIR, exist_ok=True)
SALIDA_XLSX = os.path.join(OUT_DIR, f'{MODELO}_ISIMIP3_mensual_anual.xlsx')

SEP  = '=' * 78
SEP2 = '-' * 78


# =============================================================================
#  FUNCIONES AUXILIARES
# =============================================================================

def titulo_consola(texto, nivel=1):
    if nivel == 1:
        print(f'\n{SEP}')
        print(f'  {texto}')
        print(SEP)
    else:
        print(f'\n{SEP2}')
        print(f'  {texto}')
        print(SEP2)


# ─────────────────────────────────────────────────────────────────────────────
#  BLOQUE CENTRAL ISIMIP3
# ─────────────────────────────────────────────────────────────────────────────

def _ajustar_umbral_lluvia(obs_cal, gcm_cal, umbral):
    """
    ISIMIP3 — Paso 1: Corrección de frecuencia de días secos.

    El umbral del GCM (p0_gcm) se ajusta de modo que la fracción de días
    secos en el GCM calibrado coincida con la fracción observada.

    Ecuación:
        p0_gcm* = F_gcm_cal^{-1}( F_obs_cal(p0_obs) )

    Parámetros
    ----------
    obs_cal : pd.Series  — precipitación observada en calibración
    gcm_cal : pd.Series  — precipitación del GCM en calibración
    umbral  : float      — umbral de día seco en obs (mm)

    Retorna
    -------
    p0_gcm_ajustado : float — nuevo umbral para clasificar días secos en GCM
    """
    # Fracción de días secos en observaciones
    frac_seca_obs = np.mean(obs_cal <= umbral)
    # Cuantil equivalente en el GCM: F_gcm^{-1}(frac_seca_obs)
    p0_gcm = np.quantile(gcm_cal, frac_seca_obs)
    return max(p0_gcm, 0.0)


def _mapeo_cuantil_empirico(obs_wet, gcm_cal_wet, gcm_fut_wet):
    """
    ISIMIP3 — Paso 2a: Mapeo cuantil empírico sobre días húmedos.

    X_corr(tau) = F_obs^{-1}( F_gcm_cal( X_gcm_fut ) )

    Retorna valores corregidos para gcm_fut_wet.
    """
    quantiles   = np.linspace(0, 1, N_QUANTILES + 1)
    q_obs       = np.quantile(obs_wet,     quantiles)
    q_gcm_cal   = np.quantile(gcm_cal_wet, quantiles)

    # Mapear cada valor futuro a su cuantil en la distribución de calibración
    corr = np.interp(gcm_fut_wet, q_gcm_cal, q_obs,
                     left=q_obs[0], right=q_obs[-1])
    return corr


def _ajuste_gamma(serie):
    """Ajusta distribución Gamma (shape, scale) con floc=0. Retorna (a, scale)."""
    a, _, sc = gamma_dist.fit(serie, floc=0)
    return a, sc


def _mapeo_cuantil_gamma(obs_wet, gcm_cal_wet, gcm_fut_wet):
    """
    ISIMIP3 — Paso 2b: Mapeo cuantil paramétrico con distribución Gamma.

    1. Ajusta Gamma a obs_cal y gcm_cal (días húmedos).
    2. Calcula cuantil de cada valor futuro en la Gamma del GCM.
    3. Aplica la inversa de la Gamma observada.

    X_corr = F_obs^{-1}( F_gcm_cal( X_gcm_fut ) )   — todo paramétrico
    """
    a_obs, sc_obs = _ajuste_gamma(obs_wet)
    a_cal, sc_cal = _ajuste_gamma(gcm_cal_wet)

    # Probabilidad acumulada de cada valor futuro bajo la distribución del GCM cal
    p = gamma_dist.cdf(gcm_fut_wet, a_cal, scale=sc_cal)
    p = np.clip(p, 1e-6, 1 - 1e-6)   # evitar 0 y 1 exactos

    # Inversa en la distribución observada
    corr = gamma_dist.ppf(p, a_obs, scale=sc_obs)
    return corr


def _correccion_cola_gpd(obs_wet, gcm_cal_wet, gcm_fut_wet, p_umbral=0.95):
    """
    ISIMIP3+ — Paso 2c: Corrección de extremos con Generalized Pareto Distribution.

    Para valores por encima del percentil p_umbral:
      • Ajusta GPD a los excesos de obs y gcm_cal por encima de sus umbrales.
      • Mapea los excesos futuros del GCM usando las dos GPDs.

    Ecuación de preservación de excesos:
        y_corr = u_obs + (sc_obs/sc_cal) * (x_gcm_fut - u_gcm)

    donde u = umbral (percentil p_umbral de cada distribución).
    """
    u_obs = np.quantile(obs_wet,     p_umbral)
    u_cal = np.quantile(gcm_cal_wet, p_umbral)

    exc_obs = obs_wet[obs_wet > u_obs]     - u_obs
    exc_cal = gcm_cal_wet[gcm_cal_wet > u_cal] - u_cal

    if len(exc_obs) < 5 or len(exc_cal) < 5:
        # Fallback: escalar linealmente por ratio de umbrales
        ratio = u_obs / u_cal if u_cal > 0 else 1.0
        return gcm_fut_wet * ratio

    # Ajuste GPD a los excesos
    xi_obs, _, sc_obs = genpareto.fit(exc_obs, floc=0)
    xi_cal, _, sc_cal = genpareto.fit(exc_cal, floc=0)

    sc_cal  = max(sc_cal,  1e-6)
    sc_obs  = max(sc_obs,  1e-6)

    corr = np.where(
        gcm_fut_wet > u_cal,
        u_obs + (sc_obs / sc_cal) * (gcm_fut_wet - u_cal),   # extremos
        gcm_fut_wet * (u_obs / u_cal) if u_cal > 0 else gcm_fut_wet  # no extremos
    )
    return corr


def _preservar_tendencia(gcm_fut_wet_raw, gcm_fut_wet_corr,
                          gcm_cal_wet, obs_wet,
                          modo='multiplicativo'):
    """
    ISIMIP3 — Paso 3: Preservación del cambio de señal futuro (delta scaling).

    El método ajusta la corrección para que el cambio relativo (o absoluto)
    proyectado por el GCM entre calibración y futuro se conserve.

    Modo multiplicativo (precipitación):
        delta = mean(gcm_fut_wet_raw) / mean(gcm_cal_wet)
        X_final = X_corr * delta

    Modo aditivo (temperatura):
        delta = mean(gcm_fut_wet_raw) - mean(gcm_cal_wet)
        X_final = X_corr + delta

    Parámetros
    ----------
    gcm_fut_wet_raw  : array — valores futuros del GCM SIN corregir (húmedos)
    gcm_fut_wet_corr : array — valores futuros YA corregidos por QM
    gcm_cal_wet      : array — valores del GCM en calibración (húmedos)
    obs_wet          : array — observaciones en calibración (húmedos)
    modo             : 'multiplicativo' | 'aditivo'
    """
    mean_cal = np.mean(gcm_cal_wet)
    mean_fut = np.mean(gcm_fut_wet_raw)
    mean_obs = np.mean(obs_wet)

    if modo == 'multiplicativo':
        delta = mean_fut / mean_cal if mean_cal > 0 else 1.0
        # Escalar la corrección para que preserve la tendencia del GCM
        result = gcm_fut_wet_corr * delta
        # Limitar la corrección para no amplificar extremos irrealmente
        max_ratio = 5.0
        result = np.clip(result, 0, mean_obs * max_ratio)
    else:
        delta  = mean_fut - mean_cal
        result = gcm_fut_wet_corr + delta

    return result


def isimip3_bias_correction(obs_cal, gcm_cal, gcm_fut,
                             modo='isimip3', umbral=UMBRAL_MM,
                             p_extremo=P_EXTREMO):
    """
    Corrección de sesgo ISIMIP3 completa.

    Modos disponibles:
    ------------------
    'empirica'  → ISIMIP2b: mapeo cuantil empírico sin preservación de tendencia
                  (referencia para comparación)

    'isimip3'   → ISIMIP3 estándar:
                  1. Corrección de frecuencia de días secos (umbral ajustado por cuantil)
                  2. Mapeo cuantil paramétrico con Gamma (días húmedos)
                  3. Preservación del cambio de señal futuro (delta multiplicativo)

    'isimip3+'  → ISIMIP3 con corrección de extremos:
                  Igual a 'isimip3' + ajuste GPD para valores > percentil p_extremo

    Parámetros
    ----------
    obs_cal  : pd.Series — observaciones en período de calibración
    gcm_cal  : pd.Series — GCM en período de calibración
    gcm_fut  : pd.Series — GCM en período futuro (a corregir)
    modo     : str       — 'empirica', 'isimip3', 'isimip3+'
    umbral   : float     — mm para umbral de día seco (basado en observaciones)
    p_extremo: float     — percentil para corrección de extremos GPD (solo 'isimip3+')

    Retorna
    -------
    pd.Series con los valores corregidos (mismos índices que gcm_fut)
    """
    obs_cal      = obs_cal.dropna()
    gcm_cal      = gcm_cal.dropna()
    gcm_fut_in   = gcm_fut.dropna()
    gcm_fut_vals = gcm_fut_in.values.copy().astype(float)

    # ── Modo empírico (ISIMIP2b, sin corrección de tendencia) ────────────────
    if modo == 'empirica':
        quantiles = np.linspace(0, 1, N_QUANTILES + 1)
        q_obs     = np.quantile(obs_cal,  quantiles)
        q_gcm_cal = np.quantile(gcm_cal,  quantiles)
        q_gcm_fut = np.quantile(gcm_fut_vals, quantiles)
        transfer  = q_obs - q_gcm_cal
        q_corr    = q_gcm_fut + transfer
        corr_vals = np.interp(gcm_fut_vals, q_gcm_fut, q_corr,
                              left=q_corr[0], right=q_corr[-1])
        return pd.Series(np.clip(corr_vals, 0, None), index=gcm_fut_in.index)

    # ── Modos ISIMIP3 y ISIMIP3+ ─────────────────────────────────────────────

    # --- PASO 1: Ajustar umbral de día seco para el GCM ---------------------
    p0_gcm = _ajustar_umbral_lluvia(obs_cal.values, gcm_cal.values, umbral)

    # Máscara de días húmedos/secos
    mask_obs_wet = obs_cal.values  > umbral
    mask_cal_wet = gcm_cal.values  > p0_gcm
    mask_fut_wet = gcm_fut_vals    > p0_gcm

    obs_wet     = obs_cal.values[mask_obs_wet]
    gcm_cal_wet = gcm_cal.values[mask_cal_wet]
    gcm_fut_wet = gcm_fut_vals[mask_fut_wet]

    # Fallback a empírico si hay muy pocos datos húmedos
    if len(obs_wet) < 15 or len(gcm_cal_wet) < 15 or len(gcm_fut_wet) < 5:
        return isimip3_bias_correction(obs_cal, gcm_cal, gcm_fut,
                                       modo='empirica', umbral=umbral)

    # --- PASO 2: Mapeo cuantil sobre días húmedos ---------------------------
    if modo == 'isimip3':
        corr_wet = _mapeo_cuantil_gamma(obs_wet, gcm_cal_wet, gcm_fut_wet)

    elif modo == 'isimip3+':
        # Mapeo gamma base
        corr_base = _mapeo_cuantil_gamma(obs_wet, gcm_cal_wet, gcm_fut_wet)
        # Corrección de extremos con GPD
        corr_ext  = _correccion_cola_gpd(obs_wet, gcm_cal_wet, gcm_fut_wet,
                                          p_umbral=p_extremo)
        # Combinar: usar GPD solo para valores por encima del umbral extremo
        u_cal_ext = np.quantile(gcm_cal_wet, p_extremo)
        corr_wet  = np.where(gcm_fut_wet > u_cal_ext, corr_ext, corr_base)
    else:
        raise ValueError("modo debe ser 'empirica', 'isimip3' o 'isimip3+'")

    # --- PASO 3: Preservación del cambio de señal futuro (delta scaling) ----
    corr_wet_final = _preservar_tendencia(
        gcm_fut_wet, corr_wet, gcm_cal_wet, obs_wet, modo='multiplicativo'
    )

    # --- PASO 4: Reconstruir la serie completa ------------------------------
    result = np.zeros_like(gcm_fut_vals)
    result[mask_fut_wet]  = np.clip(corr_wet_final, 0, None)
    result[~mask_fut_wet] = 0.0   # días secos → 0

    return pd.Series(result, index=gcm_fut_in.index)


# =============================================================================
#  MÉTRICAS Y EVALUACIÓN
# =============================================================================

def calcular_metricas(obs, sim, nombre=''):
    """Calcula métricas de evaluación. Métricas clave para QM: PBIAS, Error Media, Error StdDev."""
    df = pd.DataFrame({'obs': obs, 'sim': sim}).dropna()
    o  = df['obs'].values
    s  = df['sim'].values
    n  = len(o)

    if n < 3:
        return None

    nse   = 1 - np.sum((o - s)**2) / np.sum((o - np.mean(o))**2)
    r, _  = pearsonr(o, s)
    alpha = np.std(s)  / np.std(o)
    beta  = np.mean(s) / np.mean(o)
    kge   = 1 - np.sqrt((r - 1)**2 + (alpha - 1)**2 + (beta - 1)**2)
    pbias = (np.sum(s - o) / np.sum(o)) * 100
    rmse  = np.sqrt(np.mean((o - s)**2))
    mae   = np.mean(np.abs(o - s))
    r2    = r**2

    media_obs  = np.mean(o)
    media_sim  = np.mean(s)
    std_obs    = np.std(o)
    std_sim    = np.std(s)
    err_media  = abs(media_obs - media_sim) / media_obs * 100 if media_obs != 0 else 0
    err_std    = abs(std_obs  - std_sim)   / std_obs   * 100 if std_obs   != 0 else 0

    # Métricas de cuantiles (específicas ISIMIP3)
    q95_obs = np.quantile(o, 0.95)
    q95_sim = np.quantile(s, 0.95)
    q99_obs = np.quantile(o, 0.99)
    q99_sim = np.quantile(s, 0.99)
    err_q95 = abs(q95_obs - q95_sim) / q95_obs * 100 if q95_obs != 0 else 0
    err_q99 = abs(q99_obs - q99_sim) / q99_obs * 100 if q99_obs != 0 else 0

    return {
        'Nombre'          : nombre,
        'n'               : n,
        'NSE'             : round(nse,        4),
        'KGE'             : round(kge,        4),
        'r (Pearson)'     : round(r,          4),
        'R²'              : round(r2,         4),
        'PBIAS (%)'       : round(pbias,      2),
        'RMSE (mm)'       : round(rmse,       4),
        'MAE (mm)'        : round(mae,        4),
        'Media OBS'       : round(media_obs,  4),
        'Media SIM'       : round(media_sim,  4),
        'Error Media (%)'  : round(err_media,  2),
        'StdDev OBS'      : round(std_obs,    4),
        'StdDev SIM'      : round(std_sim,    4),
        'Error StdDev (%)': round(err_std,    2),
        'Q95 OBS'         : round(q95_obs,    4),
        'Q95 SIM'         : round(q95_sim,    4),
        'Error Q95 (%)'   : round(err_q95,    2),
        'Q99 OBS'         : round(q99_obs,    4),
        'Q99 SIM'         : round(q99_sim,    4),
        'Error Q99 (%)'   : round(err_q99,    2),
    }


def imprimir_metricas(m, escala=''):
    if m is None:
        print('  !! Insuficientes datos para calcular métricas.')
        return
    print(f'\n  ── {m["Nombre"]} ({escala}) ──')
    print(f'     n               : {m["n"]}')
    print(f'     NSE             : {m["NSE"]:>8.4f}   (ℹ️  informativo: QM no preserva sincronía)')
    print(f'     KGE             : {m["KGE"]:>8.4f}   (ℹ️  informativo)')
    print(f'     r  (Pearson)    : {m["r (Pearson)"]:>8.4f}   (ℹ️  informativo)')
    print(f'     R²              : {m["R²"]:>8.4f}')
    print()
    print(f'     ✅ PBIAS (%)        : {m["PBIAS (%)"]:>8.2f}   ← MÉTRICA CLAVE')
    print(f'     ✅ Error Media (%)  : {m["Error Media (%)"]:>8.2f}   ← MÉTRICA CLAVE')
    print(f'     ✅ Error StdDev (%) : {m["Error StdDev (%)"]:>8.2f}   ← MÉTRICA CLAVE')
    print(f'     ✅ Error Q95 (%)    : {m["Error Q95 (%)"]:>8.2f}   ← NUEVO en ISIMIP3')
    print(f'     ✅ Error Q99 (%)    : {m["Error Q99 (%)"]:>8.2f}   ← NUEVO en ISIMIP3')
    print()
    print(f'     Media OBS/SIM   : {m["Media OBS"]:>8.4f} / {m["Media SIM"]:>8.4f}')
    print(f'     StdDev OBS/SIM  : {m["StdDev OBS"]:>8.4f} / {m["StdDev SIM"]:>8.4f}')
    print(f'     Q95   OBS/SIM   : {m["Q95 OBS"]:>8.4f} / {m["Q95 SIM"]:>8.4f}')
    print(f'     Q99   OBS/SIM   : {m["Q99 OBS"]:>8.4f} / {m["Q99 SIM"]:>8.4f}')
    print(f'     RMSE (mm)       : {m["RMSE (mm)"]:>8.4f}')
    print(f'     MAE  (mm)       : {m["MAE (mm)"]:>8.4f}')


def clasificar_metrica_qm(m):
    """
    Clasificación de desempeño ISIMIP3.
    Incorpora errores en cuantiles extremos (Q95, Q99) como criterio adicional.
    """
    if m is None:
        return '⚠️  SIN DATOS', 'DESCONOCIDO', 'No se pudo evaluar. Insuficientes registros.'

    pbias     = abs(m['PBIAS (%)'])
    err_media = abs(m['Error Media (%)'])
    err_std   = abs(m['Error StdDev (%)'])
    err_q95   = abs(m['Error Q95 (%)'])
    err_q99   = abs(m['Error Q99 (%)'])

    # ISIMIP3 añade criterio de extremos (Q95, Q99)
    if pbias < 5 and err_media < 3 and err_std < 3 and err_q95 < 10 and err_q99 < 15:
        return ('✅✅✅', 'EXCELENTE',
                'Ajuste excelente. ISIMIP3 es ALTAMENTE RECOMENDADO. '
                'Distribución y extremos correctamente calibrados.')

    elif pbias < 10 and err_media < 5 and err_std < 5 and err_q95 < 20 and err_q99 < 25:
        return ('✅✅', 'MUY BUENO',
                'Ajuste muy bueno. ISIMIP3 es RECOMENDADO. '
                'Distribución bien calibrada; verifica extremos en FDC.')

    elif pbias < 15 and err_media < 10 and err_std < 10 and err_q95 < 30:
        return ('✅', 'BUENO',
                'Ajuste bueno. Revisa gráficos FDC y QQ-plot, '
                'especialmente en la cola superior.')

    elif pbias < 25 and err_media < 15 and err_std < 15:
        return ('⚠️', 'ACEPTABLE',
                'Ajuste moderado. Considera validación estacional o '
                'aumentar el período de calibración.')

    else:
        return ('❌', 'DEFICIENTE',
                'Ajuste no satisfactorio. Revisa calidad de datos y '
                'representatividad del GCM para la región.')


# =============================================================================
#  GRÁFICOS
# =============================================================================

COLORES = {
    'OBS'    : '#1a6e9e',
    'SIN'    : '#d94f3d',
    'EMP'    : '#7f7f7f',
    'IS3'    : '#2ca02c',
    'IS3P'   : '#ff7f0e',
}


def _guardar(fig, ruta):
    try:
        fig.tight_layout()
    except Exception:
        pass
    fig.savefig(ruta, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f'     PNG guardado → {ruta}')
    return ruta


def plot_fdc(series_dict, ruta_png, titulo):
    """Curva de duración de precipitaciones (excedencia)."""
    fig, ax = plt.subplots(figsize=(10, 5))
    estilos = ['-', '--', '-.', ':', (0, (3, 1, 1, 1))]
    colores = list(COLORES.values())

    for i, (lbl, s) in enumerate(series_dict.items()):
        data = s.dropna().sort_values(ascending=False).values
        n    = len(data)
        prob = np.arange(1, n + 1) / (n + 1) * 100
        ax.plot(prob, data,
                linestyle=estilos[i % len(estilos)],
                color=colores[i % len(colores)],
                linewidth=1.8, label=lbl)

    ax.set_xlabel('Probabilidad de excedencia (%)', fontsize=11)
    ax.set_ylabel('Precipitación (mm)', fontsize=11)
    ax.set_title(titulo, fontsize=12, fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(True, linestyle='--', alpha=0.5)
    ax.xaxis.set_major_formatter(mticker.FormatStrFormatter('%g%%'))
    return _guardar(fig, ruta_png)


def plot_hist(series_dict, ruta_png, titulo):
    """Histogramas de densidad superpuestos."""
    fig, ax = plt.subplots(figsize=(10, 5))
    colores = list(COLORES.values())

    for i, (lbl, s) in enumerate(series_dict.items()):
        data = s.dropna().values
        ax.hist(data, bins=30, density=True, alpha=0.45,
                color=colores[i % len(colores)], label=lbl, edgecolor='white')

    ax.set_xlabel('Precipitación (mm)', fontsize=11)
    ax.set_ylabel('Densidad', fontsize=11)
    ax.set_title(titulo, fontsize=12, fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(True, linestyle='--', alpha=0.5)
    return _guardar(fig, ruta_png)


def plot_qq(obs, sims_dict, ruta_png, titulo):
    """QQ-plot: cuantiles observados vs simulados."""
    quantiles = np.linspace(0, 1, 101)
    q_obs     = np.quantile(obs.dropna(), quantiles)
    colores   = [COLORES['SIN'], COLORES['EMP'], COLORES['IS3'], COLORES['IS3P']]

    fig, ax = plt.subplots(figsize=(7, 7))
    lim_max = q_obs[-1]
    for i, (lbl, s) in enumerate(sims_dict.items()):
        q_sim = np.quantile(s.dropna(), quantiles)
        ax.scatter(q_obs, q_sim, s=18, alpha=0.7,
                   color=colores[i % len(colores)], label=lbl)
        lim_max = max(lim_max, q_sim[-1])

    ax.plot([0, lim_max], [0, lim_max], 'k--', linewidth=1.2, label='1:1 (perfecto)')
    ax.set_xlabel('Cuantiles OBS (mm)', fontsize=11)
    ax.set_ylabel('Cuantiles SIM (mm)', fontsize=11)
    ax.set_title(titulo, fontsize=12, fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(True, linestyle='--', alpha=0.5)
    return _guardar(fig, ruta_png)


def plot_extremos(obs, sims_dict, ruta_png, titulo):
    """
    Gráfico específico ISIMIP3: cola superior (valores > P75).
    Permite evaluar la corrección GPD de extremos.
    """
    quantiles = np.linspace(0.75, 0.999, 60)
    q_obs     = np.quantile(obs.dropna(), quantiles)
    colores   = [COLORES['SIN'], COLORES['EMP'], COLORES['IS3'], COLORES['IS3P']]

    fig, ax = plt.subplots(figsize=(8, 5))
    for i, (lbl, s) in enumerate(sims_dict.items()):
        q_sim = np.quantile(s.dropna(), quantiles)
        ax.plot(quantiles * 100, q_sim,
                color=colores[i % len(colores)], linewidth=1.8, label=lbl)

    ax.plot(quantiles * 100, q_obs, 'k-', linewidth=2.5, label='OBS')
    ax.set_xlabel('Percentil (%)', fontsize=11)
    ax.set_ylabel('Precipitación (mm)', fontsize=11)
    ax.set_title(titulo, fontsize=12, fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(True, linestyle='--', alpha=0.5)
    return _guardar(fig, ruta_png)


# =============================================================================
#  EXPORTAR EXCEL
# =============================================================================

def df_to_sheet(ws, df, header_fill='1a6e9e'):
    fill   = PatternFill('solid', fgColor=header_fill)
    font_h = Font(bold=True, color='FFFFFF')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border    = border
            cell.alignment = Alignment(horizontal='center')
            if r_idx == 1:
                cell.fill = fill
                cell.font = font_h
        if r_idx == 1:
            ws.row_dimensions[1].height = 20

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 32)


# =============================================================================
#  EJECUCIÓN PRINCIPAL
# =============================================================================

titulo_consola(f'ISIMIP3 BIAS CORRECTION — Análisis MENSUAL y ANUAL\n  Modelo: {MODELO}')

# ─── PASO 1: Lectura de datos ─────────────────────────────────────────────────
titulo_consola('PASO 1 — Lectura de archivos de entrada', nivel=2)

print(f'  OBS  → {OBS_XLSX}')
obs_raw = pd.read_excel(OBS_XLSX, parse_dates=['fecha'], index_col='fecha')
print(f'  GCM  → {GCM_XLSX}')
gcm_raw = pd.read_excel(GCM_XLSX, parse_dates=['fecha'], index_col='fecha')

obs_d = obs_raw[COL_OBS].sort_index()
gcm_d = gcm_raw[COL_GCM].sort_index()

obs_cal_d = obs_d[CAL_INI:CAL_FIN]
gcm_cal_d = gcm_d[CAL_INI:CAL_FIN]
gcm_fut_d = gcm_d[FUT_INI:FUT_FIN]

print(f'\n  Período calibración : {CAL_INI} → {CAL_FIN}')
print(f'  Período futuro      : {FUT_INI} → {FUT_FIN}')
print(f'  Datos diarios OBS   : {len(obs_cal_d):,} registros')
print(f'  Datos diarios GCM   : {len(gcm_cal_d):,} (cal)  |  {len(gcm_fut_d):,} (fut)')

# ─── PASO 2: Agregación mensual y anual ──────────────────────────────────────
titulo_consola('PASO 2 — Agregación a escala MENSUAL y ANUAL (suma mm)', nivel=2)

obs_men   = obs_cal_d.resample('ME').sum()
gcm_cal_m = gcm_cal_d.resample('ME').sum()
gcm_fut_m = gcm_fut_d.resample('ME').sum()

obs_anu   = obs_cal_d.resample('YE').sum()
gcm_cal_a = gcm_cal_d.resample('YE').sum()
gcm_fut_a = gcm_fut_d.resample('YE').sum()

print(f'  Mensual — OBS_cal: {len(obs_men)} meses  |  GCM_cal: {len(gcm_cal_m)} meses  |  GCM_fut: {len(gcm_fut_m)} meses')
print(f'  Anual   — OBS_cal: {len(obs_anu)} años   |  GCM_cal: {len(gcm_cal_a)} años   |  GCM_fut: {len(gcm_fut_a)} años')

# ─── PASO 3: Corrección ISIMIP3 al futuro ────────────────────────────────────
titulo_consola('PASO 3 — Corrección ISIMIP3 al período FUTURO', nivel=2)
print('  (Sin OBS futuro; corrección sin evaluación directa)')
print(f'\n  Parámetros ISIMIP3:')
print(f'    Umbral día seco (obs)  : {UMBRAL_MM} mm')
print(f'    Percentil extremos GPD : {P_EXTREMO*100:.0f}%  (solo modo isimip3+)')
print(f'    Cuantiles empíricos    : {N_QUANTILES}')

for escala, obs_s, gcm_cal_s, gcm_fut_s, sfx in [
    ('Mensual', obs_men,   gcm_cal_m, gcm_fut_m, 'm'),
    ('Anual',   obs_anu,   gcm_cal_a, gcm_fut_a, 'a'),
]:
    print(f'\n  [{escala.upper()}]')

    print(f'    Empírico (referencia ISIMIP2b)...', end=' ')
    globals()[f'gcm_fut_{sfx}_emp'] = isimip3_bias_correction(
        obs_s, gcm_cal_s, gcm_fut_s, modo='empirica')
    print('OK')

    print(f'    ISIMIP3 estándar (Gamma + delta)...', end=' ')
    globals()[f'gcm_fut_{sfx}_is3'] = isimip3_bias_correction(
        obs_s, gcm_cal_s, gcm_fut_s, modo='isimip3')
    print('OK')

    print(f'    ISIMIP3+ (Gamma + GPD extremos + delta)...', end=' ')
    globals()[f'gcm_fut_{sfx}_is3p'] = isimip3_bias_correction(
        obs_s, gcm_cal_s, gcm_fut_s, modo='isimip3+', p_extremo=P_EXTREMO)
    print('OK')

# Alias para uso posterior
gcm_fut_m_emp  = globals()['gcm_fut_m_emp']
gcm_fut_m_is3  = globals()['gcm_fut_m_is3']
gcm_fut_m_is3p = globals()['gcm_fut_m_is3p']
gcm_fut_a_emp  = globals()['gcm_fut_a_emp']
gcm_fut_a_is3  = globals()['gcm_fut_a_is3']
gcm_fut_a_is3p = globals()['gcm_fut_a_is3p']

# ─── PASO 4: Retroproyección en calibración ──────────────────────────────────
titulo_consola('PASO 4 — Retroproyección en CALIBRACIÓN (para evaluar con OBS)', nivel=2)
print('  Se aplica la corrección al propio GCM_cal y se compara con OBS_cal.')

for escala, obs_s, gcm_cal_s, sfx in [
    ('Mensual', obs_men, gcm_cal_m, 'm'),
    ('Anual',   obs_anu, gcm_cal_a, 'a'),
]:
    print(f'\n  [{escala.upper()}]')

    print(f'    Empírico...', end=' ')
    globals()[f'gcm_cal_{sfx}_emp'] = isimip3_bias_correction(
        obs_s, gcm_cal_s, gcm_cal_s, modo='empirica')
    print('OK')

    print(f'    ISIMIP3...', end=' ')
    globals()[f'gcm_cal_{sfx}_is3'] = isimip3_bias_correction(
        obs_s, gcm_cal_s, gcm_cal_s, modo='isimip3')
    print('OK')

    print(f'    ISIMIP3+...', end=' ')
    globals()[f'gcm_cal_{sfx}_is3p'] = isimip3_bias_correction(
        obs_s, gcm_cal_s, gcm_cal_s, modo='isimip3+', p_extremo=P_EXTREMO)
    print('OK')

gcm_cal_m_emp  = globals()['gcm_cal_m_emp']
gcm_cal_m_is3  = globals()['gcm_cal_m_is3']
gcm_cal_m_is3p = globals()['gcm_cal_m_is3p']
gcm_cal_a_emp  = globals()['gcm_cal_a_emp']
gcm_cal_a_is3  = globals()['gcm_cal_a_is3']
gcm_cal_a_is3p = globals()['gcm_cal_a_is3p']

# ─── PASO 5: Métricas de ajuste ──────────────────────────────────────────────
titulo_consola('PASO 5 — MÉTRICAS DE AJUSTE en Calibración', nivel=2)

print("""
  ╔════════════════════════════════════════════════════════════════════════════╗
  ║  NOTA: EVALUACIÓN DE QUANTILE MAPPING (ISIMIP3)                            ║
  ╠════════════════════════════════════════════════════════════════════════════╣
  ║                                                                            ║
  ║  ISIMIP3 corrige:                                                          ║
  ║   ✅ Sesgo de distribución completa (todos los cuantiles)                 ║
  ║   ✅ Frecuencia de días secos (umbral ajustado por cuantil)               ║
  ║   ✅ Intensidad de días húmedos (ajuste Gamma paramétrico)                ║
  ║   ✅ Extremos (GPD — solo modo isimip3+)                                  ║
  ║   ✅ Tendencia futura (delta scaling multiplicativo)                       ║
  ║                                                                            ║
  ║  ISIMIP3 NO preserva (por diseño):                                         ║
  ║   ❌ Sincronía temporal (NSE/KGE son solo informativos)                   ║
  ║                                                                            ║
  ║  MÉTRICAS CLAVE para evaluar ISIMIP3:                                      ║
  ║   1. PBIAS (%)       → sesgo de volumen  (0% = perfecto)                   ║
  ║   2. Error Media (%) → error en media    (<3% = excelente)                 ║
  ║   3. Error StdDev(%) → error variabilidad (<5% = excelente)                ║
  ║   4. Error Q95 (%)   → error en extremos  (<10% = excelente) ← NUEVO      ║
  ║   5. Error Q99 (%)   → error en extremos  (<15% = excelente) ← NUEVO      ║
  ╚════════════════════════════════════════════════════════════════════════════╝
""")

titulo_consola('  ESCALA MENSUAL', nivel=2)
m_sin_m  = calcular_metricas(obs_men, gcm_cal_m,      'Sin corrección')
m_emp_m  = calcular_metricas(obs_men, gcm_cal_m_emp,  'Empírico (ISIMIP2b)')
m_is3_m  = calcular_metricas(obs_men, gcm_cal_m_is3,  'ISIMIP3 (Gamma+delta)')
m_is3p_m = calcular_metricas(obs_men, gcm_cal_m_is3p, 'ISIMIP3+ (Gamma+GPD+delta)')

for m in [m_sin_m, m_emp_m, m_is3_m, m_is3p_m]:
    imprimir_metricas(m, 'Mensual')

titulo_consola('  ESCALA ANUAL', nivel=2)
m_sin_a  = calcular_metricas(obs_anu, gcm_cal_a,      'Sin corrección')
m_emp_a  = calcular_metricas(obs_anu, gcm_cal_a_emp,  'Empírico (ISIMIP2b)')
m_is3_a  = calcular_metricas(obs_anu, gcm_cal_a_is3,  'ISIMIP3 (Gamma+delta)')
m_is3p_a = calcular_metricas(obs_anu, gcm_cal_a_is3p, 'ISIMIP3+ (Gamma+GPD+delta)')

for m in [m_sin_a, m_emp_a, m_is3_a, m_is3p_a]:
    imprimir_metricas(m, 'Anual')

# ─── PASO 6: Conclusión automática ───────────────────────────────────────────
titulo_consola('PASO 6 — CONCLUSIÓN Y COMPARACIÓN DE MÉTODOS')

print('\n  EVALUACIÓN POR ESCALA Y MÉTODO:')
print('  ' + '─' * 76)

resultados_clasificacion = {}
for escala, metricas in [
    ('Mensual', [m_sin_m, m_emp_m, m_is3_m, m_is3p_m]),
    ('Anual',   [m_sin_a, m_emp_a, m_is3_a, m_is3p_a]),
]:
    print(f'\n  ESCALA {escala.upper()}:')
    for m in metricas:
        if m is None:
            continue
        icon, nivel, desc = clasificar_metrica_qm(m)
        print(f'    {icon} {m["Nombre"]:<35} → {nivel}')
        print(f'       PBIAS:{m["PBIAS (%)"]:+.1f}% | '
              f'ErrMedia:{m["Error Media (%)"]:.1f}% | '
              f'ErrStd:{m["Error StdDev (%)"]:.1f}% | '
              f'ErrQ95:{m["Error Q95 (%)"]:.1f}% | '
              f'ErrQ99:{m["Error Q99 (%)"]:.1f}%')
        resultados_clasificacion[f'{escala}_{m["Nombre"]}'] = (icon, nivel, desc)

# Recomendación global basada en ISIMIP3
icon_best, nivel_best, desc_best = clasificar_metrica_qm(m_is3_a)
escala_ref = 'ANUAL'
if m_is3_m and abs(m_is3_m['PBIAS (%)']) < abs(m_is3_a['PBIAS (%)']):
    icon_best, nivel_best, desc_best = clasificar_metrica_qm(m_is3_m)
    escala_ref = 'MENSUAL'

print(f'\n  ' + '─' * 76)
print(f'\n  ╔══════════════════════════════════════════════════════════════════════╗')
print(f'  ║  RECOMENDACIÓN FINAL (ISIMIP3): {icon_best} {nivel_best:<40}║')
print(f'  ║  Escala prioritaria: {escala_ref:<50}║')
print(f'  ╚══════════════════════════════════════════════════════════════════════╝')
print(f'\n  {desc_best}')

print(f"""
  ╔════════════════════════════════════════════════════════════════════════════╗
  ║  CRITERIOS ISIMIP3 (incluyen evaluación de extremos)                       ║
  ╠════════════════════════════════════════════════════════════════════════════╣
  │  ✅✅✅ EXCELENTE  │ PBIAS<5%  ErrMedia<3%  ErrStd<3%  ErrQ95<10%         │
  │  ✅✅  MUY BUENO  │ PBIAS<10% ErrMedia<5%  ErrStd<5%  ErrQ95<20%         │
  │  ✅  BUENO      │ PBIAS<15% ErrMedia<10% ErrStd<10% ErrQ95<30%         │
  │  ⚠️  ACEPTABLE   │ PBIAS<25% ErrMedia<15% ErrStd<15%                     │
  │  ❌  DEFICIENTE  │ Fuera de los rangos anteriores                        │
  ╚════════════════════════════════════════════════════════════════════════════╝
""")

# ─── PASO 7: Gráficos PNG ────────────────────────────────────────────────────
titulo_consola('PASO 7 — Generando gráficos PNG', nivel=2)
pngs = []

for escala, obs_s, gcm_cal_s, emp_s, is3_s, is3p_s, sfx_label in [
    ('mensual', obs_men, gcm_cal_m, gcm_cal_m_emp, gcm_cal_m_is3, gcm_cal_m_is3p, 'Mensual'),
    ('anual',   obs_anu, gcm_cal_a, gcm_cal_a_emp, gcm_cal_a_is3, gcm_cal_a_is3p, 'Anual'),
]:
    print(f'\n  [{sfx_label.upper()}]')

    series_fdc = {
        'OBS'                      : obs_s,
        'GCM sin corrección'       : gcm_cal_s,
        'Empírico (ISIMIP2b)'      : emp_s,
        'ISIMIP3 (Gamma+delta)'    : is3_s,
        'ISIMIP3+ (Gamma+GPD+delta)': is3p_s,
    }

    p = plot_fdc(series_fdc,
                 os.path.join(PNG_DIR, f'{MODELO}_FDC_{escala}.png'),
                 f'FDC {sfx_label} — Calibración — {MODELO}')
    pngs.append({'Gráfico': f'FDC {sfx_label}', 'Ruta': p})

    p = plot_hist(series_fdc,
                  os.path.join(PNG_DIR, f'{MODELO}_Hist_{escala}.png'),
                  f'Histograma {sfx_label} — Calibración — {MODELO}')
    pngs.append({'Gráfico': f'Histograma {sfx_label}', 'Ruta': p})

    sims_qq = {k: v for k, v in series_fdc.items() if k != 'OBS'}
    p = plot_qq(obs_s, sims_qq,
                os.path.join(PNG_DIR, f'{MODELO}_QQ_{escala}.png'),
                f'QQ-plot {sfx_label} — Calibración — {MODELO}')
    pngs.append({'Gráfico': f'QQ-plot {sfx_label}', 'Ruta': p})

    # Gráfico de extremos (nuevo en ISIMIP3)
    p = plot_extremos(obs_s, sims_qq,
                      os.path.join(PNG_DIR, f'{MODELO}_Extremos_{escala}.png'),
                      f'Extremos (P75–P99.9) {sfx_label} — {MODELO}')
    pngs.append({'Gráfico': f'Extremos {sfx_label}', 'Ruta': p})

# ─── PASO 8: Exportar Excel ──────────────────────────────────────────────────
titulo_consola('PASO 8 — Exportando resultados a Excel (.xlsx)', nivel=2)

wb = Workbook()
wb.remove(wb.active)

# ── Calibración mensual ──
ws = wb.create_sheet('calibracion_mensual')
df_cal_m = pd.DataFrame({
    'obs_mm_mes'              : obs_men,
    'gcm_sin_corr_mm_mes'     : gcm_cal_m,
    'gcm_empirico_mm_mes'     : gcm_cal_m_emp,
    'gcm_isimip3_mm_mes'      : gcm_cal_m_is3,
    'gcm_isimip3plus_mm_mes'  : gcm_cal_m_is3p,
})
df_to_sheet(ws, df_cal_m)
print('  Hoja "calibracion_mensual" ......... OK')

# ── Futuro mensual ──
ws = wb.create_sheet('futuro_mensual')
df_fut_m = pd.DataFrame({
    'gcm_sin_corr_mm_mes'     : gcm_fut_m,
    'gcm_empirico_mm_mes'     : gcm_fut_m_emp,
    'gcm_isimip3_mm_mes'      : gcm_fut_m_is3,
    'gcm_isimip3plus_mm_mes'  : gcm_fut_m_is3p,
})
df_to_sheet(ws, df_fut_m)
print('  Hoja "futuro_mensual" .............. OK')

# ── Calibración anual ──
ws = wb.create_sheet('calibracion_anual')
df_cal_a = pd.DataFrame({
    'obs_mm_año'              : obs_anu,
    'gcm_sin_corr_mm_año'     : gcm_cal_a,
    'gcm_empirico_mm_año'     : gcm_cal_a_emp,
    'gcm_isimip3_mm_año'      : gcm_cal_a_is3,
    'gcm_isimip3plus_mm_año'  : gcm_cal_a_is3p,
})
df_to_sheet(ws, df_cal_a)
print('  Hoja "calibracion_anual" ........... OK')

# ── Futuro anual ──
ws = wb.create_sheet('futuro_anual')
df_fut_a = pd.DataFrame({
    'gcm_sin_corr_mm_año'     : gcm_fut_a,
    'gcm_empirico_mm_año'     : gcm_fut_a_emp,
    'gcm_isimip3_mm_año'      : gcm_fut_a_is3,
    'gcm_isimip3plus_mm_año'  : gcm_fut_a_is3p,
})
df_to_sheet(ws, df_fut_a)
print('  Hoja "futuro_anual" ................ OK')

# ── Métricas ──
ws = wb.create_sheet('metricas')
filas_met = []
for escala, lista in [
    ('Mensual', [m_sin_m, m_emp_m, m_is3_m, m_is3p_m]),
    ('Anual',   [m_sin_a, m_emp_a, m_is3_a, m_is3p_a]),
]:
    for m in lista:
        if m:
            row = {'Escala': escala}
            row.update(m)
            filas_met.append(row)
df_met = pd.DataFrame(filas_met)
df_to_sheet(ws, df_met, header_fill='1a6e9e')
print('  Hoja "metricas" .................... OK')

# ── Gráficos ──
ws = wb.create_sheet('graficos')
df_pngs = pd.DataFrame(pngs)
df_to_sheet(ws, df_pngs, header_fill='2ca02c')
print('  Hoja "graficos" .................... OK')

# ── Resumen ──
ws = wb.create_sheet('resumen')
ws.column_dimensions['A'].width = 100

lineas_resumen = [
    '═' * 95,
    f'RESUMEN EJECUTIVO — ISIMIP3 Bias Correction | Modelo: {MODELO}',
    '═' * 95,
    '',
    'PERÍODO DE ANÁLISIS:',
    f'  Calibración : {CAL_INI} → {CAL_FIN}',
    f'  Futuro      : {FUT_INI} → {FUT_FIN}',
    '',
    'MÉTODOS IMPLEMENTADOS:',
    '  1. Empírico (ISIMIP2b)          — Mapeo cuantil empírico, referencia',
    '  2. ISIMIP3  (Gamma + delta)     — Días secos por cuantil + Gamma + preservación de tendencia',
    f'  3. ISIMIP3+ (Gamma + GPD + delta) — Como ISIMIP3 + corrección GPD para P>{P_EXTREMO*100:.0f}%',
    '',
    'PARÁMETROS UTILIZADOS:',
    f'  Umbral día seco (obs)  : {UMBRAL_MM} mm',
    f'  Percentil extremos GPD : {P_EXTREMO*100:.0f}%',
    f'  Cuantiles              : {N_QUANTILES}',
    '',
    f'DIAGNÓSTICO FINAL (ISIMIP3 — escala {escala_ref}):',
    f'  {icon_best} {nivel_best}',
    f'  {desc_best}',
    '',
    'MÉTRICAS CLAVE — ISIMIP3 (Gamma+delta):',
    '',
    '  ESCALA MENSUAL:',
    f'    PBIAS         : {m_is3_m["PBIAS (%)"]:.2f}%' if m_is3_m else '    PBIAS: N/A',
    f'    Error Media   : {m_is3_m["Error Media (%)"]:.2f}%' if m_is3_m else '',
    f'    Error StdDev  : {m_is3_m["Error StdDev (%)"]:.2f}%' if m_is3_m else '',
    f'    Error Q95     : {m_is3_m["Error Q95 (%)"]:.2f}%' if m_is3_m else '',
    f'    Error Q99     : {m_is3_m["Error Q99 (%)"]:.2f}%' if m_is3_m else '',
    '',
    '  ESCALA ANUAL:',
    f'    PBIAS         : {m_is3_a["PBIAS (%)"]:.2f}%' if m_is3_a else '    PBIAS: N/A',
    f'    Error Media   : {m_is3_a["Error Media (%)"]:.2f}%' if m_is3_a else '',
    f'    Error StdDev  : {m_is3_a["Error StdDev (%)"]:.2f}%' if m_is3_a else '',
    f'    Error Q95     : {m_is3_a["Error Q95 (%)"]:.2f}%' if m_is3_a else '',
    f'    Error Q99     : {m_is3_a["Error Q99 (%)"]:.2f}%' if m_is3_a else '',
    '',
    'CRITERIOS DE CLASIFICACIÓN ISIMIP3:',
    '  ✅✅✅ EXCELENTE : PBIAS<5%  ErrMedia<3%  ErrStd<3%  ErrQ95<10%',
    '  ✅✅  MUY BUENO : PBIAS<10% ErrMedia<5%  ErrStd<5%  ErrQ95<20%',
    '  ✅  BUENO     : PBIAS<15% ErrMedia<10% ErrStd<10% ErrQ95<30%',
    '  ⚠️  ACEPTABLE  : PBIAS<25% ErrMedia<15% ErrStd<15%',
    '  ❌  DEFICIENTE : Fuera de los rangos anteriores',
    '',
    'RECOMENDACIONES ISIMIP3:',
    '  • Para proyecciones de precipitación media → usar ISIMIP3 (Gamma+delta)',
    '  • Para análisis de extremos (sequías, inundaciones) → usar ISIMIP3+',
    '  • Siempre validar con FDC, QQ-plot e histograma',
    '  • Comparar gráfico de extremos (P75–P99.9) para elegir modo óptimo',
    '',
    'ARCHIVOS GENERADOS:',
] + [f'  {p["Gráfico"]} → {os.path.basename(p["Ruta"])}' for p in pngs]

for i, linea in enumerate(lineas_resumen, 1):
    c = ws.cell(row=i, column=1, value=linea)
    if i in [1, 2, 3]:
        c.font = Font(bold=True, size=11, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1a6e9e')
    elif any(s in linea for s in ['PERÍODO', 'MÉTODOS', 'PARÁMETROS',
                                   'DIAGNÓSTICO', 'MÉTRICAS', 'CRITERIOS',
                                   'RECOMENDACIONES', 'ARCHIVOS']):
        c.font = Font(bold=True, size=10)
    else:
        c.font = Font(size=10)

wb.save(SALIDA_XLSX)
print(f'\n  ✔ Excel guardado → {SALIDA_XLSX}')

# ─── FIN ─────────────────────────────────────────────────────────────────────
titulo_consola('PROCESO COMPLETADO EXITOSAMENTE')
print(f"""
  ╔═══════════════════════════════════════════════════════════════════════════╗
  ║  ARCHIVOS GENERADOS                                                       ║
  ╠═══════════════════════════════════════════════════════════════════════════╣
  │  Excel  → {SALIDA_XLSX}
  │  PNGs   → {PNG_DIR}
  │    FDC, Histograma, QQ-plot, Extremos  ×  Mensual/Anual
  ╠═══════════════════════════════════════════════════════════════════════════╣
  ║  NUEVAS HOJAS EN EXCEL                                                    ║
  │    calibracion_mensual / futuro_mensual → incluye columnas ISIMIP3/3+     ║
  │    calibracion_anual   / futuro_anual   → incluye columnas ISIMIP3/3+     ║
  │    metricas → Error Q95 y Q99 como métricas adicionales                   ║
  │    graficos / resumen                                                      ║
  ╠═══════════════════════════════════════════════════════════════════════════╣
  ║  PRÓXIMOS PASOS                                                            ║
  │  1. Comparar gráfico Extremos: ¿ISIMIP3 o ISIMIP3+ ajusta mejor la cola? ║
  │  2. Si Error Q95/Q99 < 15% con ISIMIP3 → no es necesario usar ISIMIP3+   ║
  │  3. Usar columnas 'gcm_isimip3_*' o 'gcm_isimip3plus_*' según resultado  ║
  ╚═══════════════════════════════════════════════════════════════════════════╝
""")
