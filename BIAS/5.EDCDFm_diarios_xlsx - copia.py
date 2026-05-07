# =============================================================================
#  EDCDFm — Equidistant CDF Matching (Li et al., 2010)
#  Corrección de sesgo para PRECIPITACIÓN DIARIA
#
#  Referencia: Li, H., Sheffield, J., & Wood, E. F. (2010).
#  "Bias correction of monthly precipitation and temperature fields from
#   Intergovernmental Panel on Climate Change AR4 models using equidistant
#   quantile matching." Journal of Geophysical Research, 115, D10101.
#
#  ¿Por qué EDCDFm sobre ISIMIP para precipitación diaria?
#  ─────────────────────────────────────────────────────────
#  • ISIMIP aplica una corrección ADITIVA (delta = q_obs - q_gcm_cal), lo que
#    puede producir valores negativos o distorsiones en distribuciones asimétricas
#    como la precipitación.
#  • EDCDFm usa corrección MULTIPLICATIVA relativa: escala el cuantil futuro por
#    la razón q_obs / q_gcm_cal (en vez de la diferencia), preservando la
#    no-negatividad y las colas pesadas de lluvia extrema.
#  • Mejor desempeño en: extremos (P95, Rx1day, Rx5day), días húmedos/secos,
#    e intensidad de eventos en regiones con régimen variable.
#
#  Métodos implementados y comparados:
#  ────────────────────────────────────
#  1. EDCDFm Empírico     — corrección multiplicativa por cuantiles empíricos
#  2. EDCDFm Gamma        — corrección vía ajuste paramétrico Gamma (días húmedos)
#  3. LS (Linear Scaling) — escala lineal por razón de medias (referencia simple)
#  4. LOCI                — Linear scaling con corrección de umbral de ocurrencia
#
#  Salidas:
#  ────────
#  • Consola  : métricas detalladas + diagnóstico de calidad por método
#  • PNGs     : FDC, Histograma días húmedos, QQ-plot, índices de extremos
#  • Excel    : calibración diaria, futuro diario, métricas, extremos, resumen
# =============================================================================

import os
import warnings
import numpy as np
import pandas as pd
from scipy.stats import gamma as gamma_dist, pearsonr, ks_2samp
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')

plt.rcParams.update({
    'font.family'       : 'DejaVu Sans',
    'axes.unicode_minus': False,
    'figure.dpi'        : 150,
})

# =============================================================================
#  CONFIGURACIÓN USUARIO  ← EDITAR AQUÍ
# =============================================================================
OBS_XLSX    = r'C:/1.PYTHON/Descarga_Python/Pd_Historica_SMN.xlsx'
GCM_XLSX    = r'C:/1.PYTHON/Descarga_Python/Tabla_EC-Earth3_historico_futuro.xlsx'
COL_OBS     = 'prec'
COL_GCM     = 'pr_mm'
MODELO      = 'EC-Earth3'   # aqui colocar el nombre del modelo
CAL_INI     = '1961-01-01'
CAL_FIN     = '2024-12-31'
FUT_INI     = '2030-01-01'
FUT_FIN     = '2100-12-31'
OUT_DIR     = rf'C:/1.PYTHON/Descarga_Python/{MODELO}_salidas_EDCDFm'
N_QUANTILES = 100
UMBRAL_MM   = 0.1       # mm/día — umbral día húmedo
# =============================================================================

os.makedirs(OUT_DIR, exist_ok=True)
PNG_DIR = os.path.join(OUT_DIR, 'graficos')
os.makedirs(PNG_DIR, exist_ok=True)
SALIDA_XLSX = os.path.join(OUT_DIR, f'{MODELO}_EDCDFm_diario.xlsx')

SEP  = '=' * 78
SEP2 = '-' * 78

COLORES = {
    'OBS' : '#1a6e9e',
    'SIN' : '#d94f3d',
    'EMP' : '#2ca02c',
    'GAM' : '#ff7f0e',
    'LS'  : '#9467bd',
    'LOCI': '#8c564b',
}


# ─────────────────────────────────────────────────────────────────────────────
#  UTILIDADES CONSOLA
# ─────────────────────────────────────────────────────────────────────────────

def titulo(texto, nivel=1):
    if nivel == 1:
        print(f'\n{SEP}\n  {texto}\n{SEP}')
    else:
        print(f'\n{SEP2}\n  {texto}\n{SEP2}')


# ─────────────────────────────────────────────────────────────────────────────
#  MÉTODOS DE CORRECCIÓN
# ─────────────────────────────────────────────────────────────────────────────

def edcdfm_empirico(obs_cal, gcm_cal, gcm_fut, n_quantiles=100):
    """
    EDCDFm Empírico — Li et al. (2010).

    Corrección multiplicativa por cuantiles:
        P_corr(t) = P_fut(t) * [ F_obs^-1(p) / F_gcm_cal^-1(p) ]
    donde p = F_gcm_fut( P_fut(t) ).

    Preserva la señal de cambio del modelo y corrige la distribución.
    Evita valores negativos por ser multiplicativa.
    """
    obs_cal       = obs_cal.dropna()
    gcm_cal       = gcm_cal.dropna()
    gcm_fut_clean = gcm_fut.dropna()

    quantiles = np.linspace(0, 1, n_quantiles + 1)
    q_obs     = np.quantile(obs_cal,       quantiles)
    q_cal     = np.quantile(gcm_cal,       quantiles)
    q_fut     = np.quantile(gcm_fut_clean, quantiles)

    # Razón de corrección: obs/cal en cuantiles
    with np.errstate(divide='ignore', invalid='ignore'):
        ratio = np.where(q_cal > 0, q_obs / q_cal, 1.0)

    # Para cada valor futuro: encontrar su cuantil en fut, aplicar ratio
    p_fut     = np.interp(gcm_fut_clean.values, q_fut, quantiles,
                          left=0.0, right=1.0)
    ratio_val = np.interp(p_fut, quantiles, ratio)
    corr      = gcm_fut_clean.values * ratio_val

    return pd.Series(np.clip(corr, 0, None), index=gcm_fut_clean.index)


def edcdfm_gamma(obs_cal, gcm_cal, gcm_fut, umbral=0.1):
    """
    EDCDFm Gamma — corrección paramétrica para días húmedos.

    Ajusta distribuciones Gamma a días húmedos de OBS, GCM_cal y GCM_fut.
    Aplica corrección multiplicativa de escala via parámetros Gamma.
    Días secos (≤ umbral) se mapean a 0.

    Ventaja: más robusto que el empírico en regímenes áridos o semi-áridos
    con pocos días de lluvia.
    """
    obs_cal       = obs_cal.dropna()
    gcm_cal       = gcm_cal.dropna()
    gcm_fut_clean = gcm_fut.dropna()

    obs_wet     = obs_cal[obs_cal > umbral]
    gcm_cal_wet = gcm_cal[gcm_cal > umbral]
    gcm_fut_wet = gcm_fut_clean[gcm_fut_clean > umbral]

    # Fallback a empírico si pocos datos
    if len(obs_wet) < 10 or len(gcm_cal_wet) < 10 or len(gcm_fut_wet) < 10:
        return edcdfm_empirico(obs_cal, gcm_cal, gcm_fut)

    a_obs, _, sc_obs = gamma_dist.fit(obs_wet,     floc=0)
    a_cal, _, sc_cal = gamma_dist.fit(gcm_cal_wet, floc=0)
    a_fut, _, sc_fut = gamma_dist.fit(gcm_fut_wet, floc=0)

    # Razón multiplicativa de escala (EDCDFm: multiplicativa, no aditiva)
    with np.errstate(divide='ignore', invalid='ignore'):
        ratio_scale = sc_obs / sc_cal if sc_cal > 0 else 1.0

    sc_corr = max(sc_fut * ratio_scale, 1e-6)
    a_corr  = a_obs   # preservar forma observada

    # Mapeo por CDF: P_corr = F_obs_gamma^-1 ( F_fut_gamma(P_fut) )
    p_wet   = gamma_dist.cdf(gcm_fut_wet.values, a_fut, scale=sc_fut)
    corr_w  = gamma_dist.ppf(np.clip(p_wet, 1e-6, 1 - 1e-6),
                              a_corr, scale=sc_corr)

    result                                    = gcm_fut_clean.copy().astype(float)
    result[gcm_fut_clean > umbral]            = corr_w
    result[gcm_fut_clean <= umbral]           = 0.0

    return pd.Series(np.clip(result.values, 0, None), index=gcm_fut_clean.index)


def linear_scaling(obs_cal, gcm_cal, gcm_fut):
    """
    Linear Scaling (LS) — método de referencia simple.

    P_corr(t) = P_fut(t) * [ mean(OBS_cal) / mean(GCM_cal) ]

    Corrección global multiplicativa. Corrige la media pero no la distribución.
    Útil como línea base para comparar métodos más sofisticados.
    """
    obs_cal       = obs_cal.dropna()
    gcm_cal       = gcm_cal.dropna()
    gcm_fut_clean = gcm_fut.dropna()

    mu_obs = obs_cal.mean()
    mu_cal = gcm_cal.mean()

    with np.errstate(divide='ignore', invalid='ignore'):
        factor = mu_obs / mu_cal if mu_cal > 0 else 1.0

    corr = gcm_fut_clean.values * factor
    return pd.Series(np.clip(corr, 0, None), index=gcm_fut_clean.index)


def loci(obs_cal, gcm_cal, gcm_fut, umbral=0.1):
    """
    LOCI — Local Intensity Scaling (Schmidli et al., 2006).

    Pasos:
      1. Ajusta el umbral del GCM para igualar la fracción de días húmedos de OBS.
      2. Escala la intensidad de días húmedos con factor multiplicativo.

    Ventaja sobre LS: corrige la frecuencia de ocurrencia de lluvia,
    no solo la intensidad.
    """
    obs_cal       = obs_cal.dropna()
    gcm_cal       = gcm_cal.dropna()
    gcm_fut_clean = gcm_fut.dropna()

    # 1. Fracción días húmedos observada
    frac_obs  = (obs_cal > umbral).mean()

    # 2. Umbral del GCM que reproduce esa frecuencia
    umbral_gcm = np.quantile(gcm_cal, 1 - frac_obs) if frac_obs > 0 else umbral

    # 3. Medias de días húmedos
    mu_obs_wet = obs_cal[obs_cal > umbral].mean()
    mu_cal_wet = gcm_cal[gcm_cal > umbral_gcm].mean()

    with np.errstate(divide='ignore', invalid='ignore'):
        factor = mu_obs_wet / mu_cal_wet if mu_cal_wet > 0 else 1.0

    # 4. Aplicar al futuro
    # Umbral futuro: mantener la misma fracción húmeda del GCM futuro
    frac_fut    = (gcm_cal > umbral_gcm).mean()
    umbral_fut  = np.quantile(gcm_fut_clean, 1 - frac_fut) if frac_fut > 0 else umbral_gcm

    corr = gcm_fut_clean.copy().astype(float)
    mask_wet              = gcm_fut_clean > umbral_fut
    corr[mask_wet]        = gcm_fut_clean[mask_wet] * factor
    corr[~mask_wet]       = 0.0

    return pd.Series(np.clip(corr.values, 0, None), index=gcm_fut_clean.index)


# ─────────────────────────────────────────────────────────────────────────────
#  MÉTRICAS DE AJUSTE
# ─────────────────────────────────────────────────────────────────────────────

def calcular_metricas(obs, sim, nombre=''):
    """
    Métricas completas para evaluación de corrección de sesgo en precipitación diaria.
    Incluye métricas de distribución, índices de extremos y test estadístico KS.
    """
    df = pd.DataFrame({'obs': obs, 'sim': sim}).dropna()
    o  = df['obs'].values
    s  = df['sim'].values
    n  = len(o)

    if n < 10:
        return None

    # ── Métricas generales ──
    r, _     = pearsonr(o, s)
    r2       = r ** 2
    nse      = 1 - np.sum((o - s)**2) / np.sum((o - np.mean(o))**2)
    alpha    = np.std(s)  / np.std(o)  if np.std(o)  > 0 else np.nan
    beta     = np.mean(s) / np.mean(o) if np.mean(o) > 0 else np.nan
    kge      = 1 - np.sqrt((r-1)**2 + (alpha-1)**2 + (beta-1)**2) \
               if not (np.isnan(alpha) or np.isnan(beta)) else np.nan
    pbias    = (np.sum(s - o) / np.sum(o)) * 100 if np.sum(o) > 0 else 0.0
    rmse     = np.sqrt(np.mean((o - s)**2))
    mae      = np.mean(np.abs(o - s))

    mu_o     = np.mean(o)
    mu_s     = np.mean(s)
    std_o    = np.std(o)
    std_s    = np.std(s)
    err_mu   = abs(mu_o - mu_s)   / mu_o   * 100 if mu_o   > 0 else 0.0
    err_std  = abs(std_o - std_s) / std_o  * 100 if std_o  > 0 else 0.0

    # ── Ocurrencia ──
    frac_sec_o = np.mean(o <= UMBRAL_MM) * 100
    frac_sec_s = np.mean(s <= UMBRAL_MM) * 100
    err_fsec   = abs(frac_sec_o - frac_sec_s)

    # ── Intensidad días húmedos ──
    wet_o    = o[o > UMBRAL_MM]
    wet_s    = s[s > UMBRAL_MM]
    sdii_o   = np.mean(wet_o) if len(wet_o) > 0 else 0.0
    sdii_s   = np.mean(wet_s) if len(wet_s) > 0 else 0.0
    err_sdii = abs(sdii_o - sdii_s) / sdii_o * 100 if sdii_o > 0 else 0.0

    # ── Índices de extremos ──
    p95_o    = np.percentile(o[o > UMBRAL_MM], 95) if len(wet_o) > 0 else 0.0
    p95_s    = np.percentile(s[s > UMBRAL_MM], 95) if len(wet_s) > 0 else 0.0
    p99_o    = np.percentile(o[o > UMBRAL_MM], 99) if len(wet_o) > 0 else 0.0
    p99_s    = np.percentile(s[s > UMBRAL_MM], 99) if len(wet_s) > 0 else 0.0
    rx1_o    = np.max(o)
    rx1_s    = np.max(s)
    err_p95  = abs(p95_o - p95_s) / p95_o * 100 if p95_o > 0 else 0.0
    err_p99  = abs(p99_o - p99_s) / p99_o * 100 if p99_o > 0 else 0.0
    err_rx1  = abs(rx1_o - rx1_s) / rx1_o * 100 if rx1_o > 0 else 0.0

    # ── Test Kolmogorov-Smirnov ──
    ks_stat, ks_pval = ks_2samp(wet_o, wet_s) if len(wet_o) > 0 and len(wet_s) > 0 else (np.nan, np.nan)

    return {
        'Nombre'            : nombre,
        'n'                 : n,
        'NSE'               : round(nse,  4),
        'KGE'               : round(float(kge),   4) if not np.isnan(kge)   else None,
        'r (Pearson)'       : round(r,    4),
        'R²'                : round(r2,   4),
        'PBIAS (%)'         : round(pbias, 2),
        'RMSE (mm/día)'     : round(rmse,  4),
        'MAE (mm/día)'      : round(mae,   4),
        'Media OBS'         : round(mu_o,  4),
        'Media SIM'         : round(mu_s,  4),
        'Error Media (%)'   : round(err_mu,   2),
        'StdDev OBS'        : round(std_o,    4),
        'StdDev SIM'        : round(std_s,    4),
        'Error StdDev (%)'  : round(err_std,  2),
        'Días secos OBS (%)': round(frac_sec_o, 2),
        'Días secos SIM (%)': round(frac_sec_s, 2),
        'Error Frec. seca (pp)': round(err_fsec, 2),
        'SDII OBS (mm/día húmedo)': round(sdii_o, 4),
        'SDII SIM (mm/día húmedo)': round(sdii_s, 4),
        'Error SDII (%)'    : round(err_sdii, 2),
        'P95 OBS (mm)'      : round(p95_o, 4),
        'P95 SIM (mm)'      : round(p95_s, 4),
        'Error P95 (%)'     : round(err_p95, 2),
        'P99 OBS (mm)'      : round(p99_o, 4),
        'P99 SIM (mm)'      : round(p99_s, 4),
        'Error P99 (%)'     : round(err_p99, 2),
        'Rx1day OBS (mm)'   : round(rx1_o, 4),
        'Rx1day SIM (mm)'   : round(rx1_s, 4),
        'Error Rx1day (%)'  : round(err_rx1, 2),
        'KS estadístico'    : round(float(ks_stat), 4) if not np.isnan(ks_stat) else None,
        'KS p-valor'        : round(float(ks_pval), 4) if not np.isnan(ks_pval) else None,
    }


def imprimir_metricas(m):
    if m is None:
        print('  !! Insuficientes datos.')
        return
    ks_p  = m['KS p-valor']
    ks_ok = '✅ (distribuciones similares)' if (ks_p is not None and ks_p > 0.05) \
            else '⚠️ (distribuciones distintas)'

    print(f'\n  ┌─ {m["Nombre"]} {"─"*(54 - len(m["Nombre"]))}┐')
    print(f'  │  n = {m["n"]:,} días')
    print(f'  │')
    print(f'  │  MÉTRICAS GLOBALES (informativas para QM):')
    print(f'  │    NSE           : {m["NSE"]:>8.4f}')
    print(f'  │    KGE           : {str(m["KGE"]):>8}')
    print(f'  │    r (Pearson)   : {m["r (Pearson)"]:>8.4f}')
    print(f'  │    R²            : {m["R²"]:>8.4f}')
    print(f'  │')
    print(f'  │  ✅ MÉTRICAS CLAVE DE DISTRIBUCIÓN:')
    print(f'  │    PBIAS (%)           : {m["PBIAS (%)"]:>8.2f}%    ← 0% = perfecto')
    print(f'  │    Error Media (%)     : {m["Error Media (%)"]:>8.2f}%')
    print(f'  │    Error StdDev (%)    : {m["Error StdDev (%)"]:>8.2f}%')
    print(f'  │    RMSE (mm/día)       : {m["RMSE (mm/día)"]:>8.4f}')
    print(f'  │    MAE  (mm/día)       : {m["MAE (mm/día)"]:>8.4f}')
    print(f'  │')
    print(f'  │  🌧️ OCURRENCIA E INTENSIDAD:')
    print(f'  │    Días secos OBS/SIM  : {m["Días secos OBS (%)"]:.2f}% / {m["Días secos SIM (%)"]:.2f}%   (error: {m["Error Frec. seca (pp)"]:.2f} pp)')
    print(f'  │    SDII OBS/SIM        : {m["SDII OBS (mm/día húmedo)"]:.3f} / {m["SDII SIM (mm/día húmedo)"]:.3f} mm/día  (error: {m["Error SDII (%)"]:.2f}%)')
    print(f'  │')
    print(f'  │  ⚡ EXTREMOS:')
    print(f'  │    P95  OBS/SIM        : {m["P95 OBS (mm)"]:.3f} / {m["P95 SIM (mm)"]:.3f} mm   (error: {m["Error P95 (%)"]:.2f}%)')
    print(f'  │    P99  OBS/SIM        : {m["P99 OBS (mm)"]:.3f} / {m["P99 SIM (mm)"]:.3f} mm   (error: {m["Error P99 (%)"]:.2f}%)')
    print(f'  │    Rx1day OBS/SIM      : {m["Rx1day OBS (mm)"]:.3f} / {m["Rx1day SIM (mm)"]:.3f} mm   (error: {m["Error Rx1day (%)"]:.2f}%)')
    print(f'  │')
    print(f'  │  📊 TEST KS (días húmedos):')
    print(f'  │    Estadístico KS      : {str(m["KS estadístico"]):>8}')
    print(f'  │    p-valor             : {str(ks_p):>8}    {ks_ok}')
    print(f'  └{"─"*60}┘')


def clasificar(m):
    if m is None:
        return '⚠️', 'SIN DATOS', 'Datos insuficientes.'
    pb   = abs(m['PBIAS (%)'])
    em   = m['Error Media (%)']
    es   = m['Error StdDev (%)']
    ep95 = m['Error P95 (%)']
    efsec = m['Error Frec. seca (pp)']

    score = 0
    score += (pb < 5)  + (pb < 10)  + (pb < 15)
    score += (em < 3)  + (em < 5)   + (em < 10)
    score += (es < 5)  + (es < 10)  + (es < 15)
    score += (ep95 < 10) + (ep95 < 20)
    score += (efsec < 2)  + (efsec < 5)

    if score >= 11:
        return '✅✅✅', 'EXCELENTE',  'Corrección excelente en distribución, extremos y ocurrencia.'
    elif score >= 8:
        return '✅✅',  'MUY BUENO',  'Corrección muy buena. Método RECOMENDADO para proyecciones.'
    elif score >= 5:
        return '✅',   'BUENO',      'Corrección buena. Valida extremos con FDC/QQ-plot.'
    elif score >= 3:
        return '⚠️',  'ACEPTABLE',  'Corrección moderada. Considera validación estacional adicional.'
    else:
        return '❌',  'DEFICIENTE', 'Corrección insatisfactoria. Revisa datos o parámetros.'


# ─────────────────────────────────────────────────────────────────────────────
#  GRÁFICOS
# ─────────────────────────────────────────────────────────────────────────────

def _guardar(fig, ruta):
    try:
        fig.tight_layout()
    except Exception:
        pass
    fig.savefig(ruta, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f'     PNG → {ruta}')
    return ruta


def plot_fdc(series_dict, ruta, titulo_str):
    """FDC en escala log para destacar extremos."""
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))
    estilos = ['-', '--', '-.', ':', (0,(3,1,1,1))]
    cols    = [COLORES['OBS'], COLORES['SIN'], COLORES['EMP'],
               COLORES['GAM'], COLORES['LS'], COLORES['LOCI']]

    for ax, yscale in zip(axes, ['linear', 'log']):
        for i, (lbl, s) in enumerate(series_dict.items()):
            d    = s.dropna().sort_values(ascending=False).values
            n    = len(d)
            prob = np.arange(1, n + 1) / (n + 1) * 100
            ax.plot(prob, d, linestyle=estilos[i % 5],
                    color=cols[i % 6], linewidth=1.6, label=lbl)
        ax.set_xlabel('Probabilidad de excedencia (%)', fontsize=10)
        ax.set_ylabel('Precipitación (mm/día)', fontsize=10)
        ax.set_title(f'{titulo_str} — {"Escala lineal" if yscale == "linear" else "Escala log"}',
                     fontsize=11, fontweight='bold')
        ax.set_yscale(yscale)
        ax.legend(fontsize=8)
        ax.grid(True, linestyle='--', alpha=0.4)
        ax.xaxis.set_major_formatter(mticker.FormatStrFormatter('%g%%'))

    return _guardar(fig, ruta)


def plot_hist(series_dict, ruta, titulo_str):
    """Histograma de densidad para días húmedos."""
    fig, ax = plt.subplots(figsize=(11, 5))
    cols    = [COLORES['OBS'], COLORES['SIN'], COLORES['EMP'],
               COLORES['GAM'], COLORES['LS'], COLORES['LOCI']]
    for i, (lbl, s) in enumerate(series_dict.items()):
        data = s.dropna()
        data = data[data > UMBRAL_MM].values
        ax.hist(data, bins=40, density=True, alpha=0.45,
                color=cols[i % 6], label=lbl, edgecolor='white')
    ax.set_xlabel('Precipitación días húmedos (mm/día)', fontsize=11)
    ax.set_ylabel('Densidad', fontsize=11)
    ax.set_title(titulo_str, fontsize=12, fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(True, linestyle='--', alpha=0.4)
    return _guardar(fig, ruta)


def plot_qq(obs, sims_dict, ruta, titulo_str):
    """QQ-plot cuantiles OBS vs SIM (días húmedos)."""
    quantiles = np.linspace(0, 1, 101)
    obs_wet   = obs.dropna()
    obs_wet   = obs_wet[obs_wet > UMBRAL_MM]
    q_obs     = np.quantile(obs_wet, quantiles)

    cols = [COLORES['SIN'], COLORES['EMP'], COLORES['GAM'],
            COLORES['LS'],  COLORES['LOCI']]
    fig, ax = plt.subplots(figsize=(7, 7))
    lim_max = q_obs[-1]
    for i, (lbl, s) in enumerate(sims_dict.items()):
        wet   = s.dropna()
        wet   = wet[wet > UMBRAL_MM]
        q_sim = np.quantile(wet, quantiles)
        ax.scatter(q_obs, q_sim, s=18, alpha=0.7, color=cols[i % 5], label=lbl)
        lim_max = max(lim_max, q_sim[-1])
    ax.plot([0, lim_max], [0, lim_max], 'k--', linewidth=1.2, label='1:1 (perfecto)')
    ax.set_xlabel('Cuantiles OBS (mm/día)', fontsize=11)
    ax.set_ylabel('Cuantiles SIM (mm/día)', fontsize=11)
    ax.set_title(titulo_str, fontsize=12, fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(True, linestyle='--', alpha=0.4)
    return _guardar(fig, ruta)


def plot_extremos(metricas_list, metodos, ruta, titulo_str):
    """
    Panel de barras comparando índices de extremos entre métodos.
    P95, P99, Rx1day y SDII: valor OBS vs valor de cada método.
    """
    indices  = ['P95 SIM (mm)', 'P99 SIM (mm)', 'Rx1day SIM (mm)', 'SDII SIM (mm/día húmedo)']
    etiquetas= ['P95 (mm)', 'P99 (mm)', 'Rx1day (mm)', 'SDII (mm/día húm.)']
    refs     = ['P95 OBS (mm)', 'P99 OBS (mm)', 'Rx1day OBS (mm)', 'SDII OBS (mm/día húmedo)']

    n_idx = len(indices)
    fig, axes = plt.subplots(1, n_idx, figsize=(16, 5))
    cols = [COLORES['SIN'], COLORES['EMP'], COLORES['GAM'],
            COLORES['LS'],  COLORES['LOCI']]

    for j, (ax, idx, etq, ref) in enumerate(zip(axes, indices, etiquetas, refs)):
        obs_val = metricas_list[0][ref] if metricas_list[0] else 0
        ax.axhline(obs_val, color=COLORES['OBS'], linewidth=2.0,
                   linestyle='-', label='OBS', zorder=5)
        for i, (m, lbl) in enumerate(zip(metricas_list, metodos)):
            val = m[idx] if m else 0
            ax.bar(i, val, color=cols[i % 5], alpha=0.75, label=lbl, width=0.6)
        ax.set_title(etq, fontsize=10, fontweight='bold')
        ax.set_xticks(range(len(metodos)))
        ax.set_xticklabels(metodos, rotation=30, ha='right', fontsize=8)
        ax.set_ylabel('mm' if j < 3 else 'mm/día', fontsize=9)
        ax.grid(True, axis='y', linestyle='--', alpha=0.4)
        if j == 0:
            ax.legend(fontsize=8)

    fig.suptitle(titulo_str, fontsize=12, fontweight='bold')
    return _guardar(fig, ruta)


def plot_dias_secos(metricas_list, metodos, ruta, titulo_str):
    """Barras comparando fracción de días secos OBS vs métodos."""
    fig, ax = plt.subplots(figsize=(9, 5))
    obs_val = metricas_list[0]['Días secos OBS (%)'] if metricas_list[0] else 0
    ax.axhline(obs_val, color=COLORES['OBS'], linewidth=2.0,
               linestyle='-', label='OBS', zorder=5)
    cols = [COLORES['SIN'], COLORES['EMP'], COLORES['GAM'],
            COLORES['LS'],  COLORES['LOCI']]
    for i, (m, lbl) in enumerate(zip(metricas_list, metodos)):
        val = m['Días secos SIM (%)'] if m else 0
        ax.bar(i, val, color=cols[i % 5], alpha=0.75, width=0.6, label=lbl)
    ax.set_xticks(range(len(metodos)))
    ax.set_xticklabels(metodos, rotation=20, ha='right', fontsize=9)
    ax.set_ylabel('Fracción días secos (%)', fontsize=11)
    ax.set_title(titulo_str, fontsize=12, fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(True, axis='y', linestyle='--', alpha=0.4)
    return _guardar(fig, ruta)


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def df_to_sheet(ws, df, header_fill='1a6e9e', zebra=True):
    fill_h  = PatternFill('solid', fgColor=header_fill)
    fill_z  = PatternFill('solid', fgColor='EAF2FB')
    font_h  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    font_d  = Font(name='Arial', size=10)
    border  = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'),  bottom=Side(style='thin'))

    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if r_idx == 1:
                cell.fill = fill_h
                cell.font = font_h
            else:
                cell.font = font_d
                if zebra and r_idx % 2 == 0:
                    cell.fill = fill_z
        if r_idx == 1:
            ws.row_dimensions[1].height = 22

    for col in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 36)


# =============================================================================
#  EJECUCIÓN PRINCIPAL
# =============================================================================

titulo(f'EDCDFm BIAS CORRECTION — Precipitación DIARIA\n  Modelo: {MODELO}')

# ─── PASO 1: Lectura ──────────────────────────────────────────────────────────
titulo('PASO 1 — Lectura de archivos de entrada', nivel=2)

print(f'  OBS → {OBS_XLSX}')
obs_raw = pd.read_excel(OBS_XLSX, parse_dates=['fecha'], index_col='fecha')
print(f'  GCM → {GCM_XLSX}')
gcm_raw = pd.read_excel(GCM_XLSX, parse_dates=['fecha'], index_col='fecha')

obs_d = obs_raw[COL_OBS].sort_index()
gcm_d = gcm_raw[COL_GCM].sort_index()

obs_cal = obs_d[CAL_INI:CAL_FIN]
gcm_cal = gcm_d[CAL_INI:CAL_FIN]
gcm_fut = gcm_d[FUT_INI:FUT_FIN]

print(f'\n  Período calibración  : {CAL_INI} → {CAL_FIN}')
print(f'  Período futuro       : {FUT_INI} → {FUT_FIN}')
print(f'  OBS diarios          : {len(obs_cal):,} registros')
print(f'  GCM diarios          : {len(gcm_cal):,} (cal)  |  {len(gcm_fut):,} (fut)')
print(f'  Días húmedos OBS     : {(obs_cal > UMBRAL_MM).sum():,} ({(obs_cal > UMBRAL_MM).mean()*100:.1f}%)')
print(f'  Días húmedos GCM cal : {(gcm_cal > UMBRAL_MM).sum():,} ({(gcm_cal > UMBRAL_MM).mean()*100:.1f}%)')
print(f'  Umbral día húmedo    : {UMBRAL_MM} mm/día')

# ─── PASO 2: Corrección período FUTURO ───────────────────────────────────────
titulo('PASO 2 — Corrección al período FUTURO (4 métodos)', nivel=2)

print('  [EDCDFm Empírico]  ...', end=' ', flush=True)
fut_emp = edcdfm_empirico(obs_cal, gcm_cal, gcm_fut, N_QUANTILES)
print('OK')

print('  [EDCDFm Gamma]     ...', end=' ', flush=True)
fut_gam = edcdfm_gamma(obs_cal, gcm_cal, gcm_fut, UMBRAL_MM)
print('OK')

print('  [Linear Scaling]   ...', end=' ', flush=True)
fut_ls  = linear_scaling(obs_cal, gcm_cal, gcm_fut)
print('OK')

print('  [LOCI]             ...', end=' ', flush=True)
fut_loc = loci(obs_cal, gcm_cal, gcm_fut, UMBRAL_MM)
print('OK')

# ─── PASO 3: Retroproyección en calibración ──────────────────────────────────
titulo('PASO 3 — Retroproyección en CALIBRACIÓN (validación cruzada)', nivel=2)
print('  Se aplica corrección usando GCM_cal como "futuro" para comparar con OBS.')

print('  [EDCDFm Empírico]  ...', end=' ', flush=True)
cal_emp = edcdfm_empirico(obs_cal, gcm_cal, gcm_cal, N_QUANTILES)
print('OK')

print('  [EDCDFm Gamma]     ...', end=' ', flush=True)
cal_gam = edcdfm_gamma(obs_cal, gcm_cal, gcm_cal, UMBRAL_MM)
print('OK')

print('  [Linear Scaling]   ...', end=' ', flush=True)
cal_ls  = linear_scaling(obs_cal, gcm_cal, gcm_cal)
print('OK')

print('  [LOCI]             ...', end=' ', flush=True)
cal_loc = loci(obs_cal, gcm_cal, gcm_cal, UMBRAL_MM)
print('OK')

# ─── PASO 4: Métricas ────────────────────────────────────────────────────────
titulo('PASO 4 — MÉTRICAS DE AJUSTE EN CALIBRACIÓN')

print("""
  ╔════════════════════════════════════════════════════════════════════════════╗
  ║  GUÍA DE INTERPRETACIÓN                                                    ║
  ╠════════════════════════════════════════════════════════════════════════════╣
  ║  EDCDFm usa corrección MULTIPLICATIVA (ratio q_obs / q_gcm).               ║
  ║  Esto preserva no-negatividad y es más robusto para extremos que ISIMIP.   ║
  ║                                                                            ║
  ║  MÉTRICAS DE DISTRIBUCIÓN (prioridad para QM):                            ║
  ║    PBIAS, Error Media, Error StdDev → deben ser ≈ 0%                      ║
  ║                                                                            ║
  ║  MÉTRICAS DE OCURRENCIA:                                                  ║
  ║    Días secos OBS vs SIM → importantes para hidrología                    ║
  ║    SDII → intensidad media de días húmedos                                ║
  ║                                                                            ║
  ║  MÉTRICAS DE EXTREMOS (clave para EDCDFm):                               ║
  ║    P95, P99 → percentiles altos de días húmedos                           ║
  ║    Rx1day → máximo diario en el período                                   ║
  ║                                                                            ║
  ║  TEST KS: p > 0.05 → distribuciones estadísticamente similares            ║
  ╚════════════════════════════════════════════════════════════════════════════╝
""")

m_sin = calcular_metricas(obs_cal, gcm_cal,  'Sin corrección  (GCM raw)')
m_emp = calcular_metricas(obs_cal, cal_emp,  'EDCDFm Empírico')
m_gam = calcular_metricas(obs_cal, cal_gam,  'EDCDFm Gamma')
m_ls  = calcular_metricas(obs_cal, cal_ls,   'Linear Scaling (LS)')
m_loc = calcular_metricas(obs_cal, cal_loc,  'LOCI')

for m in [m_sin, m_emp, m_gam, m_ls, m_loc]:
    imprimir_metricas(m)

# ─── PASO 5: Diagnóstico ─────────────────────────────────────────────────────
titulo('PASO 5 — DIAGNÓSTICO COMPARATIVO DE MÉTODOS')

resultados = []
for m, lbl in [(m_sin, 'Sin corrección'),
               (m_emp, 'EDCDFm Empírico'),
               (m_gam, 'EDCDFm Gamma'),
               (m_ls,  'Linear Scaling'),
               (m_loc, 'LOCI')]:
    icon, nivel, desc = clasificar(m)
    resultados.append({'metodo': lbl, 'icon': icon, 'nivel': nivel, 'desc': desc, 'm': m})
    print(f'\n  {icon}  {lbl:<30} → {nivel}')
    if m:
        print(f'       PBIAS: {m["PBIAS (%)"]:>7.2f}% | Error μ: {m["Error Media (%)"]:>6.2f}% | Error σ: {m["Error StdDev (%)"]:>6.2f}% | Err P95: {m["Error P95 (%)"]:.2f}%')

# Mejor método
mejor = max(resultados[1:], key=lambda x: x['nivel'] if x['nivel'] else '')
print(f"""
  {'─'*76}
  ╔═══════════════════════════════════════════════════════════════════════════╗
  ║  MEJOR MÉTODO: {mejor["icon"]} {mejor["metodo"]:<28} → {mejor["nivel"]:<16}  ║
  ╚═══════════════════════════════════════════════════════════════════════════╝
  {mejor["desc"]}
""")

print("""
  CRITERIOS UTILIZADOS (EDCDFm diario):
  ────────────────────────────────────────────────────────────────────────────
  ✅✅✅ EXCELENTE : PBIAS<5%  | Err μ<3%  | Err σ<5%  | Err P95<10% | ΔFsec<2pp
  ✅✅  MUY BUENO : PBIAS<10% | Err μ<5%  | Err σ<10% | Err P95<20% | ΔFsec<5pp
  ✅   BUENO     : PBIAS<15% | Err μ<10% | Err σ<15%
  ⚠️   ACEPTABLE  : PBIAS<25% | Err μ<15% | Err σ<20%
  ❌   DEFICIENTE : fuera de los rangos anteriores
  ────────────────────────────────────────────────────────────────────────────

  ¿CUÁNDO USAR CADA MÉTODO?
  ────────────────────────────────────────────────────────────────────────────
  EDCDFm Empírico : RECOMENDADO como método principal. Robusto para extremos.
  EDCDFm Gamma    : Preferido en regímenes semi-áridos (pocos días húmedos).
  Linear Scaling  : Solo como referencia o si los datos son muy escasos.
  LOCI            : Cuando la frecuencia de ocurrencia de lluvia es incorrecta.
  ────────────────────────────────────────────────────────────────────────────
""")

# ─── PASO 6: Gráficos ────────────────────────────────────────────────────────
titulo('PASO 6 — Generando gráficos PNG', nivel=2)
pngs = []

series_cal = {
    'OBS'             : obs_cal,
    'GCM sin corr.'   : gcm_cal,
    'EDCDFm Empírico' : cal_emp,
    'EDCDFm Gamma'    : cal_gam,
    'Linear Scaling'  : cal_ls,
    'LOCI'            : cal_loc,
}

print('\n  FDC diaria...')
p = plot_fdc(series_cal,
             os.path.join(PNG_DIR, f'{MODELO}_FDC_diario.png'),
             f'FDC Diaria — Calibración — {MODELO}')
pngs.append({'Gráfico': 'FDC Diaria', 'Descripción': 'Curva de duración de frecuencia (lineal y log)', 'Ruta': p})

print('\n  Histograma días húmedos...')
p = plot_hist(series_cal,
              os.path.join(PNG_DIR, f'{MODELO}_Hist_diario.png'),
              f'Histograma (días húmedos) — Calibración — {MODELO}')
pngs.append({'Gráfico': 'Histograma', 'Descripción': 'Densidad de precipitación días húmedos', 'Ruta': p})

print('\n  QQ-plot días húmedos...')
sims_qq = {k: v for k, v in series_cal.items() if k != 'OBS'}
p = plot_qq(obs_cal, sims_qq,
            os.path.join(PNG_DIR, f'{MODELO}_QQ_diario.png'),
            f'QQ-plot Diario — Calibración — {MODELO}')
pngs.append({'Gráfico': 'QQ-plot', 'Descripción': 'Cuantiles OBS vs cada método', 'Ruta': p})

print('\n  Índices de extremos...')
metodos_ext = ['Sin corrección', 'EDCDFm Empírico', 'EDCDFm Gamma', 'Linear Scaling', 'LOCI']
metricas_ext = [m_sin, m_emp, m_gam, m_ls, m_loc]
p = plot_extremos(metricas_ext, metodos_ext,
                  os.path.join(PNG_DIR, f'{MODELO}_Extremos_diario.png'),
                  f'Índices de Extremos — Calibración — {MODELO}')
pngs.append({'Gráfico': 'Extremos', 'Descripción': 'P95, P99, Rx1day, SDII por método', 'Ruta': p})

print('\n  Fracción días secos...')
p = plot_dias_secos(metricas_ext, metodos_ext,
                    os.path.join(PNG_DIR, f'{MODELO}_DíasSecos_diario.png'),
                    f'Fracción Días Secos — Calibración — {MODELO}')
pngs.append({'Gráfico': 'Días secos', 'Descripción': 'Frecuencia de días sin lluvia por método', 'Ruta': p})

# ─── PASO 7: Excel ───────────────────────────────────────────────────────────
titulo('PASO 7 — Exportando resultados a Excel', nivel=2)

wb = Workbook()
wb.remove(wb.active)

# ── calibracion_diaria ──
ws = wb.create_sheet('calibracion_diaria')
df_cal = pd.DataFrame({
    'obs_mm_dia'            : obs_cal,
    'gcm_sin_corr'          : gcm_cal,
    'edcdfm_emp_cal'        : cal_emp,
    'edcdfm_gam_cal'        : cal_gam,
    'linear_scaling_cal'    : cal_ls,
    'loci_cal'              : cal_loc,
})
df_to_sheet(ws, df_cal)
print('  Hoja "calibracion_diaria" ........... OK')

# ── futuro_diario ──
ws = wb.create_sheet('futuro_diario')
df_fut = pd.DataFrame({
    'gcm_sin_corr'          : gcm_fut,
    'edcdfm_emp_fut'        : fut_emp,
    'edcdfm_gam_fut'        : fut_gam,
    'linear_scaling_fut'    : fut_ls,
    'loci_fut'              : fut_loc,
})
df_to_sheet(ws, df_fut)
print('  Hoja "futuro_diario" ................ OK')

# ── metricas ──
ws = wb.create_sheet('metricas')
filas = []
for m in [m_sin, m_emp, m_gam, m_ls, m_loc]:
    if m:
        filas.append(m)
df_met = pd.DataFrame(filas)
df_to_sheet(ws, df_met, header_fill='1a6e9e')
print('  Hoja "metricas" ..................... OK')

# ── extremos ──
ws = wb.create_sheet('extremos')
cols_ext = ['Nombre', 'P95 OBS (mm)', 'P95 SIM (mm)', 'Error P95 (%)',
            'P99 OBS (mm)', 'P99 SIM (mm)', 'Error P99 (%)',
            'Rx1day OBS (mm)', 'Rx1day SIM (mm)', 'Error Rx1day (%)',
            'SDII OBS (mm/día húmedo)', 'SDII SIM (mm/día húmedo)', 'Error SDII (%)',
            'Días secos OBS (%)', 'Días secos SIM (%)', 'Error Frec. seca (pp)']
filas_ext = []
for m in [m_sin, m_emp, m_gam, m_ls, m_loc]:
    if m:
        filas_ext.append({c: m.get(c) for c in cols_ext})
df_ext = pd.DataFrame(filas_ext)
df_to_sheet(ws, df_ext, header_fill='c0392b')
print('  Hoja "extremos" ..................... OK')

# ── diagnostico ──
ws = wb.create_sheet('diagnostico')
filas_diag = []
for res in resultados:
    m = res['m']
    filas_diag.append({
        'Método'           : res['metodo'],
        'Nivel'            : res['nivel'],
        'Diagnóstico'      : res['icon'] + ' ' + res['nivel'],
        'Descripción'      : res['desc'],
        'PBIAS (%)'        : m['PBIAS (%)']        if m else None,
        'Error Media (%)'  : m['Error Media (%)']  if m else None,
        'Error StdDev (%)' : m['Error StdDev (%)'] if m else None,
        'Error P95 (%)'    : m['Error P95 (%)']    if m else None,
        'KS p-valor'       : m['KS p-valor']       if m else None,
    })
df_diag = pd.DataFrame(filas_diag)
df_to_sheet(ws, df_diag, header_fill='7d3c98')
print('  Hoja "diagnostico" .................. OK')

# ── graficos ──
ws = wb.create_sheet('graficos')
df_pngs = pd.DataFrame(pngs)
df_to_sheet(ws, df_pngs, header_fill='2ca02c')
print('  Hoja "graficos" ..................... OK')

# ── resumen ──
ws = wb.create_sheet('resumen')
ws.column_dimensions['A'].width = 100

lineas = [
    '═' * 96,
    f'RESUMEN EJECUTIVO — EDCDFm Bias Correction | Precipitación Diaria | Modelo: {MODELO}',
    '═' * 96,
    '',
    f'  Calibración  : {CAL_INI} → {CAL_FIN}   ({len(obs_cal):,} días)',
    f'  Futuro       : {FUT_INI} → {FUT_FIN}   ({len(gcm_fut):,} días)',
    f'  Umbral lluvia: {UMBRAL_MM} mm/día',
    '',
    '  MÉTODOS IMPLEMENTADOS:',
    '  ─────────────────────────────────────────────────────────────────────────',
    '  1. EDCDFm Empírico   — corrección multiplicativa por cuantiles empíricos',
    '  2. EDCDFm Gamma      — corrección paramétrica Gamma para días húmedos',
    '  3. Linear Scaling    — escala global por razón de medias (referencia)',
    '  4. LOCI              — Linear scaling con corrección de frecuencia',
    '',
    '  ¿POR QUÉ EDCDFm ES MEJOR QUE ISIMIP PARA PRECIPITACIÓN?',
    '  ─────────────────────────────────────────────────────────────────────────',
    '  • ISIMIP usa corrección ADITIVA (resta): puede generar valores negativos',
    '  • EDCDFm usa corrección MULTIPLICATIVA (razón): garantiza no-negatividad',
    '  • EDCDFm preserva mejor los extremos (P95, P99, Rx1day)',
    '  • EDCDFm mantiene la señal de cambio futuro del modelo (delta)',
    '',
    '  DIAGNÓSTICO POR MÉTODO:',
    '  ─────────────────────────────────────────────────────────────────────────',
]

for res in resultados:
    m = res['m']
    if m:
        lineas.append(
            f'  {res["icon"]} {res["metodo"]:<30} | PBIAS: {m["PBIAS (%)"]:>7.2f}% | '
            f'Err μ: {m["Error Media (%)"]:>6.2f}% | Err σ: {m["Error StdDev (%)"]:>6.2f}% | '
            f'Err P95: {m["Error P95 (%)"]:.2f}%'
        )

lineas += [
    '',
    f'  MEJOR MÉTODO: {mejor["icon"]} {mejor["metodo"]} → {mejor["nivel"]}',
    f'  {mejor["desc"]}',
    '',
    '  ARCHIVOS PNG GENERADOS:',
] + [f'    {p["Gráfico"]}: {os.path.basename(p["Ruta"])}   → {p["Descripción"]}' for p in pngs]

for i, linea in enumerate(lineas, 1):
    c = ws.cell(row=i, column=1, value=linea)
    c.font = Font(name='Arial', size=10)
    if i <= 3:
        c.font = Font(bold=True, size=11, color='FFFFFF', name='Arial')
        c.fill = PatternFill('solid', fgColor='1a6e9e')
    elif any(x in linea for x in ['MÉTODOS', 'DIAGNÓSTICO', '¿POR QUÉ', 'MEJOR MÉTODO']):
        c.font = Font(bold=True, size=10, name='Arial')

wb.save(SALIDA_XLSX)
print(f'\n  ✔ Excel guardado → {SALIDA_XLSX}')

# ─── FIN ─────────────────────────────────────────────────────────────────────
titulo('PROCESO COMPLETADO EXITOSAMENTE')
print(f"""
  ╔═══════════════════════════════════════════════════════════════════════════╗
  ║  ARCHIVOS GENERADOS                                                       ║
  ╠═══════════════════════════════════════════════════════════════════════════╣
  │  Excel → {SALIDA_XLSX}
  │
  │  Hojas Excel:
  │    calibracion_diaria — OBS + 4 métodos corregidos (período cal.)
  │    futuro_diario      — GCM sin corr. + 4 métodos corregidos (futuro)
  │    metricas           — todas las métricas de ajuste
  │    extremos           — P95, P99, Rx1day, SDII, días secos por método
  │    diagnostico        — clasificación de calidad por método
  │    graficos           — rutas de los PNGs generados
  │    resumen            — resumen ejecutivo con recomendaciones
  │
  │  PNGs → {PNG_DIR}
  │    {MODELO}_FDC_diario.png        — FDC lineal + log
  │    {MODELO}_Hist_diario.png       — histograma días húmedos
  │    {MODELO}_QQ_diario.png         — QQ-plot cuantiles
  │    {MODELO}_Extremos_diario.png   — P95, P99, Rx1day, SDII
  │    {MODELO}_DíasSecos_diario.png  — fracción días secos
  ╚═══════════════════════════════════════════════════════════════════════════╝

  PRÓXIMOS PASOS:
  ─────────────────────────────────────────────────────────────────────────────
  1. Excel → hoja "diagnostico": identifica el mejor método para tus datos
  2. Excel → hoja "extremos": valida P95, P99, Rx1day en cada método
  3. PNG FDC log: verifica que las colas de la distribución coincidan con OBS
  4. PNG QQ-plot: puntos sobre la diagonal 1:1 = buen ajuste de cuantiles
  5. Usa "futuro_diario" del mejor método para proyecciones de cambio climático
  ─────────────────────────────────────────────────────────────────────────────
""")