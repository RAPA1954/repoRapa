# =============================================================================
#  ISIMIP BIAS CORRECTION — Análisis DIARIO (Precipitación)
#  Métricas de ajuste por consola + Gráficos PNG (FDC, Histograma, QQ-plot)
#  + Excel con datos diarios, métricas y resumen
# =============================================================================

import os
import warnings
import numpy as np
import pandas as pd
from scipy.stats import gamma as gamma_dist, pearsonr
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')

plt.rcParams.update({
    'font.family'       : 'DejaVu Sans',
    'axes.unicode_minus': False,
    'figure.dpi'        : 150,
})

# =============================================================================
#  CONFIGURACIÓN USUARIO  ← EDITAR AQUÍ
# =============================================================================
OBS_XLSX    = r'C:/1.PYTHON/Descarga_Python/Pd_Historica_SMN.xlsx'
GCM_XLSX    = r'C:/1.PYTHON/Descarga_Python/Tabla_CESM2_historico_futuro.xlsx'
COL_OBS     = 'prec'
COL_GCM     = 'pr_mm'
MODELO      = 'CESM2'
CAL_INI     = '1961-01-01'
CAL_FIN     = '2024-12-31'
FUT_INI     = '2030-01-01'
FUT_FIN     = '2100-12-31'
OUT_DIR     = rf'C:/1.PYTHON/Descarga_Python/{MODELO}_salidas_ISIMIP_diario'
N_QUANTILES = 100
UMBRAL_MM   = 0.1
# =============================================================================

os.makedirs(OUT_DIR, exist_ok=True)
PNG_DIR = os.path.join(OUT_DIR, 'graficos')
os.makedirs(PNG_DIR, exist_ok=True)

SALIDA_XLSX = os.path.join(OUT_DIR, f'{MODELO}_ISIMIP_diario.xlsx')

SEP  = '=' * 78
SEP2 = '-' * 78


# ─────────────────────────────────────────────────────────────────────────────
#  FUNCIONES AUXILIARES
# ─────────────────────────────────────────────────────────────────────────────

def titulo_consola(texto, nivel=1):
    if nivel == 1:
        print(f'\n{SEP}')
        print(f'  {texto}')
        print(SEP)
    else:
        print(f'\n{SEP2}')
        print(f'  {texto}')
        print(SEP2)


def isimip_bias_correction(obs_cal, gcm_cal, gcm_fut,
                            distribucion='empirica', n_quantiles=100, umbral=0.1):
    """
    ISIMIP Quantile Mapping con preservación de tendencia futura.
    Trabaja directamente sobre series diarias.
    """
    obs_cal       = obs_cal.dropna()
    gcm_cal       = gcm_cal.dropna()
    gcm_fut_clean = gcm_fut.dropna()

    if distribucion == 'empirica':
        quantiles = np.linspace(0, 1, n_quantiles + 1)
        q_obs     = np.quantile(obs_cal,       quantiles)
        q_gcm_cal = np.quantile(gcm_cal,       quantiles)
        q_gcm_fut = np.quantile(gcm_fut_clean, quantiles)
        transfer  = q_obs - q_gcm_cal
        q_corr    = q_gcm_fut + transfer
        gcm_corr  = np.interp(gcm_fut_clean.values, q_gcm_fut, q_corr,
                               left=q_corr[0], right=q_corr[-1])

    elif distribucion == 'gamma':
        obs_wet     = obs_cal[obs_cal > umbral]
        gcm_cal_wet = gcm_cal[gcm_cal > umbral]
        gcm_fut_wet = gcm_fut_clean[gcm_fut_clean > umbral]

        if len(obs_wet) < 10 or len(gcm_cal_wet) < 10 or len(gcm_fut_wet) < 10:
            return isimip_bias_correction(obs_cal, gcm_cal, gcm_fut,
                                          distribucion='empirica',
                                          n_quantiles=n_quantiles)

        a_obs, _, sc_obs = gamma_dist.fit(obs_wet,     floc=0)
        a_cal, _, sc_cal = gamma_dist.fit(gcm_cal_wet, floc=0)
        a_fut, _, sc_fut = gamma_dist.fit(gcm_fut_wet, floc=0)

        delta_scale = sc_obs - sc_cal
        sc_corr     = max(sc_fut + delta_scale, 1e-6)

        p        = gamma_dist.cdf(gcm_fut_wet.values, a_fut, scale=sc_fut)
        corr_wet = gamma_dist.ppf(p,                  a_obs, scale=sc_corr)

        gcm_corr_s                      = gcm_fut_clean.copy().astype(float)
        gcm_corr_s[gcm_fut_clean > umbral]  = corr_wet
        gcm_corr_s[gcm_fut_clean <= umbral] = 0.0
        gcm_corr = gcm_corr_s.values

    else:
        raise ValueError("distribucion debe ser 'empirica' o 'gamma'")

    return pd.Series(np.clip(gcm_corr, 0, None), index=gcm_fut_clean.index)


def calcular_metricas(obs, sim, nombre=''):
    """Calcula métricas de evaluación de QM para series diarias."""
    df = pd.DataFrame({'obs': obs, 'sim': sim}).dropna()
    o  = df['obs'].values
    s  = df['sim'].values
    n  = len(o)

    if n < 3:
        return None

    nse   = 1 - np.sum((o - s)**2) / np.sum((o - np.mean(o))**2)
    r, _  = pearsonr(o, s)
    alpha = np.std(s)  / np.std(o)
    beta  = np.mean(s) / np.mean(o) if np.mean(o) != 0 else np.nan
    kge   = 1 - np.sqrt((r-1)**2 + (alpha-1)**2 + (beta-1)**2) if not np.isnan(beta) else np.nan
    pbias = (np.sum(s - o) / np.sum(o)) * 100 if np.sum(o) != 0 else 0
    rmse  = np.sqrt(np.mean((o - s)**2))
    mae   = np.mean(np.abs(o - s))
    r2    = r**2

    media_obs  = np.mean(o)
    media_sim  = np.mean(s)
    err_media  = abs(media_obs - media_sim) / media_obs * 100 if media_obs != 0 else 0

    std_obs = np.std(o)
    std_sim = np.std(s)
    err_std = abs(std_obs - std_sim) / std_obs * 100 if std_obs != 0 else 0

    # Fracción de días secos (≤ umbral)
    frac_secos_obs = np.mean(o <= UMBRAL_MM) * 100
    frac_secos_sim = np.mean(s <= UMBRAL_MM) * 100

    return {
        'Nombre'           : nombre,
        'n'                : n,
        'NSE'              : round(nse,   4),
        'KGE'              : round(float(kge), 4) if not np.isnan(kge) else None,
        'r (Pearson)'      : round(r,     4),
        'R²'               : round(r2,    4),
        'PBIAS (%)'        : round(pbias, 2),
        'RMSE (mm/día)'    : round(rmse,  4),
        'MAE (mm/día)'     : round(mae,   4),
        'Media OBS'        : round(media_obs, 4),
        'Media SIM'        : round(media_sim, 4),
        'Error Media (%)'  : round(err_media, 2),
        'StdDev OBS'       : round(std_obs,   4),
        'StdDev SIM'       : round(std_sim,   4),
        'Error StdDev (%)' : round(err_std,   2),
        'Días secos OBS (%)': round(frac_secos_obs, 2),
        'Días secos SIM (%)': round(frac_secos_sim, 2),
    }


def imprimir_metricas(m):
    if m is None:
        print('  !! Insuficientes datos para calcular métricas.')
        return
    print(f'\n  ── {m["Nombre"]} ──')
    print(f'     n                  : {m["n"]:,}')
    print(f'     NSE                : {m["NSE"]:>8.4f}   (ℹ️  informativo: QM no preserva sincronía)')
    print(f'     KGE                : {str(m["KGE"]):>8}   (ℹ️  informativo: QM no preserva sincronía)')
    print(f'     r  (Pearson)       : {m["r (Pearson)"]:>8.4f}   (ℹ️  informativo: QM no preserva sincronía)')
    print(f'     R²                 : {m["R²"]:>8.4f}')
    print()
    print(f'     ✅ PBIAS (%)          : {m["PBIAS (%)"]:>8.2f}   ← MÉTRICA CLAVE para QM')
    print(f'     ✅ Error Media (%)    : {m["Error Media (%)"]:>8.2f}   ← MÉTRICA CLAVE para QM')
    print(f'     ✅ Error StdDev (%)   : {m["Error StdDev (%)"]:>8.2f}   ← MÉTRICA CLAVE para QM')
    print()
    print(f'     Media OBS/SIM      : {m["Media OBS"]:>8.4f} / {m["Media SIM"]:>8.4f}  mm/día')
    print(f'     StdDev OBS/SIM     : {m["StdDev OBS"]:>8.4f} / {m["StdDev SIM"]:>8.4f}  mm/día')
    print(f'     RMSE (mm/día)      : {m["RMSE (mm/día)"]:>8.4f}')
    print(f'     MAE  (mm/día)      : {m["MAE (mm/día)"]:>8.4f}')
    print(f'     Días secos OBS/SIM : {m["Días secos OBS (%)"]:>6.2f}% / {m["Días secos SIM (%)"]:>6.2f}%')


def clasificar_metrica_qm(m):
    """
    Clasifica el desempeño del QM diario según PBIAS, Error Media y Error StdDev.
    Para escala diaria los umbrales son más permisivos.
    """
    if m is None:
        return '⚠️  SIN DATOS', 'INSUFICIENTE', 'No se pudo evaluar.'

    pbias     = abs(m['PBIAS (%)'])
    err_media = m['Error Media (%)']
    err_std   = m['Error StdDev (%)']

    if pbias < 5 and err_media < 3 and err_std < 5:
        return '✅✅✅', 'EXCELENTE', (
            'Ajuste excelente a escala diaria. ISIMIP es ALTAMENTE RECOMENDADO '
            'para proyecciones de cambio climático.')

    elif pbias < 10 and err_media < 5 and err_std < 10:
        return '✅✅', 'MUY BUENO', (
            'Ajuste muy bueno. ISIMIP es RECOMENDADO. '
            'La corrección de distribución diaria es robusta.')

    elif pbias < 15 and err_media < 10 and err_std < 15:
        return '✅', 'BUENO', (
            'Ajuste bueno. ISIMIP es ACEPTABLE a escala diaria. '
            'Revisa FDC y QQ-plot para validar extremos.')

    elif pbias < 25 and err_media < 15 and err_std < 20:
        return '⚠️', 'ACEPTABLE', (
            'Ajuste moderado. La distribución diaria tiene sesgo residual. '
            'Considera complementar con validación estacional.')

    else:
        return '❌', 'DEFICIENTE', (
            'Ajuste insatisfactorio. Revisa calidad de datos, resolución '
            'del modelo o considera otros métodos de corrección.')


# ─────────────────────────────────────────────────────────────────────────────
#  FUNCIONES DE GRÁFICOS
# ─────────────────────────────────────────────────────────────────────────────

COLORES = {
    'OBS': '#1a6e9e',
    'SIN': '#d94f3d',
    'EMP': '#2ca02c',
    'GAM': '#ff7f0e',
}


def _guardar(fig, ruta):
    try:
        fig.tight_layout()
    except Exception:
        pass
    fig.savefig(ruta, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f'     PNG guardado → {ruta}')
    return ruta


def plot_fdc(series_dict, ruta_png, titulo):
    """Curva de Duración de Frecuencia (FDC / excedencia) diaria."""
    fig, ax = plt.subplots(figsize=(10, 5))
    estilos = ['-', '--', '-.', ':']
    colores = [COLORES['OBS'], COLORES['SIN'], COLORES['EMP'], COLORES['GAM']]

    for i, (lbl, s) in enumerate(series_dict.items()):
        data = s.dropna().sort_values(ascending=False).values
        n    = len(data)
        prob = np.arange(1, n + 1) / (n + 1) * 100
        ax.plot(prob, data,
                linestyle=estilos[i % 4],
                color=colores[i % 4],
                linewidth=1.5,
                label=lbl)

    ax.set_xlabel('Probabilidad de excedencia (%)', fontsize=11)
    ax.set_ylabel('Precipitación (mm/día)', fontsize=11)
    ax.set_title(titulo, fontsize=12, fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(True, linestyle='--', alpha=0.5)
    ax.xaxis.set_major_formatter(mticker.FormatStrFormatter('%g%%'))
    return _guardar(fig, ruta_png)


def plot_hist(series_dict, ruta_png, titulo):
    """Histogramas de densidad superpuestos (días húmedos)."""
    fig, ax = plt.subplots(figsize=(10, 5))
    colores = [COLORES['OBS'], COLORES['SIN'], COLORES['EMP'], COLORES['GAM']]

    for i, (lbl, s) in enumerate(series_dict.items()):
        data = s.dropna()
        data = data[data > UMBRAL_MM].values  # solo días húmedos
        ax.hist(data, bins=40, density=True, alpha=0.45,
                color=colores[i % 4], label=lbl, edgecolor='white')

    ax.set_xlabel('Precipitación días húmedos (mm/día)', fontsize=11)
    ax.set_ylabel('Densidad', fontsize=11)
    ax.set_title(titulo, fontsize=12, fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(True, linestyle='--', alpha=0.5)
    return _guardar(fig, ruta_png)


def plot_qq(obs, sims_dict, ruta_png, titulo):
    """QQ-plot: cuantiles de OBS vs cuantiles de cada SIM."""
    quantiles = np.linspace(0, 1, 101)
    q_obs     = np.quantile(obs.dropna(), quantiles)

    colores = [COLORES['SIN'], COLORES['EMP'], COLORES['GAM']]
    fig, ax = plt.subplots(figsize=(7, 7))

    lim_max = q_obs[-1]
    for i, (lbl, s) in enumerate(sims_dict.items()):
        q_sim = np.quantile(s.dropna(), quantiles)
        ax.scatter(q_obs, q_sim, s=18, alpha=0.7,
                   color=colores[i % 3], label=lbl)
        lim_max = max(lim_max, q_sim[-1])

    ax.plot([0, lim_max], [0, lim_max], 'k--', linewidth=1.2, label='1:1 (perfecto)')
    ax.set_xlabel('Cuantiles OBS (mm/día)', fontsize=11)
    ax.set_ylabel('Cuantiles SIM (mm/día)', fontsize=11)
    ax.set_title(titulo, fontsize=12, fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(True, linestyle='--', alpha=0.5)
    return _guardar(fig, ruta_png)


# ─────────────────────────────────────────────────────────────────────────────
#  FUNCIÓN EXPORTAR EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def df_to_sheet(ws, df, header_fill='1a6e9e'):
    fill   = PatternFill('solid', fgColor=header_fill)
    font_h = Font(bold=True, color='FFFFFF', name='Arial')
    font_d = Font(name='Arial', size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if r_idx == 1:
                cell.fill = fill
                cell.font = font_h
            else:
                cell.font = font_d
        if r_idx == 1:
            ws.row_dimensions[1].height = 20

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)


# =============================================================================
#  INICIO DE EJECUCIÓN
# =============================================================================

titulo_consola(f'ISIMIP BIAS CORRECTION — Análisis DIARIO (Precipitación)\n  Modelo: {MODELO}')

# ─── PASO 1: Lectura de datos ─────────────────────────────────────────────────
titulo_consola('PASO 1 — Lectura de archivos de entrada', nivel=2)

print(f'  OBS → {OBS_XLSX}')
obs_raw = pd.read_excel(OBS_XLSX, parse_dates=['fecha'], index_col='fecha')
print(f'  GCM → {GCM_XLSX}')
gcm_raw = pd.read_excel(GCM_XLSX, parse_dates=['fecha'], index_col='fecha')

obs_d = obs_raw[COL_OBS].sort_index()
gcm_d = gcm_raw[COL_GCM].sort_index()

# Series diarias por período
obs_cal_d = obs_d[CAL_INI:CAL_FIN]
gcm_cal_d = gcm_d[CAL_INI:CAL_FIN]
gcm_fut_d = gcm_d[FUT_INI:FUT_FIN]

print(f'\n  Período calibración : {CAL_INI} → {CAL_FIN}')
print(f'  Período futuro      : {FUT_INI} → {FUT_FIN}')
print(f'  Datos diarios OBS   : {len(obs_cal_d):,} registros')
print(f'  Datos diarios GCM   : {len(gcm_cal_d):,} (cal)  |  {len(gcm_fut_d):,} (fut)')
print(f'  Días húmedos OBS    : {(obs_cal_d > UMBRAL_MM).sum():,} ({(obs_cal_d > UMBRAL_MM).mean()*100:.1f}%)')
print(f'  Días húmedos GCM    : {(gcm_cal_d > UMBRAL_MM).sum():,} ({(gcm_cal_d > UMBRAL_MM).mean()*100:.1f}%)')

# ─── PASO 2: Corrección ISIMIP — Período Futuro ───────────────────────────────
titulo_consola('PASO 2 — Corrección ISIMIP al período FUTURO (escala diaria)', nivel=2)
print('  (No hay OBS futuro; corrección sin evaluación directa)')

print('  [Diario] Empírico...', end=' ', flush=True)
gcm_fut_d_emp = isimip_bias_correction(obs_cal_d, gcm_cal_d, gcm_fut_d,
                                        'empirica', N_QUANTILES)
print('OK')

print('  [Diario] Gamma...', end=' ', flush=True)
gcm_fut_d_gam = isimip_bias_correction(obs_cal_d, gcm_cal_d, gcm_fut_d,
                                        'gamma', N_QUANTILES, UMBRAL_MM)
print('OK')

# ─── PASO 3: Retroproyección en calibración ──────────────────────────────────
titulo_consola('PASO 3 — Retroproyección en CALIBRACIÓN (comparación con OBS)', nivel=2)
print('  Se corrige el propio GCM_cal usando GCM_cal como "futuro" y se compara con OBS.')

print('  [Diario] Empírico...', end=' ', flush=True)
gcm_cal_d_emp = isimip_bias_correction(obs_cal_d, gcm_cal_d, gcm_cal_d,
                                        'empirica', N_QUANTILES)
print('OK')

print('  [Diario] Gamma...', end=' ', flush=True)
gcm_cal_d_gam = isimip_bias_correction(obs_cal_d, gcm_cal_d, gcm_cal_d,
                                        'gamma', N_QUANTILES, UMBRAL_MM)
print('OK')

# ─── PASO 4: Métricas de ajuste ──────────────────────────────────────────────
titulo_consola('PASO 4 — MÉTRICAS DE AJUSTE en Calibración (escala DIARIA)')

print("""
  ╔════════════════════════════════════════════════════════════════════════════╗
  ║  NOTA CRUCIAL SOBRE EVALUACIÓN DE QUANTILE MAPPING (QM/ISIMIP)            ║
  ╠════════════════════════════════════════════════════════════════════════════╣
  ║                                                                            ║
  ║  QM es un método de CORRECCIÓN DE DISTRIBUCIÓN, no de predicción:         ║
  ║                                                                            ║
  ║  ✅ CORRIGE (métricas de distribución mejoran):                           ║
  ║     • Sesgo sistemático (PBIAS)                                            ║
  ║     • Media y varianza de la serie diaria                                  ║
  ║     • Distribución de cuantiles (incluyendo extremos)                      ║
  ║     • Fracción de días húmedos/secos                                       ║
  ║                                                                            ║
  ║  ❌ NO PRESERVA (métricas de sincronía típicamente empeoran):             ║
  ║     • Correlación día-a-día (NSE, r, KGE)                                 ║
  ║     • Secuencia exacta de eventos de lluvia                                ║
  ║                                                                            ║
  ║  📊 MÉTRICAS CLAVE para evaluar QM diario:                               ║
  ║     1. PBIAS (%)         → Sesgo de volumen total (0% = perfecto)          ║
  ║     2. Error Media (%)   → Error en precipitación media diaria             ║
  ║     3. Error StdDev (%)  → Error en variabilidad diaria                    ║
  ║     4. Días secos OBS/SIM → Fracción de días sin lluvia                   ║
  ║                                                                            ║
  ╚════════════════════════════════════════════════════════════════════════════╝
""")

m_sin_d = calcular_metricas(obs_cal_d, gcm_cal_d,     'Sin corrección')
m_emp_d = calcular_metricas(obs_cal_d, gcm_cal_d_emp, 'ISIMIP Empírico')
m_gam_d = calcular_metricas(obs_cal_d, gcm_cal_d_gam, 'ISIMIP Gamma')

for m in [m_sin_d, m_emp_d, m_gam_d]:
    imprimir_metricas(m)

# ─── PASO 5: Conclusión automática ───────────────────────────────────────────
titulo_consola('PASO 5 — CONCLUSIÓN: ¿Es correcto usar ISIMIP a escala DIARIA?')

icon_emp, nivel_emp, desc_emp = clasificar_metrica_qm(m_emp_d)
icon_gam, nivel_gam, desc_gam = clasificar_metrica_qm(m_gam_d)

print(f'\n  EVALUACIÓN DE AJUSTE DIARIO (Criterios para QM):')
print('  ' + '─' * 76)

print(f'\n  ISIMIP Empírico:')
print(f'    Diagnóstico: {icon_emp} {nivel_emp}')
if m_emp_d:
    print(f'    PBIAS: {m_emp_d["PBIAS (%)"]:.2f}% | Error Media: {m_emp_d["Error Media (%)"]:.2f}% | Error StdDev: {m_emp_d["Error StdDev (%)"]:.2f}%')

print(f'\n  ISIMIP Gamma:')
print(f'    Diagnóstico: {icon_gam} {nivel_gam}')
if m_gam_d:
    print(f'    PBIAS: {m_gam_d["PBIAS (%)"]:.2f}% | Error Media: {m_gam_d["Error Media (%)"]:.2f}% | Error StdDev: {m_gam_d["Error StdDev (%)"]:.2f}%')

# Diagnóstico global (prioriza Empírico)
icon_global  = icon_emp
nivel_global = nivel_emp
desc_global  = desc_emp

print(f'\n  ' + '─' * 76)
print(f'\n  ╔══════════════════════════════════════════════════════════════════════════╗')
print(f'  ║  RECOMENDACIÓN FINAL: {icon_global} {nivel_global:<56} ║')
print(f'  ╚══════════════════════════════════════════════════════════════════════════╝')
print(f'\n  {desc_global}')

print(f"""
  ╔════════════════════════════════════════════════════════════════════════════╗
  ║  CRITERIOS DE CLASIFICACIÓN (escala DIARIA)                               ║
  ╠════════════════════════════════════════════════════════════════════════════╣
  │  ✅✅✅ EXCELENTE  │ PBIAS<5%  y Error Media<3%  y Error Std<5%           │
  │  ✅✅  MUY BUENO  │ PBIAS<10% y Error Media<5%  y Error Std<10%          │
  │  ✅  BUENO      │ PBIAS<15% y Error Media<10% y Error Std<15%          │
  │  ⚠️  ACEPTABLE   │ PBIAS<25% y Error Media<15% y Error Std<20%          │
  │  ❌  DEFICIENTE  │ Fuera de los rangos anteriores                        │
  ╚════════════════════════════════════════════════════════════════════════════╝

  RECOMENDACIONES PRÁCTICAS:
  ─────────────────────────────────────────────────────────────────────────────
  ✔ Para PROYECCIONES DE TENDENCIA o volumen total → ISIMIP es ADECUADO
  ✔ Para ANÁLISIS DE EXTREMOS diarios → Valida con FDC y QQ-plot
  ✔ Para SIMULACIONES DÍA-A-DÍA (variabilidad temporal) → NSE/r son bajos
    por diseño del QM; evalúa por distribución, no por sincronía
  ✔ SIEMPRE → Valida visualmente con FDC, QQ-plot e histogramas
  ─────────────────────────────────────────────────────────────────────────────
""")

# ─── PASO 6: Gráficos PNG ────────────────────────────────────────────────────
titulo_consola('PASO 6 — Generando gráficos PNG (FDC, Histograma, QQ-plot)', nivel=2)
pngs = []

series_fdc = {
    'OBS'            : obs_cal_d,
    'GCM sin corr.'  : gcm_cal_d,
    'ISIMIP Empírico': gcm_cal_d_emp,
    'ISIMIP Gamma'   : gcm_cal_d_gam,
}

p = plot_fdc(series_fdc,
             os.path.join(PNG_DIR, f'{MODELO}_FDC_diario.png'),
             f'FDC Diaria — Calibración — {MODELO}')
pngs.append({'Gráfico': 'FDC Diaria', 'Ruta': p})

p = plot_hist(series_fdc,
              os.path.join(PNG_DIR, f'{MODELO}_Hist_diario.png'),
              f'Histograma Diario (días húmedos) — Calibración — {MODELO}')
pngs.append({'Gráfico': 'Histograma Diario', 'Ruta': p})

sims_qq = {
    'GCM sin corr.'  : gcm_cal_d,
    'ISIMIP Empírico': gcm_cal_d_emp,
    'ISIMIP Gamma'   : gcm_cal_d_gam,
}
p = plot_qq(obs_cal_d, sims_qq,
            os.path.join(PNG_DIR, f'{MODELO}_QQ_diario.png'),
            f'QQ-plot Diario — Calibración — {MODELO}')
pngs.append({'Gráfico': 'QQ-plot Diario', 'Ruta': p})

# ─── PASO 7: Exportar Excel ──────────────────────────────────────────────────
titulo_consola('PASO 7 — Exportando resultados a Excel (.xlsx)', nivel=2)

wb = Workbook()
wb.remove(wb.active)

# ── Hoja: calibracion_diaria ──
ws = wb.create_sheet('calibracion_diaria')
df_cal_d = pd.DataFrame({
    'obs_mm_dia'         : obs_cal_d,
    'gcm_sin_corr_mm_dia': gcm_cal_d,
    'gcm_emp_cal_mm_dia' : gcm_cal_d_emp,
    'gcm_gam_cal_mm_dia' : gcm_cal_d_gam,
})
df_to_sheet(ws, df_cal_d)
print('  Hoja "calibracion_diaria" .... OK')

# ── Hoja: futuro_diario ──
ws = wb.create_sheet('futuro_diario')
df_fut_d = pd.DataFrame({
    'gcm_sin_corr_mm_dia': gcm_fut_d,
    'gcm_emp_fut_mm_dia' : gcm_fut_d_emp,
    'gcm_gam_fut_mm_dia' : gcm_fut_d_gam,
})
df_to_sheet(ws, df_fut_d)
print('  Hoja "futuro_diario" ......... OK')

# ── Hoja: metricas ──
ws = wb.create_sheet('metricas')
filas_met = []
for m in [m_sin_d, m_emp_d, m_gam_d]:
    if m:
        row = {'Escala': 'Diaria'}
        row.update(m)
        filas_met.append(row)
df_met = pd.DataFrame(filas_met)
df_to_sheet(ws, df_met, header_fill='1a6e9e')
print('  Hoja "metricas" .............. OK')

# ── Hoja: graficos ──
ws = wb.create_sheet('graficos')
df_pngs = pd.DataFrame(pngs)
df_to_sheet(ws, df_pngs, header_fill='2ca02c')
print('  Hoja "graficos" .............. OK')

# ── Hoja: resumen ──
ws = wb.create_sheet('resumen')
ws.column_dimensions['A'].width = 100

lineas = [
    '═' * 95,
    f'RESUMEN EJECUTIVO — ISIMIP Bias Correction DIARIO | Modelo: {MODELO}',
    '═' * 95,
    '',
    'PERÍODO DE ANÁLISIS:',
    f'  Calibración : {CAL_INI} → {CAL_FIN}',
    f'  Futuro      : {FUT_INI} → {FUT_FIN}',
    '',
    'DATOS:',
    f'  OBS diarios  : {len(obs_cal_d):,} registros',
    f'  GCM diarios  : {len(gcm_cal_d):,} (cal)  |  {len(gcm_fut_d):,} (fut)',
    '',
    'DIAGNÓSTICO FINAL:',
    f'  {icon_global} {nivel_global}',
    '',
    f'  {desc_global}',
    '',
    'MÉTRICAS CLAVE (ISIMIP Empírico — escala diaria):',
    '',
]

if m_emp_d:
    lineas += [
        f'    PBIAS            : {m_emp_d["PBIAS (%)"]:.2f}%',
        f'    Error Media      : {m_emp_d["Error Media (%)"]:.2f}%',
        f'    Error StdDev     : {m_emp_d["Error StdDev (%)"]:.2f}%',
        f'    Días secos OBS   : {m_emp_d["Días secos OBS (%)"]:.2f}%',
        f'    Días secos SIM   : {m_emp_d["Días secos SIM (%)"]:.2f}%',
        f'    RMSE (mm/día)    : {m_emp_d["RMSE (mm/día)"]:.4f}',
        f'    MAE  (mm/día)    : {m_emp_d["MAE (mm/día)"]:.4f}',
    ]

lineas += [
    '',
    'CRITERIOS USADOS (QM diario):',
    '  ✅✅✅ EXCELENTE  : PBIAS<5%  y Error Media<3%  y Error Std<5%',
    '  ✅✅  MUY BUENO  : PBIAS<10% y Error Media<5%  y Error Std<10%',
    '  ✅  BUENO      : PBIAS<15% y Error Media<10% y Error Std<15%',
    '  ⚠️  ACEPTABLE   : PBIAS<25% y Error Media<15% y Error Std<20%',
    '  ❌  DEFICIENTE  : Fuera de los rangos anteriores',
    '',
    'ARCHIVOS GENERADOS:',
] + [f'  {p["Gráfico"]} → {os.path.basename(p["Ruta"])}' for p in pngs]

for i, linea in enumerate(lineas, 1):
    c = ws.cell(row=i, column=1, value=linea)
    c.font = Font(name='Arial', size=10)
    if i in [1, 2, 3]:
        c.font = Font(bold=True, size=11, color='FFFFFF', name='Arial')
        c.fill = PatternFill('solid', fgColor='1a6e9e')
    elif 'DIAGNÓSTICO' in linea or 'MÉTRICAS' in linea or 'CRITERIOS' in linea or 'PERÍODO' in linea:
        c.font = Font(bold=True, size=10, name='Arial')

wb.save(SALIDA_XLSX)
print(f'\n  ✔ Excel guardado → {SALIDA_XLSX}')

# ─── FIN ─────────────────────────────────────────────────────────────────────
titulo_consola('PROCESO COMPLETADO EXITOSAMENTE')
print(f"""
  ╔═══════════════════════════════════════════════════════════════════════════╗
  ║  ARCHIVOS GENERADOS                                                       ║
  ╠═══════════════════════════════════════════════════════════════════════════╣
  │  Excel  → {SALIDA_XLSX}
  │  PNGs   → {PNG_DIR}
  │    - {MODELO}_FDC_diario.png
  │    - {MODELO}_Hist_diario.png
  │    - {MODELO}_QQ_diario.png
  ╚═══════════════════════════════════════════════════════════════════════════╝

  PRÓXIMOS PASOS:
  ─────────────────────────────────────────────────────────────────────────────
  1. Abre el Excel → revisa hojas "metricas" y "resumen"
  2. Abre los gráficos PNG → valida FDC, QQ-plot e histogramas diarios
  3. Si FDC/QQ muestran buen ajuste → ISIMIP es CONFIABLE para el futuro
  4. Usa los datos de "futuro_diario" en tus análisis posteriores
  ─────────────────────────────────────────────────────────────────────────────
""")