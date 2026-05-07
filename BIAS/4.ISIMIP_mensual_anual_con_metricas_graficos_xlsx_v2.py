# =============================================================================
#  ISIMIP BIAS CORRECTION — ANÁLISIS COMPLETO Y PROFESIONAL v2.2
#  ✅ FIX: Corrección de read-only en rama Empírica Y Gamma
# =============================================================================

import os
import warnings
import numpy as np
import pandas as pd
from scipy.stats import gamma as gamma_dist, pearsonr
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

warnings.filterwarnings('ignore')

plt.rcParams.update({
    'font.family'       : 'DejaVu Sans',
    'axes.unicode_minus': False,
    'figure.dpi'        : 150,
})

# =============================================================================
#  ⚙️ CONFIGURACIÓN USUARIO — EDITAR AQUÍ
# =============================================================================

OBS_XLSX            = r'C:/1.PYTHON/Descarga_Python/Pd_Historica_SMN.xlsx'
GCM_XLSX            = r'C:/1.PYTHON/Descarga_Python/Tabla_CESM2_historico_futuro.xlsx'
COL_OBS             = 'prec'
COL_GCM             = 'pr_mm'
MODELO              = 'CESM2'
CAL_INI             = '1961-01-01'
CAL_FIN             = '2024-12-31'
FUT_INI             = '2030-01-01'
FUT_FIN             = '2100-12-31'
N_QUANTILES         = 100
UMBRAL_MM           = 0.1
CORREGIR_DIAS_SECOS = True
OUT_DIR             = rf'C:/1.PYTHON/Descarga_Python/{MODELO}_salidas_ISIMIP'
HEMISFERIO          = 'Sur'

# =============================================================================

os.makedirs(OUT_DIR, exist_ok=True)
PNG_DIR     = os.path.join(OUT_DIR, 'graficos')
os.makedirs(PNG_DIR, exist_ok=True)
SALIDA_XLSX = os.path.join(OUT_DIR, f'{MODELO}_ISIMIP_completo.xlsx')

SEP  = '=' * 85
SEP2 = '-' * 85

COLORES = {
    'OBS': '#1a6e9e',
    'SIN': '#d94f3d',
    'EMP': '#2ca02c',
    'GAM': '#ff7f0e',
}

# =============================================================================
#  🔧 FUNCIONES AUXILIARES
# =============================================================================

def titulo_consola(texto, nivel=1):
    if nivel == 1:
        print(f'\n{SEP}\n  {texto}\n{SEP}')
    else:
        print(f'\n{SEP2}\n  {texto}\n{SEP2}')


def asignar_estacion(fecha, hemisferio='Sur'):
    mes = fecha.month
    if hemisferio == 'Sur':
        if mes in [12, 1, 2]:  return 'Verano'
        elif mes in [3, 4, 5]: return 'Otoño'
        elif mes in [6, 7, 8]: return 'Invierno'
        else:                  return 'Primavera'
    else:
        if mes in [6, 7, 8]:   return 'Verano'
        elif mes in [9,10,11]: return 'Otoño'
        elif mes in [12, 1, 2]:return 'Invierno'
        else:                  return 'Primavera'


# =============================================================================
#  📊 FUNCIÓN PRINCIPAL ISIMIP v2.2 — TODOS LOS FIX APLICADOS
# =============================================================================

def isimip_bias_correction_v2(obs_cal, gcm_cal, gcm_fut,
                               distribucion='empirica', n_quantiles=100,
                               umbral=0.1, corregir_dias_secos=True,
                               verbose=False):
    """
    ISIMIP Quantile Mapping con corrección booleana de días secos.
    v2.2: Fix read-only en ramas empírica Y gamma.
    """

    obs_cal      = obs_cal.dropna()
    gcm_cal      = gcm_cal.dropna()
    gcm_fut_clean = gcm_fut.dropna()

    # ─────────────────────────────────────────────────────────────────────────
    # PASO 1 — Corrección Booleana (días secos / lluviosos)
    # ─────────────────────────────────────────────────────────────────────────

    if corregir_dias_secos:

        p_obs = (obs_cal > umbral).sum() / len(obs_cal)
        p_gcm = (gcm_cal > umbral).sum() / len(gcm_cal)
        ratio_prob = (p_obs / p_gcm) if p_gcm > 0.01 else 1.0

        if verbose:
            print(f"    P(lluvia) OBS={p_obs:.1%}  GCM={p_gcm:.1%}  ratio={ratio_prob:.3f}")

        # ✅ FIX 1: .copy() para que el array sea mutable
        gcm_bool = (gcm_fut_clean > umbral).astype(float).values.copy()

        if ratio_prob < 1.0:
            n_actual  = int(np.sum(gcm_bool))
            n_correcto = int(n_actual * ratio_prob)
            n_remover  = n_actual - n_correcto
            if n_remover > 0:
                idx_lluvia = np.where(gcm_bool == 1)[0]
                if len(idx_lluvia) > 0:
                    idx_rem = np.random.choice(idx_lluvia,
                                               min(n_remover, len(idx_lluvia)),
                                               replace=False)
                    gcm_bool[idx_rem] = 0           # ✅ sin error

        elif ratio_prob > 1.0:
            n_actual   = int(np.sum(gcm_bool))
            n_correcto = int(n_actual * ratio_prob)
            n_agregar  = n_correcto - n_actual
            if n_agregar > 0:
                idx_secos = np.where(gcm_bool == 0)[0]
                if len(idx_secos) > 0:
                    idx_agr = np.random.choice(idx_secos,
                                               min(n_agregar, len(idx_secos)),
                                               replace=False)
                    gcm_bool[idx_agr] = 1           # ✅ sin error

        mask_lluvia = gcm_bool == 1

    else:
        mask_lluvia = gcm_fut_clean.values > umbral

    # ─────────────────────────────────────────────────────────────────────────
    # PASO 2 — Distributional Mapping (QM)
    # ─────────────────────────────────────────────────────────────────────────

    if distribucion == 'empirica':

        quantiles   = np.linspace(0, 1, n_quantiles + 1)
        q_obs       = np.quantile(obs_cal,       quantiles)
        q_gcm_cal   = np.quantile(gcm_cal,       quantiles)
        q_gcm_fut   = np.quantile(gcm_fut_clean, quantiles)

        transfer    = q_obs - q_gcm_cal
        q_corr      = q_gcm_fut + transfer

        gcm_corr_all = np.interp(gcm_fut_clean.values, q_gcm_fut, q_corr,
                                 left=q_corr[0], right=q_corr[-1])

        gcm_corr = np.where(mask_lluvia, gcm_corr_all, 0.0)

    elif distribucion == 'gamma':

        obs_wet      = obs_cal[obs_cal > umbral]
        gcm_cal_wet  = gcm_cal[gcm_cal > umbral]
        gcm_fut_wet  = gcm_fut_clean[gcm_fut_clean > umbral]

        # Si hay pocos datos, caer a empírico
        if len(obs_wet) < 10 or len(gcm_cal_wet) < 10 or len(gcm_fut_wet) < 10:
            if verbose:
                print("    ⚠️  Pocos datos para Gamma → usando Empírico")
            return isimip_bias_correction_v2(obs_cal, gcm_cal, gcm_fut,
                                             'empirica', n_quantiles, umbral,
                                             corregir_dias_secos, verbose)

        # Ajuste de distribuciones Gamma
        a_obs, _, sc_obs = gamma_dist.fit(obs_wet,     floc=0)
        a_cal, _, sc_cal = gamma_dist.fit(gcm_cal_wet, floc=0)
        a_fut, _, sc_fut = gamma_dist.fit(gcm_fut_wet, floc=0)

        delta_scale = sc_obs - sc_cal
        sc_corr     = max(sc_fut + delta_scale, 1e-6)

        p        = gamma_dist.cdf(gcm_fut_wet.values, a_fut, scale=sc_fut)
        corr_wet = gamma_dist.ppf(p, a_obs, scale=sc_corr)

        # ✅ FIX 2: .copy() en rama Gamma para que el array sea mutable
        gcm_corr_all = gcm_fut_clean.values.astype(float).copy()
        gcm_corr_all[gcm_fut_clean.values > umbral] = corr_wet   # ✅ sin error

        gcm_corr = np.where(mask_lluvia, gcm_corr_all, 0.0)

    else:
        raise ValueError("distribucion debe ser 'empirica' o 'gamma'")

    return pd.Series(np.clip(gcm_corr, 0, None), index=gcm_fut_clean.index)


# =============================================================================
#  📈 MÉTRICAS DE DESEMPEÑO
# =============================================================================

def calcular_metricas(obs, sim, nombre=''):
    df = pd.DataFrame({'obs': obs, 'sim': sim}).dropna()
    if len(df) < 3:
        return None

    o, s = df['obs'].values, df['sim'].values

    nse   = 1 - np.sum((o - s)**2) / np.sum((o - np.mean(o))**2)
    r, _  = pearsonr(o, s)
    alpha = np.std(s)  / np.std(o)
    beta  = np.mean(s) / np.mean(o)
    kge   = 1 - np.sqrt((r-1)**2 + (alpha-1)**2 + (beta-1)**2)
    pbias = (np.sum(s - o) / np.sum(o)) * 100
    rmse  = np.sqrt(np.mean((o - s)**2))
    mae   = np.mean(np.abs(o - s))

    media_obs  = np.mean(o);  media_sim  = np.mean(s)
    std_obs    = np.std(o);   std_sim    = np.std(s)
    err_media  = abs(media_obs - media_sim) / media_obs * 100 if media_obs != 0 else 0
    err_std    = abs(std_obs   - std_sim)   / std_obs   * 100 if std_obs   != 0 else 0

    return {
        'Nombre'         : nombre,
        'n'              : len(o),
        'NSE'            : round(nse,   4),
        'KGE'            : round(kge,   4),
        'r (Pearson)'    : round(r,     4),
        'R²'             : round(r**2,  4),
        'PBIAS (%)'      : round(pbias, 2),
        'RMSE (mm)'      : round(rmse,  4),
        'MAE (mm)'       : round(mae,   4),
        'Media OBS'      : round(media_obs, 4),
        'Media SIM'      : round(media_sim, 4),
        'Error Media (%)': round(err_media, 2),
        'StdDev OBS'     : round(std_obs,   4),
        'StdDev SIM'     : round(std_sim,   4),
        'Error StdDev (%)': round(err_std,  2),
    }


def imprimir_metricas(m, escala=''):
    if m is None:
        print('  !! Insuficientes datos.'); return
    print(f'\n  ──── {m["Nombre"]} ({escala}) ────')
    print(f'     n               : {m["n"]}')
    print(f'     PBIAS (%)       : {m["PBIAS (%)"]:.2f}   ← MÉTRICA CLAVE')
    print(f'     Error Media (%) : {m["Error Media (%)"]:.2f}   ← MÉTRICA CLAVE')
    print(f'     Error StdDev(%) : {m["Error StdDev (%)"]:.2f}   ← MÉTRICA CLAVE')
    print(f'     R²              : {m["R²"]:.4f}')
    print(f'     RMSE (mm)       : {m["RMSE (mm)"]:.4f}')
    print(f'     MAE  (mm)       : {m["MAE (mm)"]:.4f}')
    print(f'     Media OBS/SIM   : {m["Media OBS"]:.2f} / {m["Media SIM"]:.2f}')
    print(f'     StdDev OBS/SIM  : {m["StdDev OBS"]:.2f} / {m["StdDev SIM"]:.2f}')


def clasificar_metrica_qm(m):
    if m is None:
        return '⚠️', 'SIN DATOS', 'Insuficientes registros.'
    pb  = abs(m['PBIAS (%)'])
    em  = abs(m['Error Media (%)'])
    es  = abs(m['Error StdDev (%)'])
    if   pb < 5  and em < 3  and es < 3:  return '✅✅✅', 'EXCELENTE',  'ISIMIP es ALTAMENTE RECOMENDADO.'
    elif pb < 10 and em < 5  and es < 5:  return '✅✅',   'MUY BUENO',  'ISIMIP es RECOMENDADO.'
    elif pb < 15 and em < 10 and es < 10: return '✅',     'BUENO',      'ISIMIP es ACEPTABLE.'
    elif pb < 25 and em < 15 and es < 15: return '⚠️',    'ACEPTABLE',  'Validar con extremos.'
    else:                                  return '❌',     'DEFICIENTE', 'Revisar datos/modelo.'


# =============================================================================
#  📊 MÉTRICAS DE EXTREMOS
# =============================================================================

def calcular_metricas_extremos(obs, sim, nombre=''):
    df = pd.DataFrame({'obs': obs, 'sim': sim}).dropna()
    if len(df) < 10:
        return None
    o, s = df['obs'].values, df['sim'].values
    resultado = {'Nombre': nombre}
    for p in [1, 5, 25, 50, 75, 95, 99]:
        po = np.percentile(o, p)
        ps = np.percentile(s, p)
        resultado[f'P{p} OBS']      = round(po, 2)
        resultado[f'P{p} SIM']      = round(ps, 2)
        resultado[f'Error P{p} (%)'] = round(((ps - po) / po * 100) if po != 0 else 0, 2)
    return resultado


# =============================================================================
#  🎨 GRÁFICOS
# =============================================================================

def _guardar(fig, ruta):
    try:    fig.tight_layout()
    except: pass
    fig.savefig(ruta, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f'     ✓ {os.path.basename(ruta)}')
    return ruta


def plot_fdc(series_dict, ruta_png, titulo):
    fig, ax = plt.subplots(figsize=(11, 6))
    estilos = ['-', '--', '-.', ':']
    cols    = [COLORES['OBS'], COLORES['SIN'], COLORES['EMP'], COLORES['GAM']]
    for i, (lbl, s) in enumerate(series_dict.items()):
        data = s.dropna().sort_values(ascending=False).values
        prob = np.arange(1, len(data)+1) / (len(data)+1) * 100
        ax.plot(prob, data, linestyle=estilos[i%4], color=cols[i%4],
                linewidth=2, label=lbl, alpha=0.85)
    ax.set_xlabel('Probabilidad de Excedencia (%)', fontsize=11, fontweight='bold')
    ax.set_ylabel('Precipitación (mm)',             fontsize=11, fontweight='bold')
    ax.set_title(titulo, fontsize=12, fontweight='bold')
    ax.legend(fontsize=10); ax.grid(True, ls='--', alpha=0.4); ax.set_xlim(0,100)
    return _guardar(fig, ruta_png)


def plot_hist(series_dict, ruta_png, titulo):
    fig, ax = plt.subplots(figsize=(11, 6))
    cols = [COLORES['OBS'], COLORES['SIN'], COLORES['EMP'], COLORES['GAM']]
    for i, (lbl, s) in enumerate(series_dict.items()):
        ax.hist(s.dropna().values, bins=35, density=True, alpha=0.5,
                color=cols[i%4], label=lbl, edgecolor='white', linewidth=0.5)
    ax.set_xlabel('Precipitación (mm)',       fontsize=11, fontweight='bold')
    ax.set_ylabel('Densidad de Probabilidad', fontsize=11, fontweight='bold')
    ax.set_title(titulo, fontsize=12, fontweight='bold')
    ax.legend(fontsize=10); ax.grid(True, ls='--', alpha=0.4, axis='y')
    return _guardar(fig, ruta_png)


def plot_qq(obs, sims_dict, ruta_png, titulo):
    quantiles = np.linspace(0, 1, 101)
    q_obs     = np.quantile(obs.dropna(), quantiles)
    cols      = [COLORES['SIN'], COLORES['EMP'], COLORES['GAM']]
    fig, ax   = plt.subplots(figsize=(8, 8))
    lim_max   = q_obs[-1]
    for i, (lbl, s) in enumerate(sims_dict.items()):
        q_sim = np.quantile(s.dropna(), quantiles)
        ax.scatter(q_obs, q_sim, s=30, alpha=0.7, color=cols[i%3],
                   label=lbl, edgecolors='white', linewidth=0.5)
        lim_max = max(lim_max, q_sim[-1])
    ax.plot([0, lim_max], [0, lim_max], 'k--', lw=1.5, label='Diagonal 1:1')
    ax.set_xlabel('Cuantiles OBS (mm)', fontsize=11, fontweight='bold')
    ax.set_ylabel('Cuantiles SIM (mm)', fontsize=11, fontweight='bold')
    ax.set_title(titulo, fontsize=12, fontweight='bold')
    ax.legend(fontsize=10); ax.grid(True, ls='--', alpha=0.4)
    ax.set_aspect('equal', adjustable='box')
    return _guardar(fig, ruta_png)


def plot_boxplot(obs, sims_dict, ruta_png, titulo):
    fig, ax = plt.subplots(figsize=(10, 6))
    data_l  = [obs.dropna().values]
    labels  = ['OBS']
    colors  = [COLORES['OBS']]
    for lbl, sim in sims_dict.items():
        data_l.append(sim.dropna().values); labels.append(lbl)
        colors.append(COLORES['EMP'] if 'mp' in lbl else
                      COLORES['GAM'] if 'am' in lbl else COLORES['SIN'])
    bp = ax.boxplot(data_l, labels=labels, patch_artist=True, showmeans=True)
    for patch, color in zip(bp['boxes'], colors):
        patch.set_facecolor(color); patch.set_alpha(0.7)
    for w in bp['whiskers']: w.set(linewidth=1.5, color='gray')
    for med in bp['medians']: med.set(linewidth=2, color='darkred')
    ax.set_ylabel('Precipitación (mm)', fontsize=11, fontweight='bold')
    ax.set_title(titulo, fontsize=12, fontweight='bold')
    ax.grid(True, ls='--', alpha=0.3, axis='y')
    return _guardar(fig, ruta_png)


def plot_serie_temporal(series_dict, ruta_png, titulo, max_puntos=500):
    fig, ax = plt.subplots(figsize=(14, 6))
    cols = [COLORES['OBS'], COLORES['SIN'], COLORES['EMP'], COLORES['GAM']]
    for i, (lbl, s) in enumerate(series_dict.items()):
        data = s.dropna().head(max_puntos)
        ax.plot(range(len(data)), data.values, color=cols[i%4],
                linewidth=1.0, label=lbl, alpha=0.8)
    ax.set_xlabel(f'Días (primeros {max_puntos})', fontsize=11, fontweight='bold')
    ax.set_ylabel('Precipitación (mm)',            fontsize=11, fontweight='bold')
    ax.set_title(titulo, fontsize=12, fontweight='bold')
    ax.legend(fontsize=10); ax.grid(True, ls='--', alpha=0.4)
    return _guardar(fig, ruta_png)


def plot_ciclo_anual(series_dict, ruta_png, titulo):
    """Ciclo anual (promedio mensual)."""
    fig, ax = plt.subplots(figsize=(11, 6))
    cols    = [COLORES['OBS'], COLORES['SIN'], COLORES['EMP'], COLORES['GAM']]
    meses   = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
    for i, (lbl, s) in enumerate(series_dict.items()):
        ciclo = s.dropna().groupby(s.dropna().index.month).mean().reindex(range(1,13))
        ax.plot(range(1,13), ciclo.values, marker='o', color=cols[i%4],
                linewidth=2, label=lbl, alpha=0.85)
    ax.set_xticks(range(1,13)); ax.set_xticklabels(meses, fontsize=10)
    ax.set_xlabel('Mes', fontsize=11, fontweight='bold')
    ax.set_ylabel('Precipitación media (mm)', fontsize=11, fontweight='bold')
    ax.set_title(titulo, fontsize=12, fontweight='bold')
    ax.legend(fontsize=10); ax.grid(True, ls='--', alpha=0.4)
    return _guardar(fig, ruta_png)


# =============================================================================
#  📋 EXPORTAR A EXCEL
# =============================================================================

def df_to_sheet(ws, df, header_fill='1a6e9e'):
    fill   = PatternFill('solid', fgColor=header_fill)
    font_h = Font(bold=True, color='FFFFFF', size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))
    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border    = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if r_idx == 1:
                cell.fill = fill; cell.font = font_h
        if r_idx == 1:
            ws.row_dimensions[1].height = 22
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 35)


# =============================================================================
#  🚀 EJECUCIÓN PRINCIPAL
# =============================================================================

titulo_consola(f'ISIMIP BIAS CORRECTION v2.2 — {MODELO}\n  Fix: read-only en ramas Empírica y Gamma')

# ─────────────────────────────────────────────────────────────────────────────
# PASO 1 — Lectura
# ─────────────────────────────────────────────────────────────────────────────

titulo_consola('PASO 1 — Lectura de Datos', nivel=2)

print(f'  OBS → {OBS_XLSX}')
obs_raw = pd.read_excel(OBS_XLSX, parse_dates=['fecha'], index_col='fecha')

print(f'  GCM → {GCM_XLSX}')
gcm_raw = pd.read_excel(GCM_XLSX, parse_dates=['fecha'], index_col='fecha')

obs_d     = obs_raw[COL_OBS].sort_index()
gcm_d     = gcm_raw[COL_GCM].sort_index()

obs_cal_d = obs_d[CAL_INI:CAL_FIN]
gcm_cal_d = gcm_d[CAL_INI:CAL_FIN]
gcm_fut_d = gcm_d[FUT_INI:FUT_FIN]

print(f'\n  Período calibración : {CAL_INI} → {CAL_FIN}')
print(f'  Período futuro      : {FUT_INI} → {FUT_FIN}')
print(f'  OBS diario          : {len(obs_cal_d):,} registros')
print(f'  GCM diario          : {len(gcm_cal_d):,} (cal)  +  {len(gcm_fut_d):,} (fut)')

# ─────────────────────────────────────────────────────────────────────────────
# PASO 2 — Agregación
# ─────────────────────────────────────────────────────────────────────────────

titulo_consola('PASO 2 — Agregación Temporal', nivel=2)

obs_men   = obs_cal_d.resample('ME').sum()
gcm_cal_m = gcm_cal_d.resample('ME').sum()
gcm_fut_m = gcm_fut_d.resample('ME').sum()

obs_anu   = obs_cal_d.resample('YE').sum()
gcm_cal_a = gcm_cal_d.resample('YE').sum()
gcm_fut_a = gcm_fut_d.resample('YE').sum()

obs_men_est_label   = obs_men.index.map(lambda x: asignar_estacion(x, HEMISFERIO))
gcm_cal_m_est_label = gcm_cal_m.index.map(lambda x: asignar_estacion(x, HEMISFERIO))
gcm_fut_m_est_label = gcm_fut_m.index.map(lambda x: asignar_estacion(x, HEMISFERIO))

print(f'  Mensual — OBS: {len(obs_men)} | GCM_cal: {len(gcm_cal_m)} | GCM_fut: {len(gcm_fut_m)}')
print(f'  Anual   — OBS: {len(obs_anu)} | GCM_cal: {len(gcm_cal_a)} | GCM_fut: {len(gcm_fut_a)}')
print(f'  Estacional: clasificación completada ({HEMISFERIO})')

# ─────────────────────────────────────────────────────────────────────────────
# PASO 3 — Corrección ISIMIP → Futuro
# ─────────────────────────────────────────────────────────────────────────────

titulo_consola('PASO 3 — Corrección ISIMIP al Período Futuro', nivel=2)

print('\n  [Mensual]')
print('    • Empírico...', end=' ', flush=True)
gcm_fut_m_emp = isimip_bias_correction_v2(obs_men, gcm_cal_m, gcm_fut_m,
                    'empirica', N_QUANTILES, UMBRAL_MM, CORREGIR_DIAS_SECOS)
print('OK')

print('    • Gamma...', end=' ', flush=True)
gcm_fut_m_gam = isimip_bias_correction_v2(obs_men, gcm_cal_m, gcm_fut_m,
                    'gamma', N_QUANTILES, UMBRAL_MM, CORREGIR_DIAS_SECOS)
print('OK')

print('\n  [Anual]')
print('    • Empírico...', end=' ', flush=True)
gcm_fut_a_emp = isimip_bias_correction_v2(obs_anu, gcm_cal_a, gcm_fut_a,
                    'empirica', N_QUANTILES, UMBRAL_MM, CORREGIR_DIAS_SECOS)
print('OK')

print('    • Gamma...', end=' ', flush=True)
gcm_fut_a_gam = isimip_bias_correction_v2(obs_anu, gcm_cal_a, gcm_fut_a,
                    'gamma', N_QUANTILES, UMBRAL_MM, CORREGIR_DIAS_SECOS)
print('OK')

print('\n  [Estacional — Futuro]')
gcm_fut_m_emp_est = {}
for est in ['Verano', 'Otoño', 'Invierno', 'Primavera']:
    mask_o  = obs_men_est_label   == est
    mask_c  = gcm_cal_m_est_label == est
    mask_f  = gcm_fut_m_est_label == est
    if mask_f.sum() > 0:
        print(f'    • {est}...', end=' ', flush=True)
        gcm_fut_m_emp_est[est] = isimip_bias_correction_v2(
            obs_men[mask_o], gcm_cal_m[mask_c], gcm_fut_m[mask_f],
            'empirica', N_QUANTILES, UMBRAL_MM, CORREGIR_DIAS_SECOS)
        print('OK')

# ─────────────────────────────────────────────────────────────────────────────
# PASO 4 — Retroproyección (validación en calibración)
# ─────────────────────────────────────────────────────────────────────────────

titulo_consola('PASO 4 — Retroproyección en Calibración (Validación)', nivel=2)

print('\n  [Mensual]')
print('    • Empírico...', end=' ', flush=True)
gcm_cal_m_emp = isimip_bias_correction_v2(obs_men, gcm_cal_m, gcm_cal_m,
                    'empirica', N_QUANTILES, UMBRAL_MM, CORREGIR_DIAS_SECOS)
print('OK')

print('    • Gamma...', end=' ', flush=True)
gcm_cal_m_gam = isimip_bias_correction_v2(obs_men, gcm_cal_m, gcm_cal_m,
                    'gamma', N_QUANTILES, UMBRAL_MM, CORREGIR_DIAS_SECOS)
print('OK')

print('\n  [Anual]')
print('    • Empírico...', end=' ', flush=True)
gcm_cal_a_emp = isimip_bias_correction_v2(obs_anu, gcm_cal_a, gcm_cal_a,
                    'empirica', N_QUANTILES, UMBRAL_MM, CORREGIR_DIAS_SECOS)
print('OK')

print('    • Gamma...', end=' ', flush=True)
gcm_cal_a_gam = isimip_bias_correction_v2(obs_anu, gcm_cal_a, gcm_cal_a,
                    'gamma', N_QUANTILES, UMBRAL_MM, CORREGIR_DIAS_SECOS)
print('OK')

# ─────────────────────────────────────────────────────────────────────────────
# PASO 5 — Métricas de Desempeño
# ─────────────────────────────────────────────────────────────────────────────

titulo_consola('PASO 5 — Métricas de Desempeño', nivel=2)

print("""
╔══════════════════════════════════════════════════════════════════════════════╗
║  NOTA: QM corrige distribución, NO sincronía temporal                      ║
║  📊 MÉTRICAS CLAVE: PBIAS, Error Media (%), Error StdDev (%)               ║
╚══════════════════════════════════════════════════════════════════════════════╝
""")

titulo_consola('Escala MENSUAL', nivel=2)
m_sin_m = calcular_metricas(obs_men, gcm_cal_m,     'Sin corrección')
m_emp_m = calcular_metricas(obs_men, gcm_cal_m_emp, 'ISIMIP Empírico')
m_gam_m = calcular_metricas(obs_men, gcm_cal_m_gam, 'ISIMIP Gamma')
for m in [m_sin_m, m_emp_m, m_gam_m]:
    imprimir_metricas(m, 'Mensual')

titulo_consola('Escala ANUAL', nivel=2)
m_sin_a = calcular_metricas(obs_anu, gcm_cal_a,     'Sin corrección')
m_emp_a = calcular_metricas(obs_anu, gcm_cal_a_emp, 'ISIMIP Empírico')
m_gam_a = calcular_metricas(obs_anu, gcm_cal_a_gam, 'ISIMIP Gamma')
for m in [m_sin_a, m_emp_a, m_gam_a]:
    imprimir_metricas(m, 'Anual')

titulo_consola('Análisis ESTACIONAL', nivel=2)
metricas_est = []
for est in ['Verano', 'Otoño', 'Invierno', 'Primavera']:
    mask_o = obs_men_est_label   == est
    mask_c = gcm_cal_m_est_label == est
    if mask_o.sum() > 0 and mask_c.sum() > 0:
        gcm_est_val = isimip_bias_correction_v2(
            obs_men[mask_o], gcm_cal_m[mask_c], gcm_cal_m[mask_c],
            'empirica', N_QUANTILES, UMBRAL_MM, CORREGIR_DIAS_SECOS)
        me = calcular_metricas(obs_men[mask_o], gcm_est_val, est)
        if me:
            metricas_est.append({'Estación': est, **me})

df_est = pd.DataFrame(metricas_est) if metricas_est else None
if df_est is not None:
    print(df_est[['Estación','PBIAS (%)','Error Media (%)','Error StdDev (%)']].to_string(index=False))

# ─────────────────────────────────────────────────────────────────────────────
# PASO 6 — Extremos
# ─────────────────────────────────────────────────────────────────────────────

titulo_consola('PASO 6 — Análisis de Extremos (Percentiles)', nivel=2)

m_ext_sin_m = calcular_metricas_extremos(obs_men, gcm_cal_m,     'Sin corrección')
m_ext_emp_m = calcular_metricas_extremos(obs_men, gcm_cal_m_emp, 'ISIMIP Empírico')
m_ext_gam_m = calcular_metricas_extremos(obs_men, gcm_cal_m_gam, 'ISIMIP Gamma')

m_ext_sin_a = calcular_metricas_extremos(obs_anu, gcm_cal_a,     'Sin corrección')
m_ext_emp_a = calcular_metricas_extremos(obs_anu, gcm_cal_a_emp, 'ISIMIP Empírico')
m_ext_gam_a = calcular_metricas_extremos(obs_anu, gcm_cal_a_gam, 'ISIMIP Gamma')

df_ext_m = pd.DataFrame([m_ext_sin_m, m_ext_emp_m, m_ext_gam_m])
df_ext_a = pd.DataFrame([m_ext_sin_a, m_ext_emp_a, m_ext_gam_a])

cols_show = ['Nombre','P5 OBS','P5 SIM','Error P5 (%)','P95 OBS','P95 SIM','Error P95 (%)']
print('\n  MENSUAL:')
print(df_ext_m[cols_show].to_string(index=False))
print('\n  ANUAL:')
print(df_ext_a[cols_show].to_string(index=False))

# ─────────────────────────────────────────────────────────────────────────────
# PASO 7 — Diagnóstico Automático
# ─────────────────────────────────────────────────────────────────────────────

titulo_consola('PASO 7 — Diagnóstico Final Automático')

icon_m, nivel_m, desc_m = clasificar_metrica_qm(m_emp_m)
icon_a, nivel_a, desc_a = clasificar_metrica_qm(m_emp_a)

print(f'\n  MENSUAL → {icon_m} {nivel_m}')
print(f'    PBIAS={m_emp_m["PBIAS (%)"]:.2f}%  |  ErrMedia={m_emp_m["Error Media (%)"]:.2f}%  |  ErrStd={m_emp_m["Error StdDev (%)"]:.2f}%')
print(f'    {desc_m}')

print(f'\n  ANUAL   → {icon_a} {nivel_a}')
print(f'    PBIAS={m_emp_a["PBIAS (%)"]:.2f}%  |  ErrMedia={m_emp_a["Error Media (%)"]:.2f}%  |  ErrStd={m_emp_a["Error StdDev (%)"]:.2f}%')
print(f'    {desc_a}')

# ─────────────────────────────────────────────────────────────────────────────
# PASO 8 — Gráficos
# ─────────────────────────────────────────────────────────────────────────────

titulo_consola('PASO 8 — Generación de Gráficos PNG', nivel=2)

pngs = []

# ── MENSUAL ──────────────────────────────────────────────────────────────────
print('\n  [Mensual]')

series_m = {
    'OBS'             : obs_men,
    'GCM sin corr.'   : gcm_cal_m,
    'ISIMIP Empírico' : gcm_cal_m_emp,
    'ISIMIP Gamma'    : gcm_cal_m_gam,
}

pngs.append({'Gráfico': 'FDC Mensual', 'Ruta': plot_fdc(
    series_m,
    os.path.join(PNG_DIR, f'{MODELO}_01_FDC_mensual.png'),
    f'FDC Mensual — Calibración — {MODELO}')})

pngs.append({'Gráfico': 'Histograma Mensual', 'Ruta': plot_hist(
    series_m,
    os.path.join(PNG_DIR, f'{MODELO}_02_Hist_mensual.png'),
    f'Histograma Mensual — Calibración — {MODELO}')})

sims_m = {k: v for k, v in series_m.items() if k != 'OBS'}
pngs.append({'Gráfico': 'QQ-plot Mensual', 'Ruta': plot_qq(
    obs_men, sims_m,
    os.path.join(PNG_DIR, f'{MODELO}_03_QQ_mensual.png'),
    f'QQ-plot Mensual — Calibración — {MODELO}')})

pngs.append({'Gráfico': 'Boxplot Mensual', 'Ruta': plot_boxplot(
    obs_men, sims_m,
    os.path.join(PNG_DIR, f'{MODELO}_04_Boxplot_mensual.png'),
    f'Boxplot Mensual — Calibración — {MODELO}')})

pngs.append({'Gráfico': 'Ciclo Anual Mensual', 'Ruta': plot_ciclo_anual(
    series_m,
    os.path.join(PNG_DIR, f'{MODELO}_05_CicloAnual_mensual.png'),
    f'Ciclo Anual Mensual — Calibración — {MODELO}')})

# ── ANUAL ─────────────────────────────────────────────────────────────────────
print('\n  [Anual]')

series_a = {
    'OBS'             : obs_anu,
    'GCM sin corr.'   : gcm_cal_a,
    'ISIMIP Empírico' : gcm_cal_a_emp,
    'ISIMIP Gamma'    : gcm_cal_a_gam,
}

pngs.append({'Gráfico': 'FDC Anual', 'Ruta': plot_fdc(
    series_a,
    os.path.join(PNG_DIR, f'{MODELO}_06_FDC_anual.png'),
    f'FDC Anual — Calibración — {MODELO}')})

pngs.append({'Gráfico': 'Histograma Anual', 'Ruta': plot_hist(
    series_a,
    os.path.join(PNG_DIR, f'{MODELO}_07_Hist_anual.png'),
    f'Histograma Anual — Calibración — {MODELO}')})

sims_a = {k: v for k, v in series_a.items() if k != 'OBS'}
pngs.append({'Gráfico': 'QQ-plot Anual', 'Ruta': plot_qq(
    obs_anu, sims_a,
    os.path.join(PNG_DIR, f'{MODELO}_08_QQ_anual.png'),
    f'QQ-plot Anual — Calibración — {MODELO}')})

pngs.append({'Gráfico': 'Boxplot Anual', 'Ruta': plot_boxplot(
    obs_anu, sims_a,
    os.path.join(PNG_DIR, f'{MODELO}_09_Boxplot_anual.png'),
    f'Boxplot Anual — Calibración — {MODELO}')})

# ── FUTURO ────────────────────────────────────────────────────────────────────
print('\n  [Futuro — Mensual]')

series_fut_m = {
    'GCM sin corr.'   : gcm_fut_m,
    'ISIMIP Empírico' : gcm_fut_m_emp,
    'ISIMIP Gamma'    : gcm_fut_m_gam,
}

pngs.append({'Gráfico': 'FDC Futuro Mensual', 'Ruta': plot_fdc(
    series_fut_m,
    os.path.join(PNG_DIR, f'{MODELO}_10_FDC_futuro_mensual.png'),
    f'FDC Mensual — Período Futuro — {MODELO}')})

pngs.append({'Gráfico': 'Ciclo Anual Futuro', 'Ruta': plot_ciclo_anual(
    series_fut_m,
    os.path.join(PNG_DIR, f'{MODELO}_11_CicloAnual_futuro.png'),
    f'Ciclo Anual — Período Futuro — {MODELO}')})

# ─────────────────────────────────────────────────────────────────────────────
# PASO 9 — Exportación Excel
# ─────────────────────────────────────────────────────────────────────────────

titulo_consola('PASO 9 — Exportación a Excel', nivel=2)

wb = Workbook()
wb.remove(wb.active)

# ── Hoja: calibracion_mensual
ws = wb.create_sheet('calibracion_mensual')
df_to_sheet(ws, pd.DataFrame({
    'obs_mm'         : obs_men,
    'gcm_sin_corr_mm': gcm_cal_m,
    'gcm_emp_mm'     : gcm_cal_m_emp,
    'gcm_gam_mm'     : gcm_cal_m_gam,
}))
print('  ✓ calibracion_mensual')

# ── Hoja: futuro_mensual
ws = wb.create_sheet('futuro_mensual')
df_to_sheet(ws, pd.DataFrame({
    'gcm_sin_corr_mm': gcm_fut_m,
    'gcm_emp_mm'     : gcm_fut_m_emp,
    'gcm_gam_mm'     : gcm_fut_m_gam,
}))
print('  ✓ futuro_mensual')

# ── Hoja: calibracion_anual
ws = wb.create_sheet('calibracion_anual')
df_to_sheet(ws, pd.DataFrame({
    'obs_mm'         : obs_anu,
    'gcm_sin_corr_mm': gcm_cal_a,
    'gcm_emp_mm'     : gcm_cal_a_emp,
    'gcm_gam_mm'     : gcm_cal_a_gam,
}))
print('  ✓ calibracion_anual')

# ── Hoja: futuro_anual
ws = wb.create_sheet('futuro_anual')
df_to_sheet(ws, pd.DataFrame({
    'gcm_sin_corr_mm': gcm_fut_a,
    'gcm_emp_mm'     : gcm_fut_a_emp,
    'gcm_gam_mm'     : gcm_fut_a_gam,
}))
print('  ✓ futuro_anual')

# ── Hoja: extremos_mensual
ws = wb.create_sheet('extremos_mensual')
df_to_sheet(ws, df_ext_m, 'd94f3d')
print('  ✓ extremos_mensual')

# ── Hoja: extremos_anual
ws = wb.create_sheet('extremos_anual')
df_to_sheet(ws, df_ext_a, 'd94f3d')
print('  ✓ extremos_anual')

# ── Hoja: estacional (si hay datos)
if df_est is not None:
    ws = wb.create_sheet('estacional')
    df_to_sheet(ws, df_est, 'ff7f0e')
    print('  ✓ estacional')

# ── Hoja: metricas
ws = wb.create_sheet('metricas')
filas = []
for escala, lista in [('Mensual', [m_sin_m, m_emp_m, m_gam_m]),
                       ('Anual',   [m_sin_a, m_emp_a, m_gam_a])]:
    for m in lista:
        if m:
            filas.append({'Escala': escala, **m})
df_to_sheet(ws, pd.DataFrame(filas))
print('  ✓ metricas')

# ── Hoja: graficos
ws = wb.create_sheet('graficos')
df_to_sheet(ws, pd.DataFrame(pngs), '2ca02c')
print('  ✓ graficos')

# ── Hoja: RESUMEN (al inicio)
ws = wb.create_sheet('RESUMEN', 0)
fecha_hora = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
resumen = [
    ['ISIMIP Bias Correction v2.2 — Resumen Ejecutivo'],
    [''],
    ['Generado'          , fecha_hora],
    ['Modelo GCM'        , MODELO],
    ['Período calibración', f'{CAL_INI} → {CAL_FIN}'],
    ['Período futuro'    , f'{FUT_INI} → {FUT_FIN}'],
    ['Hemisferio'        , HEMISFERIO],
    ['Corrección días secos', str(CORREGIR_DIAS_SECOS)],
    ['N Quantiles'       , N_QUANTILES],
    ['Umbral lluvia (mm)', UMBRAL_MM],
    [''],
    ['─── MÉTRICAS MENSUAL (ISIMIP Empírico) ───'],
    ['PBIAS (%)'         , m_emp_m['PBIAS (%)']      if m_emp_m else 'N/A'],
    ['Error Media (%)'   , m_emp_m['Error Media (%)'] if m_emp_m else 'N/A'],
    ['Error StdDev (%)'  , m_emp_m['Error StdDev (%)']if m_emp_m else 'N/A'],
    ['Diagnóstico'       , f'{icon_m} {nivel_m}'],
    [''],
    ['─── MÉTRICAS ANUAL (ISIMIP Empírico) ───'],
    ['PBIAS (%)'         , m_emp_a['PBIAS (%)']      if m_emp_a else 'N/A'],
    ['Error Media (%)'   , m_emp_a['Error Media (%)'] if m_emp_a else 'N/A'],
    ['Error StdDev (%)'  , m_emp_a['Error StdDev (%)']if m_emp_a else 'N/A'],
    ['Diagnóstico'       , f'{icon_a} {nivel_a}'],
    [''],
    ['─── GRÁFICOS GENERADOS ───'],
]
for png in pngs:
    resumen.append([png['Gráfico'], png['Ruta']])

fill_title = PatternFill('solid', fgColor='1a6e9e')
font_title = Font(bold=True, color='FFFFFF', size=12)
border     = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))

for r_idx, row in enumerate(resumen, 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.border    = border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        if r_idx == 1:
            cell.fill = fill_title
            cell.font = font_title

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 55
print('  ✓ RESUMEN')

# ── Guardar Excel
wb.save(SALIDA_XLSX)
print(f'\n  💾 Excel guardado → {SALIDA_XLSX}')

# ─────────────────────────────────────────────────────────────────────────────
# FIN
# ─────────────────────────────────────────────────────────────────────────────

titulo_consola(f'✅ PROCESO COMPLETADO — {MODELO}')
print(f"""
  📁 Directorio de salida : {OUT_DIR}
  📊 Excel completo       : {SALIDA_XLSX}
  🖼️  Gráficos PNG         : {PNG_DIR}
  📈 Gráficos generados   : {len(pngs)}

  ╔═══════════════════════════════════════════════════════╗
  ║  DIAGNÓSTICO FINAL                                    ║
  ║  Mensual → {icon_m} {nivel_m:<42}║
  ║  Anual   → {icon_a} {nivel_a:<42}║
  ╚═══════════════════════════════════════════════════════╝
""")